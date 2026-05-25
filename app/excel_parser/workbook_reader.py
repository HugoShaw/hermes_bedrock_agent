"""Workbook reader using openpyxl for cell content."""
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .models import CellBlock, SheetData, WorkbookData

logger = logging.getLogger(__name__)


def read_workbook(excel_path: str) -> WorkbookData:
    """Read workbook metadata and cell content using openpyxl."""
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    workbook_data = WorkbookData(source_path=str(excel_path))
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = SheetData(
            name=sheet_name,
            max_row=ws.max_row or 0,
            max_col=ws.max_column or 0,
        )
        
        # Merged cells
        for mc in ws.merged_cells.ranges:
            sheet_data.merged_cells.append(str(mc))
        
        # Check if sheet has drawings
        if ws._charts or ws._images:
            sheet_data.has_drawing = True
        
        # Read cell content - find non-empty regions
        cell_blocks = _extract_cell_blocks(ws, sheet_name)
        sheet_data.cell_blocks = cell_blocks
        
        workbook_data.sheets.append(sheet_data)
    
    wb.close()
    return workbook_data


def _extract_cell_blocks(ws, sheet_name: str) -> list[CellBlock]:
    """Extract non-empty cell regions as blocks."""
    blocks = []
    
    if ws.max_row is None or ws.max_column is None:
        return blocks
    
    # Find all non-empty cells
    non_empty_rows = set()
    non_empty_cols = set()
    cell_values = {}
    
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200),
                            min_col=1, max_col=min(ws.max_column, 50)):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                non_empty_rows.add(cell.row)
                non_empty_cols.add(cell.column)
                cell_values[(cell.row, cell.column)] = str(cell.value).strip()
    
    if not non_empty_rows:
        return blocks
    
    # Find contiguous row groups
    sorted_rows = sorted(non_empty_rows)
    sorted_cols = sorted(non_empty_cols)
    
    if not sorted_rows or not sorted_cols:
        return blocks
    
    # Split into blocks by gaps > 3 rows
    row_groups = []
    current_group = [sorted_rows[0]]
    for i in range(1, len(sorted_rows)):
        if sorted_rows[i] - sorted_rows[i-1] > 3:
            row_groups.append(current_group)
            current_group = [sorted_rows[i]]
        else:
            current_group.append(sorted_rows[i])
    row_groups.append(current_group)
    
    for group in row_groups:
        min_row = min(group)
        max_row = max(group)
        min_col = min(sorted_cols)
        max_col = max(sorted_cols)
        
        # Build data matrix
        data = []
        for r in range(min_row, max_row + 1):
            row_data = []
            for c in range(min_col, max_col + 1):
                row_data.append(cell_values.get((r, c), ""))
            data.append(row_data)
        
        # Determine if it looks like a table (has headers and structured data)
        is_table = _looks_like_table(data)
        
        block = CellBlock(
            sheet_name=sheet_name,
            start_row=min_row,
            start_col=min_col,
            end_row=max_row,
            end_col=max_col,
            data=data,
            is_table=is_table,
        )
        blocks.append(block)
    
    return blocks


def _looks_like_table(data: list[list[str]]) -> bool:
    """Heuristic: if multiple rows have similar column patterns, it's a table."""
    if len(data) < 2:
        return False
    
    # Check if first row has content in multiple columns (likely header)
    first_row_filled = sum(1 for c in data[0] if c)
    if first_row_filled >= 2:
        # Check if subsequent rows also have content in similar positions
        content_rows = sum(1 for row in data[1:] if sum(1 for c in row if c) >= 2)
        if content_rows >= 1:
            return True
    
    return False
