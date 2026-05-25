"""
Excel parser — openpyxlでワークブックを開き、シート・セルのメタデータを抽出する。

出力:
  - excel_workbooks.jsonl  … ワークブック単位の概要
  - excel_sheets.jsonl     … シート単位の詳細
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed — Excel parsing disabled")


class ExcelParser:
    """ワークブックとシートのメタデータを抽出するパーサー。

    Parameters
    ----------
    dataset:
        データセット名。
    run_id:
        実行ID。
    preserve_formulas:
        True の場合、数式を文字列として保持する (data_only=False)。
    include_hidden_sheets:
        隠しシートを含める場合 True。
    max_cell_scan:
        1シートあたりの最大スキャンセル数 (パフォーマンス上限)。
    """

    def __init__(
        self,
        dataset: str = "sample_20260519",
        run_id: str = "sample_20260519_evidence_v1",
        preserve_formulas: bool = True,
        include_hidden_sheets: bool = False,
        max_cell_scan: int = 5000,
    ) -> None:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required. Run: pip install openpyxl")
        self.dataset = dataset
        self.run_id = run_id
        self.preserve_formulas = preserve_formulas
        self.include_hidden_sheets = include_hidden_sheets
        self.max_cell_scan = max_cell_scan

    def parse_workbook(
        self,
        file_path: str,
        source_s3_uri: str = "",
    ) -> dict[str, Any]:
        """ワークブックを解析してワークブックレコードとシートレコードを返す。

        Parameters
        ----------
        file_path:
            ローカルファイルパス。
        source_s3_uri:
            元ファイルのS3 URI (省略可能)。

        Returns
        -------
        dict with keys:
            workbook_record, sheet_records, openpyxl_workbook, error
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext == ".xls":
            return {
                "workbook_record": None,
                "sheet_records": [],
                "openpyxl_workbook": None,
                "error": f"Legacy .xls format is not supported. File: {file_path}",
            }
        if ext not in (".xlsx", ".xlsm"):
            return {
                "workbook_record": None,
                "sheet_records": [],
                "openpyxl_workbook": None,
                "error": f"Unsupported extension: {ext}",
            }

        try:
            wb = openpyxl.load_workbook(
                file_path,
                read_only=False,
                data_only=not self.preserve_formulas,
                keep_links=False,
            )
        except Exception as exc:
            logger.error("Failed to open workbook %s: %s", file_path, exc)
            return {
                "workbook_record": None,
                "sheet_records": [],
                "openpyxl_workbook": None,
                "error": str(exc),
            }

        visible_sheets = [n for n in wb.sheetnames if wb[n].sheet_state == "visible"]
        hidden_sheets = [n for n in wb.sheetnames if wb[n].sheet_state != "visible"]

        # 定義名 (named ranges)
        defined_names: list[str] = []
        try:
            defined_names = list(wb.defined_names.definedName.keys()) if wb.defined_names else []
        except Exception:
            pass

        wb_record: dict[str, Any] = {
            "workbook_id": _wb_id(file_path, self.dataset),
            "dataset": self.dataset,
            "run_id": self.run_id,
            "source_file": str(path),
            "source_s3_uri": source_s3_uri,
            "file_name": path.name,
            "file_extension": ext,
            "workbook_name": path.stem,
            "sheet_count": len(wb.sheetnames),
            "visible_sheet_count": len(visible_sheets),
            "hidden_sheet_count": len(hidden_sheets),
            "sheet_names": list(wb.sheetnames),
            "visible_sheets": visible_sheets,
            "hidden_sheets": hidden_sheets,
            "defined_names": defined_names,
            "file_size_bytes": path.stat().st_size if path.exists() else 0,
        }

        sheet_records: list[dict[str, Any]] = []
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws: Worksheet = wb[sheet_name]
            is_visible = ws.sheet_state == "visible"
            if not is_visible and not self.include_hidden_sheets:
                continue
            sheet_rec = self._parse_sheet(ws, idx, wb_record["workbook_id"], path, source_s3_uri)
            sheet_records.append(sheet_rec)

        logger.info(
            "Parsed workbook %s: %d sheets (%d visible)",
            path.name, len(sheet_records), len(visible_sheets),
        )
        return {
            "workbook_record": wb_record,
            "sheet_records": sheet_records,
            "openpyxl_workbook": wb,
            "error": None,
        }

    def _parse_sheet(
        self,
        ws: "Worksheet",
        idx: int,
        workbook_id: str,
        workbook_path: Path,
        source_s3_uri: str,
    ) -> dict[str, Any]:
        """シート1枚のメタデータを抽出する。"""
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        non_empty = 0
        has_formula = False
        has_comments = False
        formulas: list[dict[str, str]] = []
        comments: list[dict[str, str]] = []

        # セルサンプリング (パフォーマンス制限あり)
        scan_rows = min(max_row, 200)
        scan_cols = min(max_col, 50)
        scanned = 0
        for row in ws.iter_rows(min_row=1, max_row=scan_rows, max_col=scan_cols):
            for cell in row:
                if scanned >= self.max_cell_scan:
                    break
                scanned += 1
                if cell.value is not None:
                    non_empty += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    has_formula = True
                    if len(formulas) < 50:
                        formulas.append({"cell": cell.coordinate, "formula": cell.value})
                if cell.comment:
                    has_comments = True
                    if len(comments) < 50:
                        txt = cell.comment.text or ""
                        comments.append({"cell": cell.coordinate, "comment": txt})
            if scanned >= self.max_cell_scan:
                break

        merged_ranges = [str(r) for r in ws.merged_cells.ranges] if ws.merged_cells else []

        sheet_id = _sheet_id(workbook_id, ws.title, idx)

        return {
            "sheet_id": sheet_id,
            "workbook_id": workbook_id,
            "dataset": self.dataset,
            "run_id": self.run_id,
            "source_file": str(workbook_path),
            "source_s3_uri": source_s3_uri,
            "workbook_name": workbook_path.stem,
            "sheet_name": ws.title,
            "sheet_index": idx,
            "visible": ws.sheet_state == "visible",
            "max_row": max_row,
            "max_column": max_col,
            "non_empty_cell_count": non_empty,
            "merged_cell_ranges": merged_ranges,
            "merged_cell_count": len(merged_ranges),
            "has_formula": has_formula,
            "has_comments": has_comments,
            "formula_samples": formulas,
            "comment_samples": comments,
            "dimensions": ws.dimensions or "",
            "scanned_cells": scanned,
        }

    def write_jsonl(
        self,
        workbook_records: list[dict[str, Any]],
        sheet_records: list[dict[str, Any]],
        output_dir: str,
    ) -> dict[str, str]:
        """ワークブック・シートレコードをJSONLファイルに書き出す。

        Returns
        -------
        dict with keys: workbooks_path, sheets_path
        """
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)

        wb_path = str(out / "excel_workbooks.jsonl")
        sh_path = str(out / "excel_sheets.jsonl")

        with open(wb_path, "w", encoding="utf-8") as f:
            for rec in workbook_records:
                f.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")

        with open(sh_path, "w", encoding="utf-8") as f:
            for rec in sheet_records:
                f.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")

        logger.info("Wrote %d workbook records → %s", len(workbook_records), wb_path)
        logger.info("Wrote %d sheet records → %s", len(sheet_records), sh_path)
        return {"workbooks_path": wb_path, "sheets_path": sh_path}


# ---- ID helpers -------------------------------------------------------

def _wb_id(source_path: str, dataset: str) -> str:
    import hashlib
    raw = f"wb:{dataset}:{source_path}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _sheet_id(workbook_id: str, sheet_name: str, sheet_index: int) -> str:
    import hashlib
    raw = f"sh:{workbook_id}:{sheet_name}:{sheet_index}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]
