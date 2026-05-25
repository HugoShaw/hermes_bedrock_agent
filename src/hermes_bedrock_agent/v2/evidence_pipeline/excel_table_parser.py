"""
Excel table parser — シート内のテーブル領域を検出し、行データを正規化する。

出力:
  - excel_table_regions.jsonl   … 検出テーブル領域の定義
  - excel_rows_normalized.jsonl … ヘッダー付き正規化行データ
"""
from __future__ import annotations

import hashlib
import json
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# テーブル検出の閾値
EMPTY_ROW_GAP = 3        # この行数以上の空行でテーブルを分割
MIN_REGION_ROWS = 2      # テーブルとして認める最小行数
MIN_REGION_COLS = 2      # テーブルとして認める最小列数
HEADER_TEXT_RATIO = 0.6  # この割合以上テキストならヘッダー行と判断

# 日本語/英語のテーブル種別キーワード
_TYPE_KEYWORDS: dict[str, list[str]] = {
    "api_table": ["API", "エンドポイント", "endpoint", "method", "path", "status"],
    "field_mapping_table": ["フィールド", "field", "項目", "マッピング", "mapping", "論理名", "物理名"],
    "error_code_table": ["エラーコード", "error", "code", "メッセージ", "message", "原因"],
    "test_case_table": ["テスト", "test", "ケース", "case", "期待", "結果", "result"],
    "code_master_table": ["コード", "code", "名称", "マスタ", "master", "区分"],
}


class ExcelTableParser:
    """シート内のテーブル領域を検出して行データを正規化する。

    Parameters
    ----------
    dataset, run_id:
        パイプライン識別子。
    max_rows_per_sheet:
        1シートあたりの最大処理行数。
    """

    def __init__(
        self,
        dataset: str = "sample_20260519",
        run_id: str = "sample_20260519_evidence_v1",
        max_rows_per_sheet: int = 5000,
    ) -> None:
        self.dataset = dataset
        self.run_id = run_id
        self.max_rows_per_sheet = max_rows_per_sheet

    def detect_tables(
        self,
        ws: "Worksheet",
        sheet_id: str,
        workbook_id: str,
        workbook_name: str,
        source_file: str,
        source_s3_uri: str = "",
    ) -> dict[str, Any]:
        """1シートのテーブル領域を検出する。

        Returns
        -------
        dict with keys: table_regions, normalized_rows
        """
        max_row = min(ws.max_row or 0, self.max_rows_per_sheet)
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return {"table_regions": [], "normalized_rows": []}

        # シート全体を2D配列として読み込む
        grid = _read_grid(ws, max_row, max_col)

        # 空行のギャップでテーブル候補を分割
        region_row_spans = _find_row_spans(grid, max_row)

        table_regions: list[dict[str, Any]] = []
        normalized_rows: list[dict[str, Any]] = []

        for (r_start, r_end) in region_row_spans:
            # 使用列範囲を検出
            c_start, c_end = _find_col_span(grid, r_start, r_end, max_col)
            if c_end - c_start + 1 < MIN_REGION_COLS:
                continue
            if r_end - r_start + 1 < MIN_REGION_ROWS:
                continue

            # ヘッダー行を検出 (最大3行)
            header_rows = _detect_header_rows(grid, r_start, r_end, c_start, c_end)
            if not header_rows:
                header_rows = [r_start]

            data_start = max(header_rows) + 1
            if data_start > r_end:
                continue

            col_names = _extract_column_names(grid, header_rows, c_start, c_end)
            cell_range = _range_str(r_start, c_start, r_end, c_end)
            region_id = _region_id(sheet_id, cell_range)
            region_type = _infer_region_type(col_names)
            confidence = _compute_confidence(col_names, data_start, r_end, r_start, r_end, c_start, c_end)

            region_rec: dict[str, Any] = {
                "table_region_id": region_id,
                "workbook_id": workbook_id,
                "sheet_id": sheet_id,
                "dataset": self.dataset,
                "run_id": self.run_id,
                "source_file": source_file,
                "source_s3_uri": source_s3_uri,
                "workbook_name": workbook_name,
                "sheet_name": ws.title,
                "sheet_index": (ws._id - 1) if hasattr(ws, "_id") and ws._id is not None else 0,
                "cell_range": cell_range,
                "header_rows": header_rows,
                "data_start_row": data_start,
                "data_end_row": r_end,
                "columns": col_names,
                "column_count": len(col_names),
                "data_row_count": max(0, r_end - data_start + 1),
                "confidence": confidence,
                "region_type": region_type,
            }
            table_regions.append(region_rec)

            # 各データ行を正規化
            for row_num in range(data_start, r_end + 1):
                row_vals: dict[str, Any] = {}
                row_cells: dict[str, str] = {}
                for ci, col_name in enumerate(col_names):
                    col_idx = c_start + ci
                    if col_idx > c_end:
                        break
                    raw_val = grid[row_num - 1][col_idx - 1] if row_num <= len(grid) and col_idx <= len(grid[0]) else None
                    row_vals[col_name] = raw_val
                    row_cells[col_name] = _cell_ref(row_num, col_idx)

                row_id = _row_id(sheet_id, region_id, row_num)
                row_rec: dict[str, Any] = {
                    "row_id": row_id,
                    "table_region_id": region_id,
                    "workbook_id": workbook_id,
                    "sheet_id": sheet_id,
                    "dataset": self.dataset,
                    "run_id": self.run_id,
                    "source_file": source_file,
                    "source_s3_uri": source_s3_uri,
                    "workbook_name": workbook_name,
                    "sheet_name": ws.title,
                    "row_number": row_num,
                    "column_names": col_names,
                    "values": {k: str(v) if v is not None else "" for k, v in row_vals.items()},
                    "raw_values": row_vals,
                    "cell_refs": row_cells,
                    "cell_range": _range_str(row_num, c_start, row_num, c_end),
                }
                normalized_rows.append(row_rec)

        return {"table_regions": table_regions, "normalized_rows": normalized_rows}

    def write_jsonl(
        self,
        table_regions: list[dict[str, Any]],
        normalized_rows: list[dict[str, Any]],
        output_dir: str,
    ) -> dict[str, str]:
        """テーブル領域・行データをJSONLファイルに書き出す。"""
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        tr_path = str(out / "excel_table_regions.jsonl")
        row_path = str(out / "excel_rows_normalized.jsonl")

        with open(tr_path, "w", encoding="utf-8") as f:
            for rec in table_regions:
                f.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")
        with open(row_path, "w", encoding="utf-8") as f:
            for rec in normalized_rows:
                f.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")

        logger.info("Wrote %d table regions → %s", len(table_regions), tr_path)
        logger.info("Wrote %d normalized rows → %s", len(normalized_rows), row_path)
        return {"table_regions_path": tr_path, "rows_path": row_path}


# ---- Internal helpers -------------------------------------------------

def _read_grid(ws: "Worksheet", max_row: int, max_col: int) -> list[list[Any]]:
    """シートを2D配列として読み込む (merged cells はマスター値を使用)。"""
    # マージセルのマスター座標マップ
    merged_master: dict[tuple[int, int], Any] = {}
    for merge_range in (ws.merged_cells.ranges if ws.merged_cells else []):
        master_cell = ws.cell(merge_range.min_row, merge_range.min_col)
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged_master[(row, col)] = master_cell.value

    grid: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        row_vals: list[Any] = []
        for cell in row:
            val = merged_master.get((cell.row, cell.column), cell.value)
            row_vals.append(val)
        grid.append(row_vals)
    return grid


def _is_empty_row(grid: list[list[Any]], row_idx: int) -> bool:
    if row_idx >= len(grid):
        return True
    return all(v is None or str(v).strip() == "" for v in grid[row_idx])


def _find_row_spans(grid: list[list[Any]], max_row: int) -> list[tuple[int, int]]:
    """連続空行でテーブル候補の行範囲を分割する (1始まり)。"""
    spans: list[tuple[int, int]] = []
    gap = 0
    start: int | None = None

    for i in range(max_row):
        if _is_empty_row(grid, i):
            gap += 1
            if gap >= EMPTY_ROW_GAP and start is not None:
                spans.append((start + 1, i + 1 - gap))
                start = None
                gap = 0
        else:
            if start is None:
                start = i
            gap = 0

    if start is not None:
        end = max_row - gap
        if end >= start + 1:
            spans.append((start + 1, end))

    return spans


def _find_col_span(
    grid: list[list[Any]],
    r_start: int,
    r_end: int,
    max_col: int,
) -> tuple[int, int]:
    """使用列範囲を検出する (1始まり)。"""
    min_c, max_c = max_col, 1
    for r in range(r_start - 1, min(r_end, len(grid))):
        for c, val in enumerate(grid[r]):
            if val is not None and str(val).strip() != "":
                min_c = min(min_c, c + 1)
                max_c = max(max_c, c + 1)
    if min_c > max_c:
        return 1, max_col
    return min_c, max_c


def _detect_header_rows(
    grid: list[list[Any]],
    r_start: int,
    r_end: int,
    c_start: int,
    c_end: int,
) -> list[int]:
    """テキスト比率が高い先頭行をヘッダー行と判断する。"""
    headers: list[int] = []
    width = c_end - c_start + 1
    for r in range(r_start - 1, min(r_start + 3, r_end)):  # 最大3行
        if r >= len(grid):
            break
        row = grid[r][c_start - 1: c_end]
        non_empty = [v for v in row if v is not None and str(v).strip() != ""]
        text_vals = [v for v in non_empty if not _is_numeric(v)]
        if width > 0 and len(non_empty) / width >= 0.3 and (
            len(non_empty) == 0 or len(text_vals) / max(len(non_empty), 1) >= HEADER_TEXT_RATIO
        ):
            headers.append(r + 1)
        else:
            break
    return headers


def _extract_column_names(
    grid: list[list[Any]],
    header_rows: list[int],
    c_start: int,
    c_end: int,
) -> list[str]:
    """複数ヘッダー行から列名を生成する (複数行は '/' で結合)。"""
    col_names: list[str] = []
    width = c_end - c_start + 1
    for ci in range(width):
        parts: list[str] = []
        for r in header_rows:
            if r - 1 < len(grid) and c_start - 1 + ci < len(grid[r - 1]):
                val = grid[r - 1][c_start - 1 + ci]
                if val is not None and str(val).strip():
                    parts.append(str(val).strip())
        name = "/".join(parts) if parts else f"col_{ci + 1}"
        col_names.append(name)
    return col_names


def _infer_region_type(col_names: list[str]) -> str:
    name_text = " ".join(col_names).lower()
    for region_type, keywords in _TYPE_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in name_text:
                return region_type
    return "unknown_table"


def _compute_confidence(
    col_names: list[str],
    data_start: int,
    data_end: int,
    r_start: int,
    r_end: int,
    c_start: int,
    c_end: int,
) -> float:
    score = 0.5
    named = sum(1 for c in col_names if not c.startswith("col_"))
    if named / max(len(col_names), 1) >= 0.8:
        score += 0.2
    if data_end - data_start + 1 >= 3:
        score += 0.1
    if c_end - c_start + 1 >= 3:
        score += 0.1
    if r_end - r_start + 1 >= 5:
        score += 0.1
    return min(round(score, 2), 1.0)


def _is_numeric(val: Any) -> bool:
    try:
        float(str(val))
        return True
    except (ValueError, TypeError):
        return False


def _range_str(r1: int, c1: int, r2: int, c2: int) -> str:
    return f"{_col_letter(c1)}{r1}:{_col_letter(c2)}{r2}"


def _col_letter(col: int) -> str:
    """1始まりの列番号をExcel列文字に変換 (例: 1→A, 27→AA)。"""
    result = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _cell_ref(row: int, col: int) -> str:
    return f"{_col_letter(col)}{row}"


def _region_id(sheet_id: str, cell_range: str) -> str:
    raw = f"tr:{sheet_id}:{cell_range}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _row_id(sheet_id: str, region_id: str, row_number: int) -> str:
    raw = f"row:{sheet_id}:{region_id}:{row_number}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]
