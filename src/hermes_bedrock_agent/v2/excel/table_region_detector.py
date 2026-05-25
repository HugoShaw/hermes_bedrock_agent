"""
Table region detector — identify rectangular table regions within Excel sheets.

Uses heuristics to:
- Identify non-empty cell clusters
- Group cells into rectangular regions
- Split regions by large empty row/column gaps
- Detect header row candidates
- Calculate detection confidence
"""
from __future__ import annotations

import logging
from typing import Any

from openpyxl.utils import get_column_letter

from hermes_bedrock_agent.v2.excel.excel_schema import ExcelTableRegion, ExcelSheetRecord

logger = logging.getLogger(__name__)

# Gap thresholds
EMPTY_ROW_GAP = 3  # N consecutive empty rows to split regions
MIN_REGION_ROWS = 2  # Minimum rows to consider as a table
MIN_REGION_COLS = 2  # Minimum columns to consider as a table


class TableRegionDetector:
    """Detect table regions in an Excel worksheet.

    Parameters
    ----------
    empty_row_gap : int
        Number of consecutive empty rows that triggers a region split.
    min_region_rows : int
        Minimum rows for a valid table region.
    min_region_cols : int
        Minimum columns for a valid table region.
    """

    def __init__(
        self,
        empty_row_gap: int = EMPTY_ROW_GAP,
        min_region_rows: int = MIN_REGION_ROWS,
        min_region_cols: int = MIN_REGION_COLS,
    ) -> None:
        self.empty_row_gap = empty_row_gap
        self.min_region_rows = min_region_rows
        self.min_region_cols = min_region_cols

    def detect(
        self,
        ws: Any,
        sheet_record: ExcelSheetRecord,
        workbook_id: str,
    ) -> list[ExcelTableRegion]:
        """Detect table regions in a worksheet.

        Parameters
        ----------
        ws : openpyxl Worksheet
            The worksheet.
        sheet_record : ExcelSheetRecord
            Sheet metadata record.
        workbook_id : str
            Parent workbook ID.

        Returns
        -------
        List of ExcelTableRegion records.
        """
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            return []

        # Limit scan size
        scan_max_row = min(max_row, 1000)
        scan_max_col = min(max_col, 100)

        # Build row non-empty bitmap
        row_has_data: dict[int, list[int]] = {}  # row -> list of non-empty column indices
        for row_idx in range(1, scan_max_row + 1):
            cols_with_data = []
            for col_idx in range(1, scan_max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cols_with_data.append(col_idx)
            if cols_with_data:
                row_has_data[row_idx] = cols_with_data

        if not row_has_data:
            return []

        # Split into regions by empty row gaps
        sorted_rows = sorted(row_has_data.keys())
        regions_raw: list[list[int]] = []
        current_region: list[int] = [sorted_rows[0]]

        for i in range(1, len(sorted_rows)):
            gap = sorted_rows[i] - sorted_rows[i - 1]
            if gap > self.empty_row_gap:
                regions_raw.append(current_region)
                current_region = [sorted_rows[i]]
            else:
                current_region.append(sorted_rows[i])

        if current_region:
            regions_raw.append(current_region)

        # Build table region records
        regions: list[ExcelTableRegion] = []
        for region_rows in regions_raw:
            if len(region_rows) < self.min_region_rows:
                continue

            # Determine column bounds for this region
            all_cols: set[int] = set()
            for r in region_rows:
                all_cols.update(row_has_data.get(r, []))

            if len(all_cols) < self.min_region_cols:
                continue

            min_col = min(all_cols)
            max_col_r = max(all_cols)
            min_row = min(region_rows)
            max_row_r = max(region_rows)

            # Cell range string
            cell_range = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col_r)}{max_row_r}"
            )

            # Detect header rows (first 1-3 rows of region)
            header_rows = self._detect_header_rows(ws, region_rows, min_col, max_col_r)
            data_start = header_rows[-1] + 1 if header_rows else min_row
            data_end = max_row_r

            # Extract column names from header
            columns = self._extract_column_names(ws, header_rows, min_col, max_col_r)

            # Confidence based on region quality
            confidence = self._compute_confidence(
                row_count=len(region_rows),
                col_count=len(all_cols),
                has_header=bool(header_rows),
                columns=columns,
            )

            region_id = ExcelTableRegion.generate_id(sheet_record.sheet_id, cell_range)
            region = ExcelTableRegion(
                table_region_id=region_id,
                workbook_id=workbook_id,
                sheet_id=sheet_record.sheet_id,
                sheet_name=sheet_record.sheet_name,
                cell_range=cell_range,
                header_rows=header_rows,
                data_start_row=data_start,
                data_end_row=data_end,
                columns=columns,
                confidence=confidence,
                region_type=self._infer_region_type(columns, sheet_record.guessed_sheet_type),
                metadata={
                    "row_count": len(region_rows),
                    "col_count": len(all_cols),
                    "min_col": min_col,
                    "max_col": max_col_r,
                },
            )
            regions.append(region)

        logger.info(
            "Detected %d table regions in sheet '%s'",
            len(regions), sheet_record.sheet_name,
        )
        return regions

    def _detect_header_rows(
        self,
        ws: Any,
        region_rows: list[int],
        min_col: int,
        max_col: int,
    ) -> list[int]:
        """Detect which rows are header rows.

        Heuristics:
        - First row(s) that are mostly text (not numeric)
        - Rows where all values look like column labels
        - Up to 3 header rows (multi-row header support)
        """
        header_rows: list[int] = []
        max_header_check = min(3, len(region_rows))

        for i in range(max_header_check):
            row_idx = region_rows[i]
            text_count = 0
            numeric_count = 0
            total = 0

            for col_idx in range(min_col, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                total += 1
                val = cell.value
                if isinstance(val, (int, float)):
                    numeric_count += 1
                elif isinstance(val, str):
                    # Short text is more likely a header
                    if len(val.strip()) < 50:
                        text_count += 1
                    else:
                        numeric_count += 1

            if total == 0:
                break

            # If mostly text and short, likely a header
            text_ratio = text_count / total if total > 0 else 0
            if text_ratio >= 0.6:
                header_rows.append(row_idx)
            else:
                break  # Stop at first non-header row

        return header_rows

    def _extract_column_names(
        self,
        ws: Any,
        header_rows: list[int],
        min_col: int,
        max_col: int,
    ) -> list[str]:
        """Extract column names from header rows.

        For multi-row headers, joins values with '/'.
        """
        if not header_rows:
            # Generate default column names (A, B, C, ...)
            return [get_column_letter(c) for c in range(min_col, max_col + 1)]

        columns: list[str] = []
        for col_idx in range(min_col, max_col + 1):
            parts: list[str] = []
            for row_idx in header_rows:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    parts.append(str(cell.value).strip())
            col_name = "/".join(parts) if parts else get_column_letter(col_idx)
            columns.append(col_name)

        return columns

    def _compute_confidence(
        self,
        row_count: int,
        col_count: int,
        has_header: bool,
        columns: list[str],
    ) -> float:
        """Compute confidence score for a detected table region."""
        score = 0.0

        # Size bonus
        if row_count >= 5:
            score += 0.2
        if row_count >= 20:
            score += 0.1
        if col_count >= 3:
            score += 0.2
        if col_count >= 5:
            score += 0.1

        # Header bonus
        if has_header:
            score += 0.3

        # Named columns bonus
        named_cols = [c for c in columns if not c.isalpha() or len(c) > 2]
        if named_cols:
            score += 0.1

        return min(round(score, 2), 1.0)

    def _infer_region_type(self, columns: list[str], sheet_type: str) -> str:
        """Infer region type from columns and parent sheet type."""
        cols_lower = " ".join(columns).lower()

        if any(kw in cols_lower for kw in ["mapping", "マッピング", "項目", "field", "映射"]):
            return "field_mapping_table"
        if any(kw in cols_lower for kw in ["api", "interface", "インターフェース", "接口"]):
            return "api_table"
        if any(kw in cols_lower for kw in ["error", "エラー", "message", "メッセージ", "错误"]):
            return "error_code_table"
        if any(kw in cols_lower for kw in ["test", "テスト", "ケース", "测试"]):
            return "test_case_table"

        # Fall back to sheet type
        if sheet_type != "unknown_sheet":
            return sheet_type.replace("_sheet", "_table")

        return "unknown_table"
