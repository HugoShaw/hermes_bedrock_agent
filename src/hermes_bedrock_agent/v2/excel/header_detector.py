"""
Header detector — detect single/multi-row headers and normalize column names.

Handles:
- Single-row headers
- Multi-row headers (joined with '/')
- Merged cell header groups
- Duplicated column names (appends suffix)
- Empty header cells (fills from column letter)
- Determines likely data start row
"""
from __future__ import annotations

import logging
from typing import Any

from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class HeaderDetector:
    """Detect and normalize headers in an Excel table region.

    Parameters
    ----------
    max_header_rows : int
        Maximum number of rows to consider as potential headers.
    """

    def __init__(self, max_header_rows: int = 3) -> None:
        self.max_header_rows = max_header_rows

    def detect_headers(
        self,
        ws: Any,
        start_row: int,
        end_row: int,
        min_col: int,
        max_col: int,
        merged_ranges: list[str] | None = None,
    ) -> dict[str, Any]:
        """Detect headers in the given range.

        Parameters
        ----------
        ws : openpyxl Worksheet
        start_row : int
            First row of the region.
        end_row : int
            Last row of the region.
        min_col : int
            First column (1-based).
        max_col : int
            Last column (1-based).
        merged_ranges : list[str] or None
            Merged cell range strings for context.

        Returns
        -------
        dict with keys:
            - header_rows: list[int] — row numbers that are headers
            - data_start_row: int — first data row
            - columns: list[str] — normalized column names
            - raw_headers: list[list[str]] — raw header values per row
            - column_letters: list[str] — column letter identifiers
        """
        check_rows = min(self.max_header_rows, end_row - start_row + 1)
        header_rows: list[int] = []
        raw_headers: list[list[str]] = []

        for offset in range(check_rows):
            row_idx = start_row + offset
            row_values: list[str] = []
            text_count = 0
            numeric_count = 0
            empty_count = 0
            total = max_col - min_col + 1

            for col_idx in range(min_col, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    row_values.append("")
                    empty_count += 1
                elif isinstance(cell.value, (int, float)):
                    row_values.append(str(cell.value))
                    numeric_count += 1
                else:
                    val = str(cell.value).strip()
                    row_values.append(val)
                    if len(val) < 60:
                        text_count += 1
                    else:
                        numeric_count += 1

            non_empty = total - empty_count
            if non_empty == 0:
                break

            text_ratio = text_count / non_empty if non_empty > 0 else 0

            # A row is a header if it's mostly short text
            if text_ratio >= 0.5:
                header_rows.append(row_idx)
                raw_headers.append(row_values)
            else:
                break

        # Determine data start row
        data_start_row = (header_rows[-1] + 1) if header_rows else start_row

        # Normalize columns
        columns = self._normalize_columns(raw_headers, min_col, max_col)
        column_letters = [get_column_letter(c) for c in range(min_col, max_col + 1)]

        return {
            "header_rows": header_rows,
            "data_start_row": data_start_row,
            "columns": columns,
            "raw_headers": raw_headers,
            "column_letters": column_letters,
        }

    def _normalize_columns(
        self,
        raw_headers: list[list[str]],
        min_col: int,
        max_col: int,
    ) -> list[str]:
        """Normalize column names from multi-row headers.

        - Joins multi-row values with '/'
        - Fills empty cells from merged parent or column letter
        - Deduplicates column names
        """
        col_count = max_col - min_col + 1

        if not raw_headers:
            return [get_column_letter(c) for c in range(min_col, max_col + 1)]

        columns: list[str] = []
        for col_idx in range(col_count):
            parts: list[str] = []
            for row_values in raw_headers:
                if col_idx < len(row_values) and row_values[col_idx]:
                    parts.append(row_values[col_idx])
            if parts:
                col_name = "/".join(parts)
            else:
                col_name = get_column_letter(min_col + col_idx)
            columns.append(col_name)

        # Deduplicate
        columns = self._deduplicate_columns(columns)

        return columns

    def _deduplicate_columns(self, columns: list[str]) -> list[str]:
        """Append _2, _3, etc. to duplicate column names."""
        seen: dict[str, int] = {}
        result: list[str] = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                result.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 1
                result.append(col)
        return result
