"""
Workbook loader — opens Excel workbooks using openpyxl and extracts metadata.

Supports .xlsx and .xlsm files. Reports limitations for .xls (legacy format).
Preserves formulas, comments, merged cells, and sheet visibility.
"""
from __future__ import annotations

import logging
import tempfile
from pathlib import Path
from typing import Any

from hermes_bedrock_agent.v2.excel.excel_schema import ExcelWorkbookRecord, ExcelSheetRecord

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed — Excel parsing disabled")


class WorkbookLoader:
    """Load an Excel workbook and extract structural metadata.

    Parameters
    ----------
    dataset : str
        Dataset label.
    run_id : str
        Run identifier.
    preserve_formulas : bool
        If True, read formulas as strings (data_only=False).
    """

    def __init__(
        self,
        dataset: str = "sample_20260519",
        run_id: str = "sample_20260519_excel_v1",
        preserve_formulas: bool = True,
    ) -> None:
        self.dataset = dataset
        self.run_id = run_id
        self.preserve_formulas = preserve_formulas

    def load(self, file_path: str, source_path: str | None = None) -> tuple[ExcelWorkbookRecord, list[ExcelSheetRecord], Any]:
        """Load workbook and return (workbook_record, sheet_records, openpyxl_workbook).

        Parameters
        ----------
        file_path : str
            Local filesystem path to the Excel file.
        source_path : str or None
            Original S3 key or URI. If None, uses file_path.

        Returns
        -------
        Tuple of (ExcelWorkbookRecord, list[ExcelSheetRecord], openpyxl.Workbook)

        Raises
        ------
        RuntimeError
            If openpyxl is not available.
        ValueError
            If the file extension is unsupported.
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl is not installed. Install it with: pip install openpyxl"
            )

        path = Path(file_path)
        ext = path.suffix.lower()
        if source_path is None:
            source_path = str(path)

        if ext == ".xls":
            raise ValueError(
                f"Legacy .xls format is not supported by openpyxl. "
                f"File: {file_path}. "
                f"Recommended: convert to .xlsx using LibreOffice or install xlrd for read support."
            )

        if ext not in (".xlsx", ".xlsm"):
            raise ValueError(f"Unsupported Excel extension: {ext}. Supported: .xlsx, .xlsm")

        # Open workbook — data_only=False to preserve formula strings
        wb = openpyxl.load_workbook(
            file_path,
            read_only=False,  # Need full mode for comments and merged cells
            data_only=not self.preserve_formulas,
            keep_links=False,
        )

        # Build workbook record
        workbook_id = ExcelWorkbookRecord.generate_id(source_path, self.dataset)
        visible_sheets = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
        hidden_sheets = [s for s in wb.sheetnames if wb[s].sheet_state != "visible"]

        wb_record = ExcelWorkbookRecord(
            workbook_id=workbook_id,
            dataset=self.dataset,
            run_id=self.run_id,
            source_path=source_path,
            file_name=path.name,
            file_extension=ext,
            sheet_count=len(wb.sheetnames),
            visible_sheet_count=len(visible_sheets),
            hidden_sheet_count=len(hidden_sheets),
            metadata={
                "file_size_bytes": path.stat().st_size if path.exists() else 0,
                "sheet_names": wb.sheetnames,
                "visible_sheets": visible_sheets,
                "hidden_sheets": hidden_sheets,
            },
        )

        # Build sheet records
        sheet_records: list[ExcelSheetRecord] = []
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws: Worksheet = wb[sheet_name]
            is_visible = ws.sheet_state == "visible"

            # Count non-empty cells (limited scan)
            non_empty = 0
            has_formula = False
            has_comments = False
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            # Sample cells for formula/comment detection
            sample_limit = min(max_row * max_col, 5000)
            scanned = 0
            for row in ws.iter_rows(min_row=1, max_row=min(max_row, 200), max_col=min(max_col, 50)):
                for cell in row:
                    scanned += 1
                    if cell.value is not None:
                        non_empty += 1
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        has_formula = True
                    if cell.comment:
                        has_comments = True
                    if scanned >= sample_limit:
                        break
                if scanned >= sample_limit:
                    break

            # Get merged cell ranges
            merged_ranges = [str(r) for r in ws.merged_cells.ranges] if ws.merged_cells else []

            sheet_id = ExcelSheetRecord.generate_id(workbook_id, sheet_name, idx)
            sheet_record = ExcelSheetRecord(
                sheet_id=sheet_id,
                workbook_id=workbook_id,
                sheet_name=sheet_name,
                sheet_index=idx,
                visible=is_visible,
                max_row=max_row,
                max_column=max_col,
                non_empty_cell_count=non_empty,
                merged_cell_ranges=merged_ranges,
                has_formula=has_formula,
                has_comments=has_comments,
                guessed_sheet_type="unknown_sheet",
                confidence=0.0,
                metadata={
                    "scanned_cells": scanned,
                },
            )
            sheet_records.append(sheet_record)

        logger.info(
            "Loaded workbook %s: %d sheets (%d visible, %d hidden)",
            path.name, len(sheet_records), wb_record.visible_sheet_count, wb_record.hidden_sheet_count,
        )
        return wb_record, sheet_records, wb
