#!/usr/bin/env python3
"""
Improved Excel Sheet Screenshot Pipeline
=========================================
Fixes two critical issues from the original pipeline:
1. Font: DroidSansFallbackFull cannot render ASCII/Latin → replaced with NotoSansCJK-Regular.ttc
2. Rendering: PIL text-on-canvas doesn't show original Excel look → use LibreOffice PDF export

Strategy:
- For ALL sheets: Use LibreOffice headless to export to PDF (preserves formatting, shapes, colors, merged cells)
- For very wide sheets: ALSO generate Pillow-based tile renders with correct font for structural detail
- Map PDF pages to sheets using openpyxl sheet metadata
- Stitch multi-page sheets back into single images
"""

import subprocess
import os
import sys
import shutil
import math
import json
import time
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import openpyxl

# Configuration
FONT_PATH = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
FONT_SIZE = 12
CELL_WIDTH_PX = 120
CELL_HEIGHT_PX = 22
DPI = 200  # PDF render DPI
MAX_BEDROCK_DIM = 7900  # Bedrock max dimension

OUTPUT_BASE = os.path.expanduser("~/projects/hermes_bedrock_agent/outputs/fresh_parse_20260519")


def get_workbook_info(xlsx_path):
    """Get sheet names and basic info from workbook."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = []
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        sheets.append({
            'index': i,
            'name': name,
            'max_row': ws.max_row or 0,
            'max_col': ws.max_column or 0,
        })
    wb.close()
    return sheets


def export_workbook_to_pdf(xlsx_path, output_dir):
    """Export full workbook to PDF using LibreOffice headless."""
    os.makedirs(output_dir, exist_ok=True)
    result = subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'pdf',
         '--outdir', output_dir, xlsx_path],
        capture_output=True, text=True, timeout=120
    )
    if result.returncode != 0:
        print(f"  WARNING: LibreOffice PDF export failed: {result.stderr}")
        return None
    
    # Find the output PDF
    basename = Path(xlsx_path).stem
    pdf_path = os.path.join(output_dir, f"{basename}.pdf")
    if os.path.exists(pdf_path):
        return pdf_path
    # Try to find it with different name
    pdfs = [f for f in os.listdir(output_dir) if f.endswith('.pdf')]
    if pdfs:
        return os.path.join(output_dir, pdfs[0])
    return None


def export_single_sheet_to_pdf(xlsx_path, sheet_index, output_dir):
    """Export a single sheet by hiding all others, then converting to PDF."""
    os.makedirs(output_dir, exist_ok=True)
    
    # Create temporary copy
    tmp_file = os.path.join(output_dir, f"_tmp_sheet_{sheet_index}.xlsx")
    shutil.copy2(xlsx_path, tmp_file)
    
    # Hide all sheets except target
    wb = openpyxl.load_workbook(tmp_file)
    target_name = wb.sheetnames[sheet_index]
    
    for i, name in enumerate(wb.sheetnames):
        if i == sheet_index:
            wb[name].sheet_state = 'visible'
        else:
            wb[name].sheet_state = 'hidden'
    wb.save(tmp_file)
    wb.close()
    
    # Export to PDF
    result = subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'pdf',
         '--outdir', output_dir, tmp_file],
        capture_output=True, text=True, timeout=120
    )
    
    pdf_path = os.path.join(output_dir, f"_tmp_sheet_{sheet_index}.pdf")
    os.remove(tmp_file)
    
    if os.path.exists(pdf_path):
        return pdf_path
    return None


def pdf_to_pages(pdf_path, output_prefix, dpi=200):
    """Convert PDF to PNG pages using pdftoppm."""
    result = subprocess.run(
        ['pdftoppm', '-png', '-r', str(dpi), pdf_path, output_prefix],
        capture_output=True, text=True, timeout=120
    )
    # Find generated pages
    output_dir = os.path.dirname(output_prefix)
    prefix_base = os.path.basename(output_prefix)
    pages = sorted([
        os.path.join(output_dir, f) 
        for f in os.listdir(output_dir)
        if f.startswith(prefix_base) and f.endswith('.png')
    ])
    return pages


def stitch_pages_vertical(page_paths, output_path):
    """Stitch multiple PDF pages into one tall image."""
    if not page_paths:
        return None
    
    if len(page_paths) == 1:
        shutil.copy2(page_paths[0], output_path)
        return output_path
    
    images = [Image.open(p) for p in page_paths]
    
    # Calculate total dimensions
    max_width = max(img.width for img in images)
    total_height = sum(img.height for img in images)
    
    # Create stitched image
    stitched = Image.new('RGB', (max_width, total_height), 'white')
    y_offset = 0
    for img in images:
        stitched.paste(img, (0, y_offset))
        y_offset += img.height
    
    # Save
    stitched.save(output_path, 'PNG')
    
    for img in images:
        img.close()
    
    return output_path


def render_tile_pillow(ws, max_row, max_col, start_row, end_row, start_col, end_col):
    """Render a portion of an Excel sheet using PIL with NotoSansCJK font."""
    end_row = min(end_row, max_row)
    end_col = min(end_col, max_col)
    
    num_rows = end_row - start_row + 1
    num_cols = end_col - start_col + 1
    
    img_width = num_cols * CELL_WIDTH_PX
    img_height = num_rows * CELL_HEIGHT_PX
    
    # Cap to reasonable size
    img_width = min(img_width, 12000)
    img_height = min(img_height, 12000)
    
    img = Image.new('RGB', (img_width, img_height), 'white')
    draw = ImageDraw.Draw(img)
    
    try:
        font = ImageFont.truetype(FONT_PATH, FONT_SIZE)
    except Exception as e:
        print(f"  WARNING: Font load failed ({e}), using default")
        font = ImageFont.load_default()
    
    # Draw grid and cell values
    for row_idx in range(num_rows):
        actual_row = start_row + row_idx
        y = row_idx * CELL_HEIGHT_PX
        
        # Horizontal grid line
        draw.line([(0, y), (img_width, y)], fill='#CCCCCC', width=1)
        
        for col_idx in range(num_cols):
            actual_col = start_col + col_idx
            x = col_idx * CELL_WIDTH_PX
            
            # Vertical grid line (only on first row)
            if row_idx == 0:
                draw.line([(x, 0), (x, img_height)], fill='#CCCCCC', width=1)
            
            try:
                cell = ws.cell(actual_row, actual_col)
                value = cell.value
                if value is not None:
                    text = str(value)[:20]  # Show more chars than before (was 12)
                    # Style based on cell properties
                    is_header = actual_row <= 3 or (cell.font and cell.font.bold)
                    color = '#000066' if is_header else '#333333'
                    draw.text((x + 2, y + 3), text, fill=color, font=font)
            except Exception:
                pass
    
    # Final borders
    draw.line([(0, img_height - 1), (img_width - 1, img_height - 1)], fill='#CCCCCC', width=1)
    draw.line([(img_width - 1, 0), (img_width - 1, img_height - 1)], fill='#CCCCCC', width=1)
    
    return img


def generate_pillow_tiles(ws, max_row, max_col, output_dir):
    """Generate tiled screenshots using Pillow with correct font."""
    os.makedirs(output_dir, exist_ok=True)
    tiles = []
    
    cols_per_tile = 50
    rows_per_tile = 150
    
    num_col_tiles = math.ceil(max_col / cols_per_tile)
    num_row_tiles = math.ceil(max_row / rows_per_tile)
    
    tile_idx = 0
    for row_tile in range(num_row_tiles):
        for col_tile in range(num_col_tiles):
            start_row = row_tile * rows_per_tile + 1
            end_row = min((row_tile + 1) * rows_per_tile, max_row)
            start_col = col_tile * cols_per_tile + 1
            end_col = min((col_tile + 1) * cols_per_tile, max_col)
            
            tile_idx += 1
            tile_path = os.path.join(output_dir, f"tile_{tile_idx:03d}.png")
            
            img = render_tile_pillow(ws, max_row, max_col, start_row, end_row, start_col, end_col)
            img.save(tile_path, 'PNG')
            img.close()
            
            tiles.append({
                'path': tile_path,
                'row_tile': row_tile,
                'col_tile': col_tile,
                'start_row': start_row,
                'end_row': end_row,
                'start_col': start_col,
                'end_col': end_col,
            })
    
    return tiles


def stitch_tiles(tiles, output_path):
    """Stitch tiles back into one full image."""
    if not tiles:
        return None
    
    # Group by row_tile
    max_row_tile = max(t['row_tile'] for t in tiles)
    max_col_tile = max(t['col_tile'] for t in tiles)
    
    # Open all tile images
    tile_images = {}
    for t in tiles:
        tile_images[(t['row_tile'], t['col_tile'])] = Image.open(t['path'])
    
    # Calculate dimensions
    row_heights = []
    for rt in range(max_row_tile + 1):
        h = max(tile_images[(rt, ct)].height for ct in range(max_col_tile + 1) if (rt, ct) in tile_images)
        row_heights.append(h)
    
    col_widths = []
    for ct in range(max_col_tile + 1):
        w = max(tile_images[(rt, ct)].width for rt in range(max_row_tile + 1) if (rt, ct) in tile_images)
        col_widths.append(w)
    
    total_width = sum(col_widths)
    total_height = sum(row_heights)
    
    # Stitch
    stitched = Image.new('RGB', (total_width, total_height), 'white')
    y = 0
    for rt in range(max_row_tile + 1):
        x = 0
        for ct in range(max_col_tile + 1):
            if (rt, ct) in tile_images:
                stitched.paste(tile_images[(rt, ct)], (x, y))
            x += col_widths[ct]
        y += row_heights[rt]
    
    # Close tile images
    for img in tile_images.values():
        img.close()
    
    # Resize if too large for Bedrock
    w, h = stitched.size
    if max(w, h) > MAX_BEDROCK_DIM:
        ratio = MAX_BEDROCK_DIM / max(w, h)
        new_w = int(w * ratio)
        new_h = int(h * ratio)
        stitched = stitched.resize((new_w, new_h), Image.LANCZOS)
    
    stitched.save(output_path, 'PNG')
    stitched.close()
    return output_path


def process_sheet(xlsx_path, sheet_info, workbook_output_dir, tmp_dir):
    """Process a single sheet: generate LibreOffice PDF screenshot + optional Pillow tiles."""
    idx = sheet_info['index']
    name = sheet_info['name']
    max_row = sheet_info['max_row']
    max_col = sheet_info['max_col']
    
    safe_name = name.replace('/', '_').replace('\\', '_').replace(' ', '_')[:40]
    sheet_dir = os.path.join(workbook_output_dir, "sheets", f"{idx:02d}_{safe_name}")
    screenshots_dir = os.path.join(sheet_dir, "screenshots")
    os.makedirs(screenshots_dir, exist_ok=True)
    
    print(f"  Processing sheet {idx}: {name} ({max_row}x{max_col})")
    
    # Strategy 1: LibreOffice per-sheet PDF export
    pdf_path = export_single_sheet_to_pdf(xlsx_path, idx, tmp_dir)
    lo_pages = []
    
    if pdf_path and os.path.exists(pdf_path):
        page_prefix = os.path.join(tmp_dir, f"sheet_{idx}_page")
        lo_pages = pdf_to_pages(pdf_path, page_prefix, dpi=DPI)
        
        # Save individual pages as tiles
        for pi, page_path in enumerate(lo_pages):
            tile_dest = os.path.join(screenshots_dir, f"tile_{pi+1:03d}.png")
            shutil.copy2(page_path, tile_dest)
        
        # Stitch pages into full sheet image
        stitched_path = os.path.join(screenshots_dir, "stitched_full_sheet.png")
        stitch_pages_vertical(lo_pages, stitched_path)
        
        # Clean tmp
        os.remove(pdf_path)
        for p in lo_pages:
            os.remove(p)
        
        print(f"    LibreOffice export: {len(lo_pages)} pages → stitched")
    else:
        print(f"    WARNING: LibreOffice export failed, falling back to Pillow render")
        lo_pages = []
    
    # Strategy 2: If sheet is wide (>50 cols) or LO failed, also generate Pillow tiles
    is_wide = max_col > 50
    needs_pillow = is_wide or not lo_pages
    
    if needs_pillow and max_row > 0 and max_col > 0:
        print(f"    Also generating Pillow tiles (wide sheet: {max_col} cols)")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[name]
        
        # Force read_only sheet to cache data
        pillow_dir = os.path.join(screenshots_dir, "pillow_tiles")
        os.makedirs(pillow_dir, exist_ok=True)
        
        # For read_only mode we need to iterate first
        wb2 = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        ws2 = wb2[name]
        
        tiles = generate_pillow_tiles(ws2, max_row, max_col, pillow_dir)
        
        if not lo_pages:
            # If LO failed, use Pillow stitched as main
            stitched_path = os.path.join(screenshots_dir, "stitched_full_sheet.png")
            stitch_tiles(tiles, stitched_path)
            # Also copy tiles to main screenshots dir
            for t in tiles:
                tile_name = os.path.basename(t['path'])
                shutil.copy2(t['path'], os.path.join(screenshots_dir, tile_name))
        
        wb2.close()
        wb.close()
    
    return sheet_dir


def process_workbook(xlsx_path, workbook_name, tmp_base):
    """Process all sheets in a workbook."""
    workbook_dir = os.path.join(OUTPUT_BASE, workbook_name)
    tmp_dir = os.path.join(tmp_base, workbook_name)
    os.makedirs(tmp_dir, exist_ok=True)
    
    print(f"\nProcessing workbook: {workbook_name}")
    sheets = get_workbook_info(xlsx_path)
    print(f"  Found {len(sheets)} sheets")
    
    for sheet_info in sheets:
        try:
            process_sheet(xlsx_path, sheet_info, workbook_dir, tmp_dir)
        except Exception as e:
            print(f"  ERROR on sheet {sheet_info['index']} ({sheet_info['name']}): {e}")
    
    # Clean tmp
    shutil.rmtree(tmp_dir, ignore_errors=True)


def main():
    """Main entry: re-render all screenshots for both workbooks."""
    source_dir = "/tmp/s3_downloads/サンプル20260519"
    tmp_base = "/tmp/screenshot_fix_tmp"
    os.makedirs(tmp_base, exist_ok=True)
    
    workbooks = [
        ("01_基本設計/M社様_DSSスクリプト改修概要_フローチャート.xlsx",
         "M社様_DSSスクリプト改修概要_フローチャート"),
        ("02_詳細設計/MW_IFマッピング定義書_205_発注情報(登録・変更・取消).xlsx",
         "MW_IFマッピング定義書_205_発注情報(登録・変更・取消)"),
    ]
    
    start = time.time()
    
    for rel_path, wb_name in workbooks:
        xlsx_path = os.path.join(source_dir, rel_path)
        if not os.path.exists(xlsx_path):
            print(f"SKIP: {xlsx_path} not found")
            continue
        process_workbook(xlsx_path, wb_name, tmp_base)
    
    elapsed = time.time() - start
    print(f"\nDone in {elapsed:.0f}s")
    
    # Clean up
    shutil.rmtree(tmp_base, ignore_errors=True)


if __name__ == '__main__':
    main()
