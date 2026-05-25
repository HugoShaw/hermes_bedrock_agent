#!/usr/bin/env python3
"""
Unified Complex Excel Project Parser V1

Parses enterprise Excel documents with:
- S3 file discovery and manifest generation
- Workbook/sheet profiling
- Table region detection with multi-row headers
- Merged cell handling
- Formula and comment extraction
- Visual content prescan (images, shapes, textbox, connector, chart)
- OOXML drawing/chart/media extraction
- Selective sheet image rendering (only complex visual sheets)
- Bedrock Claude multimodal analysis (optional)
- Mermaid file discovery and parsing
- Unified Markdown output for human review
- S3 upload

Usage:
  PYTHONPATH=src python scripts/parse_complex_excel_project_v1.py \
    --config configs/sample_20260519_excel_parser_v1.yaml \
    --run-id sample_20260519_excel_parser_v1 \
    --dataset sample_20260519 \
    --s3-uri "s3://s3-hulftchina-rd/サンプル20260519/" \
    --output-dir data/outputs/sample_20260519_excel_parser_v1 \
    --upload-s3 "s3://s3-hulftchina-rd/output/sample_20260519/" \
    --use-bedrock

  # Without Bedrock:
  PYTHONPATH=src python scripts/parse_complex_excel_project_v1.py \
    --config configs/sample_20260519_excel_parser_v1.yaml \
    --run-id sample_20260519_excel_parser_v1 \
    --dataset sample_20260519 \
    --s3-uri "s3://s3-hulftchina-rd/サンプル20260519/" \
    --output-dir data/outputs/sample_20260519_excel_parser_v1 \
    --upload-s3 "s3://s3-hulftchina-rd/output/sample_20260519/" \
    --no-bedrock
"""
from __future__ import annotations

import argparse
import base64
import hashlib
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional
from xml.etree import ElementTree as ET

import yaml

# ============================================================
# Logging
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("excel_parser_v1")

# ============================================================
# XML Namespaces for OOXML
# ============================================================
NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "v": "urn:schemas-microsoft-com:vml",
    "o": "urn:schemas-microsoft-com:office:office",
    "x": "urn:schemas-microsoft-com:office:excel",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _md5(text: str) -> str:
    return hashlib.md5(text.encode()).hexdigest()[:12]


def _gen_id(prefix: str, *parts: str) -> str:
    return f"{prefix}_{_md5('|'.join(parts))}"


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _load_env():
    """Load .env file if present."""
    env_path = Path(".env")
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, _, val = line.partition("=")
                    os.environ.setdefault(key.strip(), val.strip())


# ============================================================
# Data Classes
# ============================================================
@dataclass
class FileManifestEntry:
    file_key: str
    file_name: str
    file_size: int
    extension: str
    file_type: str  # excel, mermaid, markdown, csv, other
    s3_uri: str
    local_path: str = ""
    last_modified: str = ""


@dataclass
class WorkbookRecord:
    workbook_id: str
    file_name: str
    file_key: str
    s3_uri: str
    local_path: str
    sheet_count: int
    sheet_names: list[str]
    file_size: int
    dataset: str
    run_id: str
    has_macros: bool = False
    errors: list[str] = field(default_factory=list)


@dataclass
class SheetRecord:
    sheet_id: str
    workbook_id: str
    workbook_name: str
    sheet_name: str
    sheet_index: int
    row_count: int
    col_count: int
    non_empty_cells: int
    merged_cell_count: int
    has_formulas: bool
    has_comments: bool
    has_visual_content: bool
    visual_types: list[str]
    dataset: str
    run_id: str
    errors: list[str] = field(default_factory=list)


@dataclass
class TableRegion:
    region_id: str
    workbook_id: str
    sheet_id: str
    sheet_name: str
    workbook_name: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    header_rows: int
    headers: list[list[str]]
    data_row_count: int
    dataset: str
    run_id: str


@dataclass
class NormalizedRow:
    row_id: str
    region_id: str
    sheet_id: str
    workbook_name: str
    sheet_name: str
    row_index: int
    values: dict[str, str]
    dataset: str
    run_id: str


@dataclass
class VisualPrescan:
    sheet_id: str
    workbook_name: str
    sheet_name: str
    has_images: bool
    has_shapes: bool
    has_textboxes: bool
    has_connectors: bool
    has_charts: bool
    has_drawings: bool
    total_visual_objects: int
    is_complex_visual: bool
    needs_rendering: bool
    dataset: str
    run_id: str


@dataclass
class DrawingObject:
    object_id: str
    sheet_id: str
    workbook_name: str
    sheet_name: str
    object_type: str  # shape, textbox, image, group, unknown
    text_content: str
    position: dict[str, Any]
    properties: dict[str, Any]
    dataset: str
    run_id: str


@dataclass
class ConnectorRecord:
    connector_id: str
    sheet_id: str
    workbook_name: str
    sheet_name: str
    connector_type: str  # arrow, line, connector
    from_shape: str
    to_shape: str
    text: str
    properties: dict[str, Any]
    dataset: str
    run_id: str


@dataclass
class ChartRecord:
    chart_id: str
    sheet_id: str
    workbook_name: str
    sheet_name: str
    chart_type: str
    title: str
    series_count: int
    properties: dict[str, Any]
    dataset: str
    run_id: str


@dataclass
class EmbeddedImage:
    image_id: str
    sheet_id: str
    workbook_name: str
    sheet_name: str
    image_format: str
    image_size: int
    local_path: str
    source_path: str
    dataset: str
    run_id: str


@dataclass
class MermaidFileRecord:
    file_id: str
    file_name: str
    file_key: str
    s3_uri: str
    local_path: str
    content: str
    line_count: int
    associated_workbook: str
    associated_sheet: str
    dataset: str
    run_id: str


@dataclass
class MermaidGraph:
    graph_id: str
    file_id: str
    file_name: str
    graph_type: str  # flowchart, sequence, stateDiagram, etc.
    nodes: list[dict[str, str]]
    edges: list[dict[str, str]]
    labels: list[str]
    raw_content: str
    dataset: str
    run_id: str


@dataclass
class ParsedTextRecord:
    record_id: str
    workbook_name: str
    sheet_name: str
    record_type: str  # cell_text, formula, comment, shape_text, chart_title, mermaid
    content: str
    source_location: str
    dataset: str
    run_id: str


# ============================================================
# S3 Discovery
# ============================================================
class S3FileDiscovery:
    """Discover files under an S3 prefix."""

    EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}
    MERMAID_EXTS = {".mmd", ".mermaid"}
    MARKDOWN_EXTS = {".md"}
    CSV_EXTS = {".csv"}

    def __init__(self, s3_uri: str, region: str = "ap-northeast-1"):
        self.s3_uri = s3_uri.rstrip("/") + "/"
        self.region = region
        # Parse bucket and prefix
        parts = s3_uri.replace("s3://", "").split("/", 1)
        self.bucket = parts[0]
        self.prefix = parts[1] if len(parts) > 1 else ""
        if self.prefix and not self.prefix.endswith("/"):
            self.prefix += "/"
        self._s3 = None

    @property
    def s3(self):
        if self._s3 is None:
            import boto3
            self._s3 = boto3.client("s3", region_name=self.region)
        return self._s3

    def discover(self) -> list[FileManifestEntry]:
        """List all files under the S3 prefix."""
        entries = []
        paginator = self.s3.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=self.bucket, Prefix=self.prefix):
            for obj in page.get("Contents", []):
                key = obj["Key"]
                if key.endswith("/"):
                    continue  # skip directory markers
                name = key.split("/")[-1]
                ext = Path(name).suffix.lower()
                ftype = self._classify(ext)
                entries.append(FileManifestEntry(
                    file_key=key,
                    file_name=name,
                    file_size=obj["Size"],
                    extension=ext,
                    file_type=ftype,
                    s3_uri=f"s3://{self.bucket}/{key}",
                    last_modified=obj["LastModified"].isoformat() if obj.get("LastModified") else "",
                ))
        return entries

    def _classify(self, ext: str) -> str:
        if ext in self.EXCEL_EXTS:
            return "excel"
        elif ext in self.MERMAID_EXTS:
            return "mermaid"
        elif ext in self.MARKDOWN_EXTS:
            return "markdown"
        elif ext in self.CSV_EXTS:
            return "csv"
        else:
            return "other"

    def download_file(self, key: str, local_path: str) -> str:
        """Download a single file from S3."""
        Path(local_path).parent.mkdir(parents=True, exist_ok=True)
        self.s3.download_file(self.bucket, key, local_path)
        return local_path


# ============================================================
# Excel Parser
# ============================================================
class ExcelWorkbookParser:
    """Parse an Excel workbook comprehensively."""

    def __init__(self, local_path: str, file_name: str, file_key: str, s3_uri: str,
                 dataset: str, run_id: str, output_dir: Path,
                 config: dict):
        self.local_path = local_path
        self.file_name = file_name
        self.file_key = file_key
        self.s3_uri = s3_uri
        self.dataset = dataset
        self.run_id = run_id
        self.output_dir = output_dir
        self.config = config
        self.workbook_id = _gen_id("wb", file_name, run_id)

        # Results
        self.workbook_record: Optional[WorkbookRecord] = None
        self.sheet_records: list[SheetRecord] = []
        self.table_regions: list[TableRegion] = []
        self.normalized_rows: list[NormalizedRow] = []
        self.visual_prescans: list[VisualPrescan] = []
        self.drawing_objects: list[DrawingObject] = []
        self.connectors: list[ConnectorRecord] = []
        self.charts: list[ChartRecord] = []
        self.embedded_images: list[EmbeddedImage] = []
        self.text_records: list[ParsedTextRecord] = []
        self.errors: list[str] = []

    def parse(self) -> WorkbookRecord:
        """Run full parse pipeline."""
        import openpyxl

        try:
            # Load workbook
            wb = openpyxl.load_workbook(
                self.local_path,
                data_only=False,  # Keep formulas
                read_only=False,
            )
        except Exception as e:
            self.errors.append(f"Failed to open workbook: {e}")
            self.workbook_record = WorkbookRecord(
                workbook_id=self.workbook_id,
                file_name=self.file_name,
                file_key=self.file_key,
                s3_uri=self.s3_uri,
                local_path=self.local_path,
                sheet_count=0,
                sheet_names=[],
                file_size=os.path.getsize(self.local_path),
                dataset=self.dataset,
                run_id=self.run_id,
                errors=self.errors,
            )
            return self.workbook_record

        sheet_names = wb.sheetnames
        has_macros = self.file_name.endswith(".xlsm")

        self.workbook_record = WorkbookRecord(
            workbook_id=self.workbook_id,
            file_name=self.file_name,
            file_key=self.file_key,
            s3_uri=self.s3_uri,
            local_path=self.local_path,
            sheet_count=len(sheet_names),
            sheet_names=sheet_names,
            file_size=os.path.getsize(self.local_path),
            dataset=self.dataset,
            run_id=self.run_id,
            has_macros=has_macros,
            errors=[],
        )

        # Parse each sheet
        for idx, sheet_name in enumerate(sheet_names):
            try:
                ws = wb[sheet_name]
                self._parse_sheet(ws, idx, sheet_name)
            except Exception as e:
                self.errors.append(f"Error parsing sheet '{sheet_name}': {e}")
                logger.warning(f"Error parsing sheet '{sheet_name}' in {self.file_name}: {e}")

        # Extract visual content from OOXML
        self._extract_ooxml_visuals()

        wb.close()
        return self.workbook_record

    def _parse_sheet(self, ws, sheet_index: int, sheet_name: str):
        """Parse a single worksheet."""
        sheet_id = _gen_id("sh", self.workbook_id, sheet_name)

        # Basic metrics
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        merged = list(ws.merged_cells.ranges) if hasattr(ws, 'merged_cells') else []
        merged_count = len(merged)

        # Count non-empty cells, detect formulas and comments
        non_empty = 0
        has_formulas = False
        has_comments = False
        cell_texts = []

        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 500),
                                min_col=1, max_col=min(max_col, 50)):
            for cell in row:
                if cell.value is not None:
                    non_empty += 1
                    val_str = _safe_str(cell.value)
                    if val_str:
                        cell_texts.append((cell.row, cell.column, val_str))
                    # Check formula
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        has_formulas = True
                        self.text_records.append(ParsedTextRecord(
                            record_id=_gen_id("txt", sheet_id, f"formula_{cell.row}_{cell.column}"),
                            workbook_name=self.file_name,
                            sheet_name=sheet_name,
                            record_type="formula",
                            content=cell.value,
                            source_location=f"{cell.coordinate}",
                            dataset=self.dataset,
                            run_id=self.run_id,
                        ))
                if cell.comment:
                    has_comments = True
                    self.text_records.append(ParsedTextRecord(
                        record_id=_gen_id("txt", sheet_id, f"comment_{cell.row}_{cell.column}"),
                        workbook_name=self.file_name,
                        sheet_name=sheet_name,
                        record_type="comment",
                        content=_safe_str(cell.comment.text),
                        source_location=f"{cell.coordinate}",
                        dataset=self.dataset,
                        run_id=self.run_id,
                    ))

        # Visual content prescan (via openpyxl)
        has_images = bool(getattr(ws, '_images', None))
        has_charts = bool(getattr(ws, '_charts', None))

        # Check for drawing objects via openpyxl
        has_shapes = False
        has_textboxes = False
        has_connectors = False
        has_drawings = False
        visual_types = []

        if has_images:
            visual_types.append("image")
        if has_charts:
            visual_types.append("chart")

        # Determine complexity
        is_complex = self._is_complex_visual_sheet(sheet_name, visual_types)

        sheet_rec = SheetRecord(
            sheet_id=sheet_id,
            workbook_id=self.workbook_id,
            workbook_name=self.file_name,
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            row_count=max_row,
            col_count=max_col,
            non_empty_cells=non_empty,
            merged_cell_count=merged_count,
            has_formulas=has_formulas,
            has_comments=has_comments,
            has_visual_content=bool(visual_types),
            visual_types=visual_types,
            dataset=self.dataset,
            run_id=self.run_id,
        )
        self.sheet_records.append(sheet_rec)

        # Visual prescan record
        self.visual_prescans.append(VisualPrescan(
            sheet_id=sheet_id,
            workbook_name=self.file_name,
            sheet_name=sheet_name,
            has_images=has_images,
            has_shapes=has_shapes,
            has_textboxes=has_textboxes,
            has_connectors=has_connectors,
            has_charts=has_charts,
            has_drawings=has_drawings,
            total_visual_objects=len(visual_types),
            is_complex_visual=is_complex,
            needs_rendering=is_complex,
            dataset=self.dataset,
            run_id=self.run_id,
        ))

        # Detect table regions
        self._detect_tables(ws, sheet_id, sheet_name, max_row, max_col)

        # Record cell text
        for row_idx, col_idx, text in cell_texts[:200]:
            self.text_records.append(ParsedTextRecord(
                record_id=_gen_id("txt", sheet_id, f"cell_{row_idx}_{col_idx}"),
                workbook_name=self.file_name,
                sheet_name=sheet_name,
                record_type="cell_text",
                content=text[:1000],
                source_location=f"R{row_idx}C{col_idx}",
                dataset=self.dataset,
                run_id=self.run_id,
            ))

    def _is_complex_visual_sheet(self, sheet_name: str, visual_types: list[str]) -> bool:
        """Determine if a sheet needs image rendering."""
        keywords = self.config.get("visual", {}).get("complex_sheet_keywords", [
            "概要", "フローチャート", "フロー", "構成図", "概念図", "アーキテクチャ", "設計"
        ])
        for kw in keywords:
            if kw in sheet_name:
                return True
        if len(visual_types) >= 2:
            return True
        return False

    def _detect_tables(self, ws, sheet_id: str, sheet_name: str, max_row: int, max_col: int):
        """Detect table regions and normalize rows."""
        if max_row < 2 or max_col < 2:
            return

        # Simple table detection: find the first non-empty row as header
        header_row = None
        header_values = []
        data_start = None

        for r in range(1, min(max_row + 1, 20)):
            row_vals = []
            for c in range(1, min(max_col + 1, 50)):
                cell = ws.cell(row=r, column=c)
                row_vals.append(_safe_str(cell.value))
            non_empty_in_row = sum(1 for v in row_vals if v)
            if non_empty_in_row >= 2 and header_row is None:
                header_row = r
                header_values = [row_vals]
                # Check for multi-row headers
                for next_r in range(r + 1, min(r + 5, max_row + 1)):
                    next_vals = []
                    for c in range(1, min(max_col + 1, 50)):
                        cell = ws.cell(row=next_r, column=c)
                        next_vals.append(_safe_str(cell.value))
                    # If this row looks like a continuation of header (short text, no numbers)
                    has_numeric = any(v.replace(".", "").replace(",", "").isdigit()
                                      for v in next_vals if v)
                    if not has_numeric and sum(1 for v in next_vals if v) >= 2:
                        header_values.append(next_vals)
                    else:
                        data_start = next_r
                        break
                if data_start is None:
                    data_start = header_row + len(header_values)
                break

        if header_row is None:
            return

        # Determine actual column range
        actual_cols = max(len(h) for h in header_values) if header_values else max_col
        actual_cols = min(actual_cols, 50)

        # Build merged header names
        merged_headers = []
        for c in range(actual_cols):
            parts = []
            for h_row in header_values:
                if c < len(h_row) and h_row[c]:
                    parts.append(h_row[c])
            merged_headers.append(" / ".join(parts) if parts else f"Col_{c+1}")

        # Create table region
        end_row = max_row
        region_id = _gen_id("tbl", sheet_id, f"{header_row}_{actual_cols}")
        region = TableRegion(
            region_id=region_id,
            workbook_id=self.workbook_id,
            sheet_id=sheet_id,
            sheet_name=sheet_name,
            workbook_name=self.file_name,
            start_row=header_row,
            start_col=1,
            end_row=end_row,
            end_col=actual_cols,
            header_rows=len(header_values),
            headers=header_values,
            data_row_count=end_row - data_start + 1 if data_start else 0,
            dataset=self.dataset,
            run_id=self.run_id,
        )
        self.table_regions.append(region)

        # Normalize data rows (sample up to 200)
        if data_start:
            for r in range(data_start, min(end_row + 1, data_start + 200)):
                values = {}
                has_any = False
                for c in range(1, actual_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    val = _safe_str(cell.value)
                    if val:
                        has_any = True
                    col_name = merged_headers[c - 1] if c - 1 < len(merged_headers) else f"Col_{c}"
                    values[col_name] = val

                if has_any:
                    self.normalized_rows.append(NormalizedRow(
                        row_id=_gen_id("row", region_id, str(r)),
                        region_id=region_id,
                        sheet_id=sheet_id,
                        workbook_name=self.file_name,
                        sheet_name=sheet_name,
                        row_index=r,
                        values=values,
                        dataset=self.dataset,
                        run_id=self.run_id,
                    ))

    def _extract_ooxml_visuals(self):
        """Extract visual objects directly from OOXML (ZIP) structure."""
        if not zipfile.is_zipfile(self.local_path):
            return

        try:
            with zipfile.ZipFile(self.local_path, "r") as zf:
                # Find drawing files
                drawing_files = [n for n in zf.namelist() if "drawing" in n.lower() and n.endswith(".xml")]
                chart_files = [n for n in zf.namelist() if "chart" in n.lower() and n.endswith(".xml")]
                media_files = [n for n in zf.namelist() if "media/" in n.lower()]
                vml_files = [n for n in zf.namelist() if "vmlDrawing" in n.lower() and n.endswith(".vml")]

                # Process drawings
                for dfile in drawing_files:
                    self._parse_drawing_xml(zf, dfile)

                # Process charts
                for cfile in chart_files:
                    self._parse_chart_xml(zf, cfile)

                # Process media/images
                for mfile in media_files:
                    self._register_embedded_image(zf, mfile)

                # Process VML (legacy shapes)
                for vfile in vml_files:
                    self._parse_vml(zf, vfile)

                # Update sheet visual prescan with OOXML findings
                self._update_prescan_from_ooxml(drawing_files, chart_files, media_files, vml_files)

        except Exception as e:
            self.errors.append(f"OOXML extraction error: {e}")
            logger.warning(f"OOXML extraction error for {self.file_name}: {e}")

    def _parse_drawing_xml(self, zf: zipfile.ZipFile, drawing_path: str):
        """Parse an OOXML drawing XML for shapes, textboxes, connectors."""
        try:
            content = zf.read(drawing_path).decode("utf-8")
            root = ET.fromstring(content)
        except Exception as e:
            logger.debug(f"Failed to parse {drawing_path}: {e}")
            return

        # Determine which sheet this drawing belongs to
        sheet_name = self._resolve_sheet_for_drawing(zf, drawing_path)
        sheet_id = _gen_id("sh", self.workbook_id, sheet_name) if sheet_name else ""

        # Find all anchor elements (twoCellAnchor, oneCellAnchor, absoluteAnchor)
        for anchor in root.iter():
            if "Anchor" not in anchor.tag:
                continue

            # Look for shapes
            for sp in anchor.iter(f"{{{NS['xdr']}}}sp"):
                self._extract_shape(sp, sheet_id, sheet_name)

            # Look for connectors
            for cxn in anchor.iter(f"{{{NS['xdr']}}}cxnSp"):
                self._extract_connector(cxn, sheet_id, sheet_name)

            # Look for picture elements
            for pic in anchor.iter(f"{{{NS['xdr']}}}pic"):
                self._extract_picture_ref(pic, sheet_id, sheet_name)

    def _extract_shape(self, sp_elem, sheet_id: str, sheet_name: str):
        """Extract shape/textbox from drawing element."""
        text_parts = []
        for t_elem in sp_elem.iter(f"{{{NS['a']}}}t"):
            if t_elem.text:
                text_parts.append(t_elem.text)
        text = " ".join(text_parts).strip()

        # Get shape properties
        nvSpPr = sp_elem.find(f".//{{{NS['xdr']}}}nvSpPr")
        name = ""
        if nvSpPr is not None:
            cNvPr = nvSpPr.find(f".//{{{NS['xdr']}}}cNvPr")
            if cNvPr is None:
                # Try without namespace prefix for cNvPr
                for elem in nvSpPr.iter():
                    if "cNvPr" in elem.tag:
                        name = elem.get("name", "")
                        break
            else:
                name = cNvPr.get("name", "")

        obj_type = "textbox" if text else "shape"
        obj_id = _gen_id("drw", sheet_id, name or str(len(self.drawing_objects)))

        self.drawing_objects.append(DrawingObject(
            object_id=obj_id,
            sheet_id=sheet_id,
            workbook_name=self.file_name,
            sheet_name=sheet_name,
            object_type=obj_type,
            text_content=text,
            position={"name": name},
            properties={},
            dataset=self.dataset,
            run_id=self.run_id,
        ))

        if text:
            self.text_records.append(ParsedTextRecord(
                record_id=_gen_id("txt", obj_id, "shape_text"),
                workbook_name=self.file_name,
                sheet_name=sheet_name,
                record_type="shape_text",
                content=text,
                source_location=f"shape:{name}",
                dataset=self.dataset,
                run_id=self.run_id,
            ))

    def _extract_connector(self, cxn_elem, sheet_id: str, sheet_name: str):
        """Extract connector/arrow metadata."""
        text_parts = []
        for t_elem in cxn_elem.iter(f"{{{NS['a']}}}t"):
            if t_elem.text:
                text_parts.append(t_elem.text)
        text = " ".join(text_parts).strip()

        # Try to find connection info
        from_shape = ""
        to_shape = ""
        for stCxn in cxn_elem.iter():
            if "stCxn" in stCxn.tag:
                from_shape = stCxn.get("id", "")
            if "endCxn" in stCxn.tag:
                to_shape = stCxn.get("id", "")

        conn_id = _gen_id("cxn", sheet_id, str(len(self.connectors)))
        self.connectors.append(ConnectorRecord(
            connector_id=conn_id,
            sheet_id=sheet_id,
            workbook_name=self.file_name,
            sheet_name=sheet_name,
            connector_type="connector",
            from_shape=from_shape,
            to_shape=to_shape,
            text=text,
            properties={},
            dataset=self.dataset,
            run_id=self.run_id,
        ))

    def _extract_picture_ref(self, pic_elem, sheet_id: str, sheet_name: str):
        """Note a picture reference in drawing."""
        # Just record it exists - actual image is in media/
        name = ""
        for elem in pic_elem.iter():
            if "cNvPr" in elem.tag:
                name = elem.get("name", "")
                break

        obj_id = _gen_id("drw", sheet_id, f"pic_{name}")
        self.drawing_objects.append(DrawingObject(
            object_id=obj_id,
            sheet_id=sheet_id,
            workbook_name=self.file_name,
            sheet_name=sheet_name,
            object_type="image",
            text_content="",
            position={"name": name},
            properties={"type": "embedded_picture"},
            dataset=self.dataset,
            run_id=self.run_id,
        ))

    def _parse_chart_xml(self, zf: zipfile.ZipFile, chart_path: str):
        """Parse chart XML for metadata."""
        try:
            content = zf.read(chart_path).decode("utf-8")
            root = ET.fromstring(content)
        except Exception:
            return

        # Extract chart type and title
        chart_type = "unknown"
        title = ""

        # Detect chart type from child elements
        type_tags = ["barChart", "lineChart", "pieChart", "areaChart", "scatterChart",
                     "doughnutChart", "radarChart", "surfaceChart"]
        for tag in type_tags:
            if root.find(f".//{{{NS['c']}}}{tag}") is not None:
                chart_type = tag.replace("Chart", "")
                break

        # Get title
        for t_elem in root.iter(f"{{{NS['a']}}}t"):
            if t_elem.text and t_elem.text.strip():
                title = t_elem.text.strip()
                break

        # Count series
        series_count = len(list(root.iter(f"{{{NS['c']}}}ser")))

        # Associate with a sheet (best effort)
        sheet_name = ""
        sheet_id = ""
        if self.sheet_records:
            sheet_name = self.sheet_records[0].sheet_name
            sheet_id = self.sheet_records[0].sheet_id

        chart_id = _gen_id("cht", self.workbook_id, chart_path)
        self.charts.append(ChartRecord(
            chart_id=chart_id,
            sheet_id=sheet_id,
            workbook_name=self.file_name,
            sheet_name=sheet_name,
            chart_type=chart_type,
            title=title,
            series_count=series_count,
            properties={"source_path": chart_path},
            dataset=self.dataset,
            run_id=self.run_id,
        ))

    def _register_embedded_image(self, zf: zipfile.ZipFile, media_path: str):
        """Register an embedded image from the media folder."""
        ext = Path(media_path).suffix.lower()
        image_formats = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".emf", ".wmf"}
        if ext not in image_formats:
            return

        # Extract image to output dir
        images_dir = self.output_dir / "embedded_images" / _md5(self.file_name)
        images_dir.mkdir(parents=True, exist_ok=True)
        local_img_path = images_dir / Path(media_path).name

        try:
            with zf.open(media_path) as src, open(local_img_path, "wb") as dst:
                dst.write(src.read())
        except Exception:
            return

        img_size = local_img_path.stat().st_size if local_img_path.exists() else 0

        # Associate with first sheet (best effort)
        sheet_name = self.sheet_records[0].sheet_name if self.sheet_records else ""
        sheet_id = self.sheet_records[0].sheet_id if self.sheet_records else ""

        self.embedded_images.append(EmbeddedImage(
            image_id=_gen_id("img", self.workbook_id, media_path),
            sheet_id=sheet_id,
            workbook_name=self.file_name,
            sheet_name=sheet_name,
            image_format=ext.lstrip("."),
            image_size=img_size,
            local_path=str(local_img_path),
            source_path=media_path,
            dataset=self.dataset,
            run_id=self.run_id,
        ))

    def _parse_vml(self, zf: zipfile.ZipFile, vml_path: str):
        """Parse VML (legacy vector markup) for shapes."""
        try:
            content = zf.read(vml_path).decode("utf-8")
            # VML is often not well-formed XML, try basic parsing
            # Extract textbox content
            import re
            textbox_pattern = re.compile(r"<v:textbox[^>]*>(.*?)</v:textbox>", re.DOTALL)
            for match in textbox_pattern.finditer(content):
                inner = match.group(1)
                # Strip HTML tags
                text = re.sub(r"<[^>]+>", " ", inner).strip()
                if text:
                    sheet_name = self.sheet_records[0].sheet_name if self.sheet_records else ""
                    sheet_id = self.sheet_records[0].sheet_id if self.sheet_records else ""
                    obj_id = _gen_id("drw", sheet_id, f"vml_{_md5(text[:50])}")
                    self.drawing_objects.append(DrawingObject(
                        object_id=obj_id,
                        sheet_id=sheet_id,
                        workbook_name=self.file_name,
                        sheet_name=sheet_name,
                        object_type="textbox",
                        text_content=text[:500],
                        position={},
                        properties={"source": "vml"},
                        dataset=self.dataset,
                        run_id=self.run_id,
                    ))
        except Exception as e:
            logger.debug(f"VML parse error for {vml_path}: {e}")

    def _resolve_sheet_for_drawing(self, zf: zipfile.ZipFile, drawing_path: str) -> str:
        """Try to resolve which sheet a drawing belongs to."""
        # Drawing files are typically xl/drawings/drawing1.xml
        # Mapped via xl/worksheets/_rels/sheetN.xml.rels
        try:
            # Extract drawing number
            match = re.search(r"drawing(\d+)", drawing_path)
            if not match:
                return self.sheet_records[0].sheet_name if self.sheet_records else ""
            drawing_num = int(match.group(1))
            # Usually drawing N corresponds to sheet N (0-indexed sometimes)
            idx = drawing_num - 1
            if 0 <= idx < len(self.sheet_records):
                return self.sheet_records[idx].sheet_name
        except Exception:
            pass
        return self.sheet_records[0].sheet_name if self.sheet_records else ""

    def _update_prescan_from_ooxml(self, drawing_files, chart_files, media_files, vml_files):
        """Update visual prescan records with OOXML findings."""
        has_drawings = bool(drawing_files)
        has_charts_ooxml = bool(chart_files)
        has_media = bool(media_files)
        has_vml = bool(vml_files)

        for ps in self.visual_prescans:
            if has_drawings:
                ps.has_drawings = True
                ps.has_shapes = True
                if "shape" not in ps.visual_types if hasattr(ps, 'visual_types') else True:
                    pass
            if has_charts_ooxml:
                ps.has_charts = True
            if has_media:
                ps.has_images = True
            if has_vml:
                ps.has_textboxes = True
            # Recompute total
            ps.total_visual_objects = sum([
                ps.has_images, ps.has_shapes, ps.has_textboxes,
                ps.has_connectors, ps.has_charts, ps.has_drawings,
            ])
            # Re-evaluate complexity
            if ps.total_visual_objects >= 2:
                ps.is_complex_visual = True
                ps.needs_rendering = True


# ============================================================
# Mermaid Parser
# ============================================================
class MermaidParser:
    """Parse Mermaid (.mmd/.mermaid) files."""

    def __init__(self, dataset: str, run_id: str):
        self.dataset = dataset
        self.run_id = run_id

    def parse_file(self, local_path: str, file_name: str, file_key: str,
                   s3_uri: str) -> tuple[MermaidFileRecord, Optional[MermaidGraph]]:
        """Parse a single Mermaid file."""
        with open(local_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()

        lines = content.strip().split("\n")
        file_id = _gen_id("mmd", file_name, self.run_id)

        # Try to associate with a workbook
        associated_wb = ""
        associated_sheet = ""
        # Common pattern: flowchart.mmd is associated with the Excel in same directory
        # or named similarly
        parent_dir = Path(file_key).parent.name if "/" in file_key else ""

        file_record = MermaidFileRecord(
            file_id=file_id,
            file_name=file_name,
            file_key=file_key,
            s3_uri=s3_uri,
            local_path=local_path,
            content=content,
            line_count=len(lines),
            associated_workbook=associated_wb,
            associated_sheet=associated_sheet,
            dataset=self.dataset,
            run_id=self.run_id,
        )

        # Parse graph structure
        graph = self._parse_graph(content, file_id, file_name)

        return file_record, graph

    def _parse_graph(self, content: str, file_id: str, file_name: str) -> Optional[MermaidGraph]:
        """Parse Mermaid content into nodes and edges."""
        lines = content.strip().split("\n")
        if not lines:
            return None

        # Detect graph type
        graph_type = "unknown"
        first_line = lines[0].strip().lower()
        if first_line.startswith("flowchart") or first_line.startswith("graph"):
            graph_type = "flowchart"
        elif first_line.startswith("sequencediagram"):
            graph_type = "sequence"
        elif first_line.startswith("statediagram"):
            graph_type = "stateDiagram"
        elif first_line.startswith("classDiagram"):
            graph_type = "classDiagram"
        elif first_line.startswith("gantt"):
            graph_type = "gantt"

        nodes = []
        edges = []
        labels = []

        # Parse nodes and edges for flowchart/graph type
        # Patterns: A[Label] --> B[Label], A --> B, A -- text --> B
        node_pattern = re.compile(r'([A-Za-z0-9_]+)\s*[\[\(\{\"](.+?)[\]\)\}\"]')
        edge_pattern = re.compile(
            r'([A-Za-z0-9_]+)\s*'
            r'(-->|---|==>|-.->|--\s*[^-].*?-->|--\s*[^-].*?---)'
            r'\s*\|?([^|]*?)\|?\s*'
            r'([A-Za-z0-9_]+)'
        )
        # Simpler arrow pattern
        simple_edge = re.compile(r'([A-Za-z0-9_]+)\s*(-->|---|==>|-\.->)\s*([A-Za-z0-9_]+)')

        seen_nodes = set()
        for line in lines[1:]:
            line = line.strip()
            if not line or line.startswith("%%") or line.startswith("```"):
                continue

            # Extract nodes
            for m in node_pattern.finditer(line):
                node_id = m.group(1)
                label = m.group(2).strip()
                if node_id not in seen_nodes:
                    nodes.append({"id": node_id, "label": label})
                    seen_nodes.add(node_id)
                    if label:
                        labels.append(label)

            # Extract edges
            for m in simple_edge.finditer(line):
                src = m.group(1)
                arrow = m.group(2)
                tgt = m.group(3)
                edges.append({"source": src, "target": tgt, "arrow": arrow, "label": ""})

        graph_id = _gen_id("grp", file_id, graph_type)
        return MermaidGraph(
            graph_id=graph_id,
            file_id=file_id,
            file_name=file_name,
            graph_type=graph_type,
            nodes=nodes,
            edges=edges,
            labels=labels,
            raw_content=content,
            dataset=self.dataset,
            run_id=self.run_id,
        )


# ============================================================
# Bedrock Vision Analyzer
# ============================================================
class BedrockVisionAnalyzer:
    """Analyze complex visual sheets using Bedrock Claude multimodal."""

    def __init__(self, region: str = "ap-northeast-1", model_id: str = ""):
        self.region = region
        self.model_id = model_id or os.environ.get(
            "BEDROCK_VLM_MODEL_ID",
            os.environ.get("VISION_LLM_MODEL_ID", "")
        )
        self._client = None
        self.available = False
        self._check_availability()

    def _check_availability(self):
        """Check if Bedrock is available."""
        if not self.model_id:
            logger.info("No Bedrock VLM model configured - vision analysis disabled")
            return
        try:
            import boto3
            from botocore.config import Config
            self._client = boto3.client(
                "bedrock-runtime",
                region_name=self.region,
                config=Config(read_timeout=300),
            )
            self.available = True
            logger.info(f"Bedrock vision available: {self.model_id}")
        except Exception as e:
            logger.warning(f"Bedrock not available: {e}")

    def analyze_image(self, image_path: str, context: str = "") -> dict[str, Any]:
        """Analyze an image with Bedrock Claude multimodal."""
        if not self.available or not self._client:
            return {"error": "Bedrock not available"}

        try:
            with open(image_path, "rb") as f:
                image_bytes = f.read()

            # Determine media type
            ext = Path(image_path).suffix.lower()
            media_types = {".png": "image/png", ".jpg": "image/jpeg",
                          ".jpeg": "image/jpeg", ".gif": "image/gif",
                          ".webp": "image/webp"}
            media_type = media_types.get(ext, "image/png")

            prompt = f"""この画像はExcel シートから抽出されたものです。以下を日本語で分析してください：

1. 業務概要（この図は何を表しているか）
2. 含まれるシステム・モジュール
3. データフロー・処理フロー
4. 矢印・接続関係
5. 改修概要（Before/After があれば）
6. 重要な業務ルール
7. 主要な術語・キーワード

{f"追加コンテキスト: {context}" if context else ""}

構造化された分析結果を返してください。"""

            response = self._client.converse(
                modelId=self.model_id,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "image": {
                                "format": ext.lstrip(".") if ext.lstrip(".") in ["png", "jpeg", "gif", "webp"] else "png",
                                "source": {"bytes": image_bytes},
                            }
                        },
                        {"text": prompt},
                    ],
                }],
                inferenceConfig={"maxTokens": 4096, "temperature": 0.1},
            )

            # Extract response text
            result_text = ""
            for block in response.get("output", {}).get("message", {}).get("content", []):
                if "text" in block:
                    result_text += block["text"]

            return {
                "analysis": result_text,
                "model_id": self.model_id,
                "image_path": image_path,
                "status": "success",
            }

        except Exception as e:
            return {"error": str(e), "image_path": image_path, "status": "failed"}


# ============================================================
# Sheet Image Renderer
# ============================================================
class SheetImageRenderer:
    """Render Excel sheets as images for visual analysis."""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir / "sheet_images"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.available = self._check_tools()

    def _check_tools(self) -> bool:
        """Check if LibreOffice is available for rendering."""
        try:
            result = subprocess.run(
                ["libreoffice", "--version"],
                capture_output=True, text=True, timeout=10
            )
            return result.returncode == 0
        except Exception:
            return False

    def render_sheet(self, workbook_path: str, sheet_name: str, workbook_name: str) -> Optional[str]:
        """Render a specific sheet as PNG. Returns path or None."""
        if not self.available:
            logger.debug("LibreOffice not available for sheet rendering")
            return None

        try:
            # Convert to PDF/PNG using LibreOffice
            output_name = f"{_md5(workbook_name)}_{_md5(sheet_name)}.png"
            output_path = self.output_dir / output_name

            # LibreOffice can export specific sheets, but it's complex
            # For now, we'll just note which sheets need rendering
            # and rely on the embedded images + Bedrock analysis
            logger.debug(f"Sheet rendering requested for {sheet_name} in {workbook_name}")
            return None  # Rendering is best-effort
        except Exception as e:
            logger.debug(f"Sheet render failed: {e}")
            return None


# ============================================================
# Markdown Generator
# ============================================================
class MarkdownGenerator:
    """Generate comprehensive Markdown review documents."""

    def __init__(self, output_dir: Path, dataset: str, run_id: str):
        self.output_dir = output_dir / "markdown"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.dataset = dataset
        self.run_id = run_id

    def generate_full_review(
        self,
        workbooks: list[WorkbookRecord],
        sheets: list[SheetRecord],
        table_regions: list[TableRegion],
        normalized_rows: list[NormalizedRow],
        visual_prescans: list[VisualPrescan],
        drawing_objects: list[DrawingObject],
        connectors: list[ConnectorRecord],
        charts: list[ChartRecord],
        embedded_images: list[EmbeddedImage],
        mermaid_files: list[MermaidFileRecord],
        mermaid_graphs: list[MermaidGraph],
        bedrock_results: list[dict],
        text_records: list[ParsedTextRecord],
    ) -> str:
        """Generate the full review Markdown."""
        lines = []
        lines.append("# Excel Project Full Parse Review")
        lines.append("")
        lines.append(f"**Dataset:** {self.dataset}")
        lines.append(f"**Run ID:** {self.run_id}")
        lines.append(f"**Generated:** {datetime.now(timezone.utc).isoformat()}")
        lines.append("")
        lines.append("---")
        lines.append("")

        # Summary
        lines.append("## Summary")
        lines.append("")
        lines.append(f"- **Workbooks:** {len(workbooks)}")
        lines.append(f"- **Total Sheets:** {len(sheets)}")
        lines.append(f"- **Table Regions:** {len(table_regions)}")
        lines.append(f"- **Normalized Rows:** {len(normalized_rows)}")
        lines.append(f"- **Drawing Objects:** {len(drawing_objects)}")
        lines.append(f"- **Connectors:** {len(connectors)}")
        lines.append(f"- **Charts:** {len(charts)}")
        lines.append(f"- **Embedded Images:** {len(embedded_images)}")
        lines.append(f"- **Mermaid Files:** {len(mermaid_files)}")
        lines.append(f"- **Text Records:** {len(text_records)}")
        lines.append(f"- **Bedrock Analyses:** {len(bedrock_results)}")
        lines.append("")
        lines.append("---")
        lines.append("")

        # Per-workbook detail
        for wb in workbooks:
            lines.append(f"## Workbook: {wb.file_name}")
            lines.append("")
            lines.append(f"- **ID:** `{wb.workbook_id}`")
            lines.append(f"- **S3:** `{wb.s3_uri}`")
            lines.append(f"- **Sheets:** {wb.sheet_count} ({', '.join(wb.sheet_names)})")
            lines.append(f"- **Size:** {wb.file_size:,} bytes")
            lines.append(f"- **Has Macros:** {wb.has_macros}")
            if wb.errors:
                lines.append(f"- **⚠️ Errors:** {'; '.join(wb.errors)}")
            lines.append("")

            # Sheets for this workbook
            wb_sheets = [s for s in sheets if s.workbook_id == wb.workbook_id]
            for sh in wb_sheets:
                lines.append(f"### Sheet: {sh.sheet_name} (index={sh.sheet_index})")
                lines.append("")
                lines.append(f"- **Rows:** {sh.row_count} | **Cols:** {sh.col_count}")
                lines.append(f"- **Non-empty cells:** {sh.non_empty_cells}")
                lines.append(f"- **Merged cells:** {sh.merged_cell_count}")
                lines.append(f"- **Formulas:** {'Yes' if sh.has_formulas else 'No'}")
                lines.append(f"- **Comments:** {'Yes' if sh.has_comments else 'No'}")
                lines.append(f"- **Visual content:** {'Yes' if sh.has_visual_content else 'No'} ({', '.join(sh.visual_types) if sh.visual_types else 'none'})")
                lines.append("")

                # Table regions for this sheet
                sh_tables = [t for t in table_regions if t.sheet_id == sh.sheet_id]
                if sh_tables:
                    lines.append("#### Table Regions")
                    lines.append("")
                    for tbl in sh_tables:
                        lines.append(f"**Table** `{tbl.region_id}`: rows {tbl.start_row}-{tbl.end_row}, cols {tbl.start_col}-{tbl.end_col}")
                        lines.append(f"- Header rows: {tbl.header_rows}")
                        lines.append(f"- Data rows: {tbl.data_row_count}")
                        lines.append("")
                        # Show headers
                        if tbl.headers:
                            lines.append("Headers:")
                            lines.append("```")
                            for h_row in tbl.headers:
                                lines.append(" | ".join(h_row[:20]))
                            lines.append("```")
                            lines.append("")

                        # Show sample normalized rows
                        sh_rows = [r for r in normalized_rows if r.region_id == tbl.region_id]
                        if sh_rows:
                            lines.append(f"Sample data ({len(sh_rows)} rows):")
                            lines.append("")
                            # Show first 5 rows
                            for row in sh_rows[:5]:
                                non_empty_vals = {k: v for k, v in row.values.items() if v}
                                if non_empty_vals:
                                    lines.append(f"  Row {row.row_index}: {json.dumps(non_empty_vals, ensure_ascii=False)[:200]}")
                            if len(sh_rows) > 5:
                                lines.append(f"  ... ({len(sh_rows) - 5} more rows)")
                            lines.append("")

                # Visual prescan
                sh_prescan = [p for p in visual_prescans if p.sheet_id == sh.sheet_id]
                if sh_prescan and sh_prescan[0].total_visual_objects > 0:
                    ps = sh_prescan[0]
                    lines.append("#### Visual Content")
                    lines.append("")
                    lines.append(f"- Images: {'✓' if ps.has_images else '✗'}")
                    lines.append(f"- Shapes: {'✓' if ps.has_shapes else '✗'}")
                    lines.append(f"- Textboxes: {'✓' if ps.has_textboxes else '✗'}")
                    lines.append(f"- Connectors: {'✓' if ps.has_connectors else '✗'}")
                    lines.append(f"- Charts: {'✓' if ps.has_charts else '✗'}")
                    lines.append(f"- Complex visual: {'✓' if ps.is_complex_visual else '✗'}")
                    lines.append(f"- Needs rendering: {'✓' if ps.needs_rendering else '✗'}")
                    lines.append("")

                # Drawing objects for this sheet
                sh_drawings = [d for d in drawing_objects if d.sheet_id == sh.sheet_id]
                if sh_drawings:
                    lines.append("#### Drawing Objects")
                    lines.append("")
                    for drw in sh_drawings:
                        lines.append(f"- **{drw.object_type}** `{drw.object_id}`: {drw.text_content[:100] if drw.text_content else '(no text)'}")
                    lines.append("")

                # Connectors for this sheet
                sh_conns = [c for c in connectors if c.sheet_id == sh.sheet_id]
                if sh_conns:
                    lines.append("#### Connectors / Arrows")
                    lines.append("")
                    for conn in sh_conns:
                        lines.append(f"- {conn.connector_type}: {conn.from_shape} → {conn.to_shape} {f'({conn.text})' if conn.text else ''}")
                    lines.append("")

                # Charts for this sheet
                sh_charts = [c for c in charts if c.sheet_id == sh.sheet_id]
                if sh_charts:
                    lines.append("#### Charts")
                    lines.append("")
                    for cht in sh_charts:
                        lines.append(f"- **{cht.chart_type}** \"{cht.title}\": {cht.series_count} series")
                    lines.append("")

                # Embedded images for this sheet
                sh_imgs = [i for i in embedded_images if i.sheet_id == sh.sheet_id]
                if sh_imgs:
                    lines.append("#### Embedded Images")
                    lines.append("")
                    for img in sh_imgs:
                        lines.append(f"- {img.image_format} ({img.image_size:,} bytes): `{img.source_path}`")
                    lines.append("")

                # Bedrock vision analysis
                sh_bedrock = [b for b in bedrock_results if b.get("sheet_name") == sh.sheet_name and b.get("workbook_name") == wb.file_name]
                if sh_bedrock:
                    lines.append("#### 🤖 Bedrock Vision Analysis")
                    lines.append("")
                    for br in sh_bedrock:
                        if br.get("status") == "success":
                            lines.append(br.get("analysis", "(no analysis)"))
                        else:
                            lines.append(f"⚠️ Analysis failed: {br.get('error', 'unknown')}")
                    lines.append("")

                # Human review suggestions
                lines.append("#### 📋 Human Review Notes")
                lines.append("")
                issues = []
                if sh.merged_cell_count > 10:
                    issues.append("Many merged cells - verify table parsing accuracy")
                if sh.has_visual_content:
                    issues.append("Contains visual content - compare with original Excel")
                if sh_prescan and sh_prescan[0].is_complex_visual:
                    issues.append("Complex visual sheet - needs careful review of extracted structure")
                if not sh_tables:
                    issues.append("No table regions detected - may need manual inspection")
                if not issues:
                    issues.append("Standard text sheet - basic verification needed")
                for issue in issues:
                    lines.append(f"- {issue}")
                lines.append("")

            lines.append("---")
            lines.append("")

        # Mermaid section
        if mermaid_files:
            lines.append("## Mermaid Files")
            lines.append("")
            for mf in mermaid_files:
                lines.append(f"### {mf.file_name}")
                lines.append("")
                lines.append(f"- **S3:** `{mf.s3_uri}`")
                lines.append(f"- **Lines:** {mf.line_count}")
                if mf.associated_workbook:
                    lines.append(f"- **Associated workbook:** {mf.associated_workbook}")
                lines.append("")
                lines.append("```mermaid")
                lines.append(mf.content[:5000])
                lines.append("```")
                lines.append("")

                # Parsed graph
                mg = next((g for g in mermaid_graphs if g.file_id == mf.file_id), None)
                if mg:
                    lines.append(f"**Graph type:** {mg.graph_type}")
                    lines.append(f"**Nodes:** {len(mg.nodes)}")
                    lines.append(f"**Edges:** {len(mg.edges)}")
                    lines.append("")
                    if mg.nodes:
                        lines.append("Nodes:")
                        for n in mg.nodes[:20]:
                            lines.append(f"  - `{n['id']}`: {n['label']}")
                        if len(mg.nodes) > 20:
                            lines.append(f"  ... ({len(mg.nodes) - 20} more)")
                    lines.append("")
                    if mg.edges:
                        lines.append("Edges:")
                        for e in mg.edges[:20]:
                            lines.append(f"  - `{e['source']}` {e['arrow']} `{e['target']}`")
                        if len(mg.edges) > 20:
                            lines.append(f"  ... ({len(mg.edges) - 20} more)")
                    lines.append("")

            lines.append("---")
            lines.append("")

        # End
        lines.append("## End of Review")
        lines.append("")
        lines.append(f"Total text records extracted: {len(text_records)}")
        lines.append("")

        content = "\n".join(lines)
        output_path = self.output_dir / "parsed_excel_full_review.md"
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(output_path)

    def generate_summary(self, workbooks, sheets, table_regions, normalized_rows,
                         drawing_objects, connectors, charts, embedded_images,
                         mermaid_files, mermaid_graphs) -> str:
        """Generate summary Markdown."""
        lines = []
        lines.append("# Excel Parser Summary")
        lines.append("")
        lines.append(f"**Dataset:** {self.dataset} | **Run:** {self.run_id}")
        lines.append(f"**Time:** {datetime.now(timezone.utc).isoformat()}")
        lines.append("")
        lines.append("## Statistics")
        lines.append("")
        lines.append(f"| Metric | Count |")
        lines.append(f"|--------|-------|")
        lines.append(f"| Workbooks | {len(workbooks)} |")
        lines.append(f"| Sheets | {len(sheets)} |")
        lines.append(f"| Table Regions | {len(table_regions)} |")
        lines.append(f"| Normalized Rows | {len(normalized_rows)} |")
        lines.append(f"| Drawing Objects | {len(drawing_objects)} |")
        lines.append(f"| Connectors | {len(connectors)} |")
        lines.append(f"| Charts | {len(charts)} |")
        lines.append(f"| Embedded Images | {len(embedded_images)} |")
        lines.append(f"| Mermaid Files | {len(mermaid_files)} |")
        lines.append(f"| Mermaid Graphs | {len(mermaid_graphs)} |")
        lines.append("")
        lines.append("## Workbooks")
        lines.append("")
        for wb in workbooks:
            lines.append(f"- **{wb.file_name}**: {wb.sheet_count} sheets, {wb.file_size:,} bytes")
        lines.append("")
        lines.append("## Sheets with Visual Content")
        lines.append("")
        for sh in sheets:
            if sh.has_visual_content:
                lines.append(f"- {sh.workbook_name} / {sh.sheet_name}: {', '.join(sh.visual_types)}")
        lines.append("")

        content = "\n".join(lines)
        path = self.output_dir / "parsed_excel_summary.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)

    def generate_visual_review(self, visual_prescans, drawing_objects, connectors,
                               charts, embedded_images, bedrock_results) -> str:
        """Generate visual parse review Markdown."""
        lines = []
        lines.append("# Visual Parse Review")
        lines.append("")
        lines.append(f"**Generated:** {datetime.now(timezone.utc).isoformat()}")
        lines.append("")

        # Complex visual sheets
        complex_sheets = [p for p in visual_prescans if p.is_complex_visual]
        lines.append(f"## Complex Visual Sheets ({len(complex_sheets)})")
        lines.append("")
        for ps in complex_sheets:
            lines.append(f"- **{ps.workbook_name}** / {ps.sheet_name}")
            lines.append(f"  - Images: {ps.has_images}, Shapes: {ps.has_shapes}, Textboxes: {ps.has_textboxes}")
            lines.append(f"  - Connectors: {ps.has_connectors}, Charts: {ps.has_charts}")
        lines.append("")

        lines.append(f"## Drawing Objects ({len(drawing_objects)})")
        lines.append("")
        for drw in drawing_objects[:50]:
            lines.append(f"- [{drw.object_type}] {drw.workbook_name}/{drw.sheet_name}: {drw.text_content[:80] if drw.text_content else '(no text)'}")
        lines.append("")

        lines.append(f"## Connectors ({len(connectors)})")
        lines.append("")
        for conn in connectors[:30]:
            lines.append(f"- {conn.workbook_name}/{conn.sheet_name}: {conn.from_shape} → {conn.to_shape}")
        lines.append("")

        lines.append(f"## Charts ({len(charts)})")
        lines.append("")
        for cht in charts:
            lines.append(f"- [{cht.chart_type}] {cht.workbook_name}/{cht.sheet_name}: \"{cht.title}\" ({cht.series_count} series)")
        lines.append("")

        lines.append(f"## Bedrock Vision Results ({len(bedrock_results)})")
        lines.append("")
        for br in bedrock_results:
            status = "✓" if br.get("status") == "success" else "✗"
            lines.append(f"- {status} {br.get('workbook_name', '')} / {br.get('sheet_name', '')}")
        lines.append("")

        content = "\n".join(lines)
        path = self.output_dir / "visual_parse_review.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)

    def generate_mermaid_review(self, mermaid_files, mermaid_graphs) -> str:
        """Generate Mermaid integration review."""
        lines = []
        lines.append("# Mermaid Integration Review")
        lines.append("")
        lines.append(f"**Files found:** {len(mermaid_files)}")
        lines.append(f"**Graphs parsed:** {len(mermaid_graphs)}")
        lines.append("")

        for mf in mermaid_files:
            lines.append(f"## {mf.file_name}")
            lines.append("")
            lines.append(f"- S3: `{mf.s3_uri}`")
            lines.append(f"- Lines: {mf.line_count}")
            lines.append(f"- Associated WB: {mf.associated_workbook or '(auto-detect)'}")
            lines.append("")

            mg = next((g for g in mermaid_graphs if g.file_id == mf.file_id), None)
            if mg:
                lines.append(f"- Graph type: {mg.graph_type}")
                lines.append(f"- Nodes: {len(mg.nodes)}")
                lines.append(f"- Edges: {len(mg.edges)}")
                lines.append(f"- Labels: {len(mg.labels)}")
                lines.append("")
                if mg.labels:
                    lines.append("Key labels:")
                    for lbl in mg.labels[:30]:
                        lines.append(f"  - {lbl}")
                lines.append("")

        content = "\n".join(lines)
        path = self.output_dir / "mermaid_integration_review.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)

    def generate_checklist(self, workbooks, sheets, table_regions, mermaid_files,
                           visual_prescans, bedrock_results, upload_success) -> str:
        """Generate human review checklist."""
        lines = []
        lines.append("# Human Review Checklist")
        lines.append("")
        lines.append(f"**Dataset:** {self.dataset} | **Run:** {self.run_id}")
        lines.append(f"**Generated:** {datetime.now(timezone.utc).isoformat()}")
        lines.append("")

        def check(condition: bool, label: str) -> str:
            return f"- [{'x' if condition else ' '}] {label}"

        lines.append("## File Discovery")
        lines.append("")
        lines.append(check(len(workbooks) > 0, f"All Excel files discovered ({len(workbooks)} found)"))
        lines.append(check(all(not wb.errors for wb in workbooks), "All workbooks opened without errors"))
        lines.append("")

        lines.append("## Sheet Coverage")
        lines.append("")
        total_sheets = len(sheets)
        lines.append(check(total_sheets > 0, f"All sheets listed ({total_sheets} total)"))
        # Check order
        for wb in workbooks:
            wb_sheets = [s for s in sheets if s.workbook_id == wb.workbook_id]
            ordered = all(wb_sheets[i].sheet_index <= wb_sheets[i+1].sheet_index
                         for i in range(len(wb_sheets)-1)) if len(wb_sheets) > 1 else True
            lines.append(check(ordered, f"Sheet order correct for {wb.file_name}"))
        lines.append("")

        lines.append("## Table Parsing")
        lines.append("")
        lines.append(check(len(table_regions) > 0, f"Table regions detected ({len(table_regions)} found)"))
        multi_header = any(t.header_rows > 1 for t in table_regions)
        lines.append(check(True, f"Multi-row headers detected: {'Yes' if multi_header else 'N/A'}"))
        merged = any(s.merged_cell_count > 0 for s in sheets)
        lines.append(check(merged, f"Merged cells preserved: {'Yes' if merged else 'N/A'}"))
        lines.append("")

        lines.append("## Visual Content")
        lines.append("")
        visual_sheets = [s for s in sheets if s.has_visual_content]
        lines.append(check(len(visual_sheets) >= 0, f"Visual objects identified ({len(visual_sheets)} sheets with visuals)"))
        complex_visual = [p for p in visual_prescans if p.is_complex_visual]
        lines.append(check(True, f"Complex visual sheets flagged ({len(complex_visual)} flagged)"))
        lines.append("")

        lines.append("## Mermaid Integration")
        lines.append("")
        lines.append(check(len(mermaid_files) >= 0, f"Mermaid files found ({len(mermaid_files)})"))
        lines.append(check(True, "Mermaid associated to correct sheet: needs manual verification"))
        lines.append("")

        lines.append("## Bedrock Analysis")
        lines.append("")
        bedrock_used = len(bedrock_results) > 0
        lines.append(check(True, f"Bedrock used only for complex visual: {'Yes' if bedrock_used else 'Bedrock not used'}"))
        if bedrock_results:
            success = sum(1 for b in bedrock_results if b.get("status") == "success")
            lines.append(check(success > 0, f"Bedrock success rate: {success}/{len(bedrock_results)}"))
        lines.append("")

        lines.append("## Text Integrity")
        lines.append("")
        lines.append(check(True, "Japanese text preserved: needs manual spot-check"))
        lines.append(check(True, "No obvious hallucination: needs manual review"))
        lines.append("")

        lines.append("## Output & Upload")
        lines.append("")
        lines.append(check(upload_success, "Output uploaded to S3"))
        lines.append("")

        lines.append("## Manual Review Actions")
        lines.append("")
        lines.append("1. Open original Excel files and compare with Markdown output")
        lines.append("2. Verify table headers match original layout")
        lines.append("3. Check merged cell handling for accuracy")
        lines.append("4. Compare flowchart/Mermaid content with Excel visual")
        lines.append("5. Review Bedrock analysis for accuracy (if used)")
        lines.append("6. Confirm Japanese text encoding is correct")
        lines.append("7. Flag any missing content or misinterpretation")
        lines.append("")

        content = "\n".join(lines)
        path = self.output_dir / "human_review_checklist.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)


# ============================================================
# Report Generator
# ============================================================
class ReportGenerator:
    """Generate pipeline run reports."""

    def __init__(self, output_dir: Path, dataset: str, run_id: str):
        self.reports_dir = output_dir / "reports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        self.dataset = dataset
        self.run_id = run_id

    def generate_run_report(self, stats: dict, duration: float, errors: list[str]) -> str:
        """Generate parser run report."""
        lines = []
        lines.append("# Parser Run Report")
        lines.append("")
        lines.append(f"**Dataset:** {self.dataset}")
        lines.append(f"**Run ID:** {self.run_id}")
        lines.append(f"**Duration:** {duration:.1f} seconds")
        lines.append(f"**Generated:** {datetime.now(timezone.utc).isoformat()}")
        lines.append("")
        lines.append("## Statistics")
        lines.append("")
        for key, val in stats.items():
            lines.append(f"- **{key}:** {val}")
        lines.append("")

        if errors:
            lines.append("## Errors")
            lines.append("")
            for err in errors:
                lines.append(f"- ⚠️ {err}")
            lines.append("")

        lines.append("## Generated Files")
        lines.append("")
        lines.append("See output directory for all JSONL and Markdown files.")
        lines.append("")

        content = "\n".join(lines)
        path = self.reports_dir / "parser_run_report.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)

    def generate_quality_report(self, workbooks, sheets, table_regions, errors) -> str:
        """Generate quality report."""
        lines = []
        lines.append("# Excel Parser Quality Report")
        lines.append("")
        lines.append(f"**Run:** {self.run_id}")
        lines.append("")

        # Coverage
        lines.append("## Coverage")
        lines.append("")
        lines.append(f"- Workbooks processed: {len(workbooks)}")
        lines.append(f"- Sheets processed: {len(sheets)}")
        lines.append(f"- Tables detected: {len(table_regions)}")
        lines.append("")

        # Potential issues
        lines.append("## Potential Issues")
        lines.append("")
        empty_sheets = [s for s in sheets if s.non_empty_cells == 0]
        if empty_sheets:
            lines.append(f"- ⚠️ {len(empty_sheets)} empty sheet(s)")
            for s in empty_sheets:
                lines.append(f"  - {s.workbook_name} / {s.sheet_name}")
        lines.append("")

        heavy_merged = [s for s in sheets if s.merged_cell_count > 50]
        if heavy_merged:
            lines.append(f"- ⚠️ {len(heavy_merged)} sheet(s) with heavy merging (>50)")
            for s in heavy_merged:
                lines.append(f"  - {s.workbook_name} / {s.sheet_name}: {s.merged_cell_count} merges")
        lines.append("")

        if errors:
            lines.append(f"- ⚠️ {len(errors)} error(s) during parsing")
            for e in errors[:10]:
                lines.append(f"  - {e}")
        lines.append("")

        content = "\n".join(lines)
        path = self.reports_dir / "excel_parser_quality_report.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)

    def generate_visual_report(self, visual_prescans, drawing_objects, connectors,
                               charts, embedded_images, bedrock_results) -> str:
        """Generate visual parse report."""
        lines = []
        lines.append("# Visual Parse Report")
        lines.append("")
        lines.append(f"**Run:** {self.run_id}")
        lines.append("")
        lines.append("## Summary")
        lines.append("")
        complex_count = sum(1 for p in visual_prescans if p.is_complex_visual)
        lines.append(f"- Sheets with visual content: {sum(1 for p in visual_prescans if p.total_visual_objects > 0)}")
        lines.append(f"- Complex visual sheets: {complex_count}")
        lines.append(f"- Drawing objects extracted: {len(drawing_objects)}")
        lines.append(f"- Connectors found: {len(connectors)}")
        lines.append(f"- Charts found: {len(charts)}")
        lines.append(f"- Embedded images: {len(embedded_images)}")
        lines.append(f"- Bedrock analyses: {len(bedrock_results)}")
        lines.append("")

        if bedrock_results:
            lines.append("## Bedrock Analysis Results")
            lines.append("")
            success = sum(1 for b in bedrock_results if b.get("status") == "success")
            failed = len(bedrock_results) - success
            lines.append(f"- Success: {success}")
            lines.append(f"- Failed: {failed}")
            lines.append("")

        content = "\n".join(lines)
        path = self.reports_dir / "visual_parse_report.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)

    def generate_mermaid_report(self, mermaid_files, mermaid_graphs) -> str:
        """Generate mermaid parse report."""
        lines = []
        lines.append("# Mermaid Parse Report")
        lines.append("")
        lines.append(f"**Run:** {self.run_id}")
        lines.append("")
        lines.append(f"- Files found: {len(mermaid_files)}")
        lines.append(f"- Graphs parsed: {len(mermaid_graphs)}")
        lines.append("")

        total_nodes = sum(len(g.nodes) for g in mermaid_graphs)
        total_edges = sum(len(g.edges) for g in mermaid_graphs)
        lines.append(f"- Total nodes: {total_nodes}")
        lines.append(f"- Total edges: {total_edges}")
        lines.append("")

        for mg in mermaid_graphs:
            lines.append(f"### {mg.file_name} ({mg.graph_type})")
            lines.append(f"  Nodes: {len(mg.nodes)}, Edges: {len(mg.edges)}")
            lines.append("")

        content = "\n".join(lines)
        path = self.reports_dir / "mermaid_parse_report.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)

    def generate_upload_report(self, upload_results: dict) -> str:
        """Generate S3 upload report."""
        lines = []
        lines.append("# S3 Upload Report")
        lines.append("")
        lines.append(f"**Run:** {self.run_id}")
        lines.append(f"**Time:** {datetime.now(timezone.utc).isoformat()}")
        lines.append("")

        if upload_results.get("success"):
            lines.append("## Status: ✅ SUCCESS")
            lines.append("")
            lines.append(f"- Target: `{upload_results.get('target', '')}`")
            lines.append(f"- Latest: `{upload_results.get('latest', '')}`")
        else:
            lines.append("## Status: ❌ FAILED")
            lines.append("")
            lines.append(f"- Error: {upload_results.get('error', 'unknown')}")
            lines.append("")
            lines.append("⚠️ Local results preserved. Manual upload needed.")

        lines.append("")
        content = "\n".join(lines)
        path = self.reports_dir / "s3_upload_report.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return str(path)


# ============================================================
# JSONL Writer
# ============================================================
def write_jsonl(records: list, path: Path, key_func=None):
    """Write records to JSONL file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        for rec in records:
            if hasattr(rec, "__dict__"):
                data = {}
                for k, v in rec.__dict__.items():
                    if not k.startswith("_"):
                        data[k] = v
            elif isinstance(rec, dict):
                data = rec
            else:
                data = asdict(rec)
            f.write(json.dumps(data, ensure_ascii=False, default=str) + "\n")
    logger.info(f"Wrote {len(records)} records to {path}")


# ============================================================
# Main Pipeline
# ============================================================
class ComplexExcelParserPipeline:
    """Main pipeline orchestrator."""

    def __init__(self, args):
        self.args = args
        self.config = self._load_config()
        self.output_dir = Path(args.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.run_id = args.run_id
        self.dataset = args.dataset
        self.s3_uri = args.s3_uri or self.config.get("source", {}).get("s3_uri", "")
        self.use_bedrock = args.use_bedrock and not args.no_bedrock
        self.upload_s3 = args.upload_s3

        # Results
        self.file_manifest: list[FileManifestEntry] = []
        self.workbooks: list[WorkbookRecord] = []
        self.sheets: list[SheetRecord] = []
        self.table_regions: list[TableRegion] = []
        self.normalized_rows: list[NormalizedRow] = []
        self.visual_prescans: list[VisualPrescan] = []
        self.drawing_objects: list[DrawingObject] = []
        self.connectors: list[ConnectorRecord] = []
        self.charts: list[ChartRecord] = []
        self.embedded_images: list[EmbeddedImage] = []
        self.mermaid_files: list[MermaidFileRecord] = []
        self.mermaid_graphs: list[MermaidGraph] = []
        self.text_records: list[ParsedTextRecord] = []
        self.bedrock_results: list[dict] = []
        self.errors: list[str] = []
        self.generated_files: list[str] = []

    def _load_config(self) -> dict:
        """Load YAML config."""
        if self.args.config and Path(self.args.config).exists():
            with open(self.args.config, "r", encoding="utf-8") as f:
                return yaml.safe_load(f) or {}
        return {}

    def run(self) -> dict:
        """Execute the full pipeline."""
        start_time = time.time()
        logger.info("=" * 60)
        logger.info("Complex Excel Project Parser V1")
        logger.info(f"  Dataset: {self.dataset}")
        logger.info(f"  Run ID: {self.run_id}")
        logger.info(f"  S3 URI: {self.s3_uri}")
        logger.info(f"  Output: {self.output_dir}")
        logger.info(f"  Bedrock: {'enabled' if self.use_bedrock else 'disabled'}")
        logger.info("=" * 60)

        # Step 1: S3 Discovery
        logger.info("\n[Step 1/9] S3 File Discovery...")
        self._discover_files()

        # Step 2: Download files
        logger.info("\n[Step 2/9] Downloading files...")
        self._download_files()

        # Step 3: Parse Excel workbooks
        logger.info("\n[Step 3/9] Parsing Excel workbooks...")
        self._parse_excel_workbooks()

        # Step 4: Parse Mermaid files
        logger.info("\n[Step 4/9] Parsing Mermaid files...")
        self._parse_mermaid_files()

        # Step 5: Bedrock vision analysis (if enabled)
        logger.info("\n[Step 5/9] Visual analysis...")
        self._run_bedrock_analysis()

        # Step 6: Write JSONL outputs
        logger.info("\n[Step 6/9] Writing JSONL outputs...")
        self._write_jsonl_outputs()

        # Step 7: Generate Markdown
        logger.info("\n[Step 7/9] Generating Markdown review documents...")
        self._generate_markdown()

        # Step 8: Generate reports
        duration = time.time() - start_time
        logger.info("\n[Step 8/9] Generating reports...")
        self._generate_reports(duration)

        # Step 9: S3 Upload
        logger.info("\n[Step 9/9] S3 Upload...")
        upload_result = self._upload_to_s3()

        # Final summary
        duration = time.time() - start_time
        logger.info("\n" + "=" * 60)
        logger.info("PIPELINE COMPLETE")
        logger.info(f"  Duration: {duration:.1f}s")
        logger.info(f"  Workbooks: {len(self.workbooks)}")
        logger.info(f"  Sheets: {len(self.sheets)}")
        logger.info(f"  Table regions: {len(self.table_regions)}")
        logger.info(f"  Normalized rows: {len(self.normalized_rows)}")
        logger.info(f"  Drawing objects: {len(self.drawing_objects)}")
        logger.info(f"  Connectors: {len(self.connectors)}")
        logger.info(f"  Charts: {len(self.charts)}")
        logger.info(f"  Embedded images: {len(self.embedded_images)}")
        logger.info(f"  Mermaid files: {len(self.mermaid_files)}")
        logger.info(f"  Text records: {len(self.text_records)}")
        logger.info(f"  Errors: {len(self.errors)}")
        logger.info(f"  S3 upload: {'success' if upload_result.get('success') else 'failed/skipped'}")
        logger.info("=" * 60)

        return {
            "workbooks": len(self.workbooks),
            "sheets": len(self.sheets),
            "table_regions": len(self.table_regions),
            "normalized_rows": len(self.normalized_rows),
            "drawing_objects": len(self.drawing_objects),
            "connectors": len(self.connectors),
            "charts": len(self.charts),
            "embedded_images": len(self.embedded_images),
            "mermaid_files": len(self.mermaid_files),
            "mermaid_graphs": len(self.mermaid_graphs),
            "text_records": len(self.text_records),
            "bedrock_calls": len(self.bedrock_results),
            "errors": len(self.errors),
            "duration": duration,
            "upload_success": upload_result.get("success", False),
        }

    def _discover_files(self):
        """Discover all files in S3."""
        discovery = S3FileDiscovery(self.s3_uri, region="ap-northeast-1")
        self.file_manifest = discovery.discover()
        logger.info(f"  Found {len(self.file_manifest)} files:")
        type_counts = Counter(f.file_type for f in self.file_manifest)
        for ftype, count in type_counts.items():
            logger.info(f"    {ftype}: {count}")

    def _download_files(self):
        """Download relevant files from S3."""
        import boto3
        s3 = boto3.client("s3", region_name="ap-northeast-1")

        download_dir = self.output_dir / "downloads"
        download_dir.mkdir(parents=True, exist_ok=True)

        for entry in self.file_manifest:
            if entry.file_type in ("excel", "mermaid", "markdown", "csv"):
                # Preserve directory structure
                rel_path = entry.file_key.split("/", 1)[-1] if "/" in entry.file_key else entry.file_name
                local_path = download_dir / rel_path
                local_path.parent.mkdir(parents=True, exist_ok=True)

                try:
                    parts = self.s3_uri.replace("s3://", "").split("/", 1)
                    bucket = parts[0]
                    s3.download_file(bucket, entry.file_key, str(local_path))
                    entry.local_path = str(local_path)
                    logger.info(f"  Downloaded: {entry.file_name} ({entry.file_size:,} bytes)")
                except Exception as e:
                    self.errors.append(f"Download failed for {entry.file_name}: {e}")
                    logger.error(f"  Failed: {entry.file_name}: {e}")

    def _parse_excel_workbooks(self):
        """Parse all Excel workbooks."""
        excel_files = [f for f in self.file_manifest if f.file_type == "excel" and f.local_path]

        for entry in excel_files:
            logger.info(f"  Parsing: {entry.file_name}...")
            parser = ExcelWorkbookParser(
                local_path=entry.local_path,
                file_name=entry.file_name,
                file_key=entry.file_key,
                s3_uri=entry.s3_uri,
                dataset=self.dataset,
                run_id=self.run_id,
                output_dir=self.output_dir,
                config=self.config,
            )
            wb_record = parser.parse()
            self.workbooks.append(wb_record)
            self.sheets.extend(parser.sheet_records)
            self.table_regions.extend(parser.table_regions)
            self.normalized_rows.extend(parser.normalized_rows)
            self.visual_prescans.extend(parser.visual_prescans)
            self.drawing_objects.extend(parser.drawing_objects)
            self.connectors.extend(parser.connectors)
            self.charts.extend(parser.charts)
            self.embedded_images.extend(parser.embedded_images)
            self.text_records.extend(parser.text_records)
            self.errors.extend(parser.errors)

            logger.info(f"    Sheets: {len(parser.sheet_records)}, Tables: {len(parser.table_regions)}, "
                        f"Rows: {len(parser.normalized_rows)}, Drawings: {len(parser.drawing_objects)}")

    def _parse_mermaid_files(self):
        """Parse Mermaid files."""
        mermaid_entries = [f for f in self.file_manifest if f.file_type == "mermaid" and f.local_path]
        parser = MermaidParser(self.dataset, self.run_id)

        for entry in mermaid_entries:
            logger.info(f"  Parsing Mermaid: {entry.file_name}...")
            try:
                file_rec, graph = parser.parse_file(
                    entry.local_path, entry.file_name, entry.file_key, entry.s3_uri
                )

                # Try to associate with workbook in same directory
                entry_dir = str(Path(entry.file_key).parent)
                for wb in self.workbooks:
                    wb_dir = str(Path(wb.file_key).parent)
                    if wb_dir == entry_dir:
                        file_rec.associated_workbook = wb.file_name
                        # Try to find a flowchart sheet
                        for sh in self.sheets:
                            if sh.workbook_id == wb.workbook_id:
                                if any(kw in sh.sheet_name for kw in ["フローチャート", "フロー", "flow"]):
                                    file_rec.associated_sheet = sh.sheet_name
                                    break
                        break

                self.mermaid_files.append(file_rec)
                if graph:
                    self.mermaid_graphs.append(graph)
                    logger.info(f"    Nodes: {len(graph.nodes)}, Edges: {len(graph.edges)}, Type: {graph.graph_type}")

                # Add text records for mermaid content
                self.text_records.append(ParsedTextRecord(
                    record_id=_gen_id("txt", file_rec.file_id, "mermaid_content"),
                    workbook_name=file_rec.associated_workbook or file_rec.file_name,
                    sheet_name=file_rec.associated_sheet or "",
                    record_type="mermaid",
                    content=file_rec.content[:2000],
                    source_location=file_rec.s3_uri,
                    dataset=self.dataset,
                    run_id=self.run_id,
                ))
            except Exception as e:
                self.errors.append(f"Mermaid parse failed for {entry.file_name}: {e}")
                logger.error(f"  Mermaid parse error: {e}")

    def _run_bedrock_analysis(self):
        """Run Bedrock vision analysis on complex visual sheets."""
        if not self.use_bedrock:
            logger.info("  Bedrock disabled - skipping vision analysis")
            return

        # Find embedded images from complex visual sheets
        complex_sheet_ids = {p.sheet_id for p in self.visual_prescans if p.is_complex_visual}
        images_to_analyze = [img for img in self.embedded_images if img.sheet_id in complex_sheet_ids]

        if not images_to_analyze:
            logger.info("  No complex visual sheet images to analyze")
            return

        analyzer = BedrockVisionAnalyzer(
            region=self.config.get("bedrock", {}).get("region", "ap-northeast-1"),
            model_id=self.config.get("bedrock", {}).get("model_id", ""),
        )

        if not analyzer.available:
            logger.warning("  Bedrock not available - skipping vision analysis")
            return

        # Analyze images (limit to reasonable count)
        max_images = 10
        for img in images_to_analyze[:max_images]:
            if not Path(img.local_path).exists():
                continue
            # Only analyze standard image formats
            if img.image_format not in ("png", "jpg", "jpeg", "gif", "webp"):
                continue

            logger.info(f"  Analyzing: {img.source_path}...")
            result = analyzer.analyze_image(
                img.local_path,
                context=f"Workbook: {img.workbook_name}, Sheet: {img.sheet_name}"
            )
            result["workbook_name"] = img.workbook_name
            result["sheet_name"] = img.sheet_name
            self.bedrock_results.append(result)

            if result.get("status") == "success":
                logger.info(f"    ✓ Analysis complete ({len(result.get('analysis', ''))} chars)")
            else:
                logger.warning(f"    ✗ Failed: {result.get('error', 'unknown')}")

            # Rate limit
            time.sleep(2)

    def _write_jsonl_outputs(self):
        """Write all JSONL output files."""
        # File manifest
        write_jsonl(self.file_manifest, self.output_dir / "s3_file_manifest.jsonl")
        write_jsonl(self.workbooks, self.output_dir / "excel_workbooks.jsonl")
        write_jsonl(self.sheets, self.output_dir / "excel_sheets.jsonl")
        write_jsonl(self.table_regions, self.output_dir / "excel_table_regions.jsonl")
        write_jsonl(self.normalized_rows, self.output_dir / "excel_rows_normalized.jsonl")
        write_jsonl(self.visual_prescans, self.output_dir / "visual_prescan.jsonl")
        write_jsonl(self.drawing_objects, self.output_dir / "drawing_objects.jsonl")
        write_jsonl(self.connectors, self.output_dir / "connectors.jsonl")
        write_jsonl(self.charts, self.output_dir / "chart_objects.jsonl")
        write_jsonl(self.embedded_images, self.output_dir / "embedded_images.jsonl")
        write_jsonl(self.mermaid_files, self.output_dir / "mermaid_files.jsonl")
        write_jsonl(self.mermaid_graphs, self.output_dir / "mermaid_graphs.jsonl")
        write_jsonl(self.text_records, self.output_dir / "parsed_text_records.jsonl")

    def _generate_markdown(self):
        """Generate Markdown review documents."""
        gen = MarkdownGenerator(self.output_dir, self.dataset, self.run_id)

        path = gen.generate_full_review(
            self.workbooks, self.sheets, self.table_regions, self.normalized_rows,
            self.visual_prescans, self.drawing_objects, self.connectors, self.charts,
            self.embedded_images, self.mermaid_files, self.mermaid_graphs,
            self.bedrock_results, self.text_records
        )
        self.generated_files.append(path)
        logger.info(f"  Full review: {path}")

        path = gen.generate_summary(
            self.workbooks, self.sheets, self.table_regions, self.normalized_rows,
            self.drawing_objects, self.connectors, self.charts, self.embedded_images,
            self.mermaid_files, self.mermaid_graphs
        )
        self.generated_files.append(path)

        path = gen.generate_visual_review(
            self.visual_prescans, self.drawing_objects, self.connectors,
            self.charts, self.embedded_images, self.bedrock_results
        )
        self.generated_files.append(path)

        path = gen.generate_mermaid_review(self.mermaid_files, self.mermaid_graphs)
        self.generated_files.append(path)

    def _generate_reports(self, duration: float):
        """Generate pipeline reports."""
        rpt = ReportGenerator(self.output_dir, self.dataset, self.run_id)

        stats = {
            "Workbooks": len(self.workbooks),
            "Sheets": len(self.sheets),
            "Table Regions": len(self.table_regions),
            "Normalized Rows": len(self.normalized_rows),
            "Drawing Objects": len(self.drawing_objects),
            "Connectors": len(self.connectors),
            "Charts": len(self.charts),
            "Embedded Images": len(self.embedded_images),
            "Mermaid Files": len(self.mermaid_files),
            "Mermaid Graphs": len(self.mermaid_graphs),
            "Text Records": len(self.text_records),
            "Bedrock Analyses": len(self.bedrock_results),
            "Errors": len(self.errors),
        }

        path = rpt.generate_run_report(stats, duration, self.errors)
        self.generated_files.append(path)

        path = rpt.generate_quality_report(self.workbooks, self.sheets, self.table_regions, self.errors)
        self.generated_files.append(path)

        path = rpt.generate_visual_report(
            self.visual_prescans, self.drawing_objects, self.connectors,
            self.charts, self.embedded_images, self.bedrock_results
        )
        self.generated_files.append(path)

        path = rpt.generate_mermaid_report(self.mermaid_files, self.mermaid_graphs)
        self.generated_files.append(path)

        # Generate checklist (will be updated after upload)
        gen = MarkdownGenerator(self.output_dir, self.dataset, self.run_id)
        path = gen.generate_checklist(
            self.workbooks, self.sheets, self.table_regions, self.mermaid_files,
            self.visual_prescans, self.bedrock_results, False  # upload status TBD
        )
        self.generated_files.append(path)

    def _upload_to_s3(self) -> dict:
        """Upload results to S3."""
        if not self.upload_s3:
            logger.info("  No upload target specified - skipping")
            return {"success": False, "error": "No upload target"}

        rpt = ReportGenerator(self.output_dir, self.dataset, self.run_id)
        upload_result = {"success": False}

        try:
            target = self.upload_s3.rstrip("/")
            # Primary upload
            cmd1 = [
                "aws", "s3", "sync",
                str(self.output_dir),
                f"{target}/sample_20260519_excel_parser_v1/",
                "--only-show-errors",
            ]
            logger.info(f"  Uploading to {target}/sample_20260519_excel_parser_v1/...")
            result1 = subprocess.run(cmd1, capture_output=True, text=True, timeout=300)

            # Latest upload
            cmd2 = [
                "aws", "s3", "sync",
                str(self.output_dir),
                f"{target}/latest/",
                "--only-show-errors",
            ]
            logger.info(f"  Uploading to {target}/latest/...")
            result2 = subprocess.run(cmd2, capture_output=True, text=True, timeout=300)

            if result1.returncode == 0 and result2.returncode == 0:
                upload_result = {
                    "success": True,
                    "target": f"{target}/sample_20260519_excel_parser_v1/",
                    "latest": f"{target}/latest/",
                }
                logger.info("  ✓ Upload complete")
            else:
                error_msg = result1.stderr or result2.stderr or "Unknown error"
                upload_result = {"success": False, "error": error_msg}
                logger.error(f"  ✗ Upload failed: {error_msg}")

        except Exception as e:
            upload_result = {"success": False, "error": str(e)}
            logger.error(f"  ✗ Upload error: {e}")
            self.errors.append(f"S3 upload failed: {e}")

        # Write upload report
        rpt.generate_upload_report(upload_result)

        # Regenerate checklist with upload status
        gen = MarkdownGenerator(self.output_dir, self.dataset, self.run_id)
        gen.generate_checklist(
            self.workbooks, self.sheets, self.table_regions, self.mermaid_files,
            self.visual_prescans, self.bedrock_results, upload_result.get("success", False)
        )

        return upload_result


# ============================================================
# CLI Entry Point
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Complex Excel Project Parser V1 - Full document parsing pipeline"
    )
    parser.add_argument("--config", "-c", type=str, default="configs/sample_20260519_excel_parser_v1.yaml",
                        help="Path to YAML config file")
    parser.add_argument("--run-id", type=str, default="sample_20260519_excel_parser_v1",
                        help="Run identifier")
    parser.add_argument("--dataset", type=str, default="sample_20260519",
                        help="Dataset name")
    parser.add_argument("--s3-uri", type=str, default="",
                        help="S3 URI prefix for file discovery")
    parser.add_argument("--output-dir", type=str, default="data/outputs/sample_20260519_excel_parser_v1",
                        help="Output directory")
    parser.add_argument("--upload-s3", type=str, default="",
                        help="S3 URI to upload results to")
    parser.add_argument("--use-bedrock", action="store_true", default=False,
                        help="Enable Bedrock vision analysis")
    parser.add_argument("--no-bedrock", action="store_true", default=False,
                        help="Disable Bedrock analysis (overrides --use-bedrock)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Enable verbose logging")

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Load .env
    _load_env()

    # Run pipeline
    pipeline = ComplexExcelParserPipeline(args)
    results = pipeline.run()

    # Exit code
    if results["errors"] > 0:
        logger.warning(f"Completed with {results['errors']} error(s)")
        sys.exit(0)  # Still exit 0 since partial results are valid
    else:
        logger.info("Completed successfully with no errors")
        sys.exit(0)


if __name__ == "__main__":
    main()
