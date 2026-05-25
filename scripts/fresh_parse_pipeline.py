#!/usr/bin/env python3
"""
Fresh Excel Parse Pipeline - VLM-enhanced parsing of Japanese enterprise Excel design documents.
Combines structural extraction + screenshot rendering + Claude VLM analysis.
"""

import os
import sys
import json
import time
import re
import math
import traceback
from pathlib import Path
from datetime import datetime
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import boto3
from botocore.config import Config

# Configuration
OUTPUT_DIR = Path(os.path.expanduser("~/projects/hermes_bedrock_agent/outputs/fresh_parse_20260519"))
DOWNLOADS_DIR = Path("/tmp/s3_downloads/サンプル20260519")
BEDROCK_MODEL = "jp.anthropic.claude-sonnet-4-6"
BEDROCK_REGION = "ap-northeast-1"

# Image rendering settings
CELL_WIDTH_PX = 80  # pixels per column
CELL_HEIGHT_PX = 20  # pixels per row
MAX_TILE_WIDTH = 4000  # max tile width in pixels
MAX_TILE_HEIGHT = 4000  # max tile height in pixels
MAX_IMAGE_BYTES = 4_500_000  # ~4.5MB max for Bedrock API
FONT_PATH = "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf"
FONT_SIZE = 11

# Bedrock client
bedrock_config = Config(
    region_name=BEDROCK_REGION,
    read_timeout=600,
    retries={'max_attempts': 3}
)
bedrock_client = boto3.client('bedrock-runtime', config=bedrock_config)


def safe_filename(name: str) -> str:
    """Convert sheet name to filesystem-safe name."""
    # Replace problematic chars
    safe = re.sub(r'[\\/:*?"<>|]', '_', name)
    safe = re.sub(r'[\s]+', '_', safe)
    safe = safe.strip('_')
    if len(safe) > 60:
        safe = safe[:60]
    return safe


def classify_sheet_heuristic(ws_name: str, max_row: int, max_col: int, cell_sample: list) -> dict:
    """Heuristically classify a sheet based on name, size, and sample cells."""
    classification = {
        "sheet_type": "unknown",
        "has_flowchart": False,
        "has_mapping": False,
        "has_tables": False,
        "is_overview": False,
        "is_wide": max_col > 50,
        "is_tall": max_row > 100,
        "needs_tiling": False,
        "confidence": 0.5
    }
    
    name_lower = ws_name.lower()
    
    # Detect by name patterns
    if any(k in ws_name for k in ['概要', '変更履歴', '補足', '目次', 'TOC']):
        classification["sheet_type"] = "overview"
        classification["is_overview"] = True
        classification["confidence"] = 0.8
    elif any(k in ws_name for k in ['フローチャート', 'フロー', 'flow']):
        classification["sheet_type"] = "flowchart"
        classification["has_flowchart"] = True
        classification["confidence"] = 0.8
    elif any(k in ws_name for k in ['マッピング', 'mapping', 'マッピングシート']):
        classification["sheet_type"] = "mapping"
        classification["has_mapping"] = True
        classification["confidence"] = 0.9
    elif any(k in ws_name for k in ['データ取得条件', '条件']):
        classification["sheet_type"] = "condition_table"
        classification["has_tables"] = True
        classification["confidence"] = 0.8
    elif any(k in ws_name for k in ['API', 'DataSpider', '開発仕様']):
        classification["sheet_type"] = "specification"
        classification["has_tables"] = True
        classification["confidence"] = 0.7
    
    # Size-based classification adjustments
    if max_col > 100:
        classification["has_mapping"] = True
        classification["is_wide"] = True
        classification["needs_tiling"] = True
    
    if max_row > 50 or max_col > 30:
        classification["needs_tiling"] = True
    
    return classification


def extract_cell_data(ws, max_row: int, max_col: int, sample_rows: int = 30) -> list:
    """Extract cell data from worksheet for structural analysis."""
    cells = []
    row_limit = min(max_row, sample_rows)
    col_limit = min(max_col, 50)  # Cap columns for sampling
    
    for row in range(1, row_limit + 1):
        for col in range(1, col_limit + 1):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cells.append({
                    "row": row,
                    "col": col,
                    "value": str(cell.value)[:100],
                    "col_letter": get_column_letter(col)
                })
    return cells


def render_sheet_image(ws, max_row: int, max_col: int, 
                        start_row: int = 1, end_row: int = None,
                        start_col: int = 1, end_col: int = None) -> Image.Image:
    """Render a portion of an Excel sheet to a PIL Image."""
    if end_row is None:
        end_row = max_row
    if end_col is None:
        end_col = max_col
    
    # Limit render size
    end_row = min(end_row, start_row + 200)  # Max 200 rows per tile
    end_col = min(end_col, start_col + 60)   # Max 60 cols per tile
    
    num_rows = end_row - start_row + 1
    num_cols = end_col - start_col + 1
    
    img_width = num_cols * CELL_WIDTH_PX
    img_height = num_rows * CELL_HEIGHT_PX
    
    # Cap image size
    img_width = min(img_width, MAX_TILE_WIDTH)
    img_height = min(img_height, MAX_TILE_HEIGHT)
    
    img = Image.new('RGB', (img_width, img_height), 'white')
    draw = ImageDraw.Draw(img)
    
    try:
        font = ImageFont.truetype(FONT_PATH, FONT_SIZE)
    except:
        font = ImageFont.load_default()
    
    # Draw grid and cell values
    for row_idx in range(num_rows):
        actual_row = start_row + row_idx
        y = row_idx * CELL_HEIGHT_PX
        
        # Draw horizontal line
        draw.line([(0, y), (img_width, y)], fill='#CCCCCC', width=1)
        
        for col_idx in range(num_cols):
            actual_col = start_col + col_idx
            x = col_idx * CELL_WIDTH_PX
            
            # Draw vertical line
            if row_idx == 0:
                draw.line([(x, 0), (x, img_height)], fill='#CCCCCC', width=1)
            
            # Get cell value
            try:
                cell = ws.cell(actual_row, actual_col)
                value = cell.value
                if value is not None:
                    text = str(value)[:12]  # Truncate for display
                    # Check if header (bold or first few rows)
                    is_header = actual_row <= 3 or (cell.font and cell.font.bold)
                    color = '#000066' if is_header else '#333333'
                    draw.text((x + 2, y + 3), text, fill=color, font=font)
            except Exception:
                pass
    
    # Draw final borders
    draw.line([(0, img_height - 1), (img_width - 1, img_height - 1)], fill='#CCCCCC', width=1)
    draw.line([(img_width - 1, 0), (img_width - 1, img_height - 1)], fill='#CCCCCC', width=1)
    
    return img


def generate_tiles(ws, max_row: int, max_col: int) -> list:
    """Generate image tiles for a sheet, splitting if needed."""
    tiles = []
    
    # Calculate how many tiles we need
    cols_per_tile = 50  # ~50 columns per tile
    rows_per_tile = 150  # ~150 rows per tile
    
    num_col_tiles = math.ceil(max_col / cols_per_tile)
    num_row_tiles = math.ceil(max_row / rows_per_tile)
    
    for row_tile in range(num_row_tiles):
        for col_tile in range(num_col_tiles):
            start_row = row_tile * rows_per_tile + 1
            end_row = min((row_tile + 1) * rows_per_tile, max_row)
            start_col = col_tile * cols_per_tile + 1
            end_col = min((col_tile + 1) * cols_per_tile, max_col)
            
            tile_img = render_sheet_image(ws, max_row, max_col, 
                                          start_row, end_row, start_col, end_col)
            tiles.append({
                "image": tile_img,
                "start_row": start_row,
                "end_row": end_row,
                "start_col": start_col,
                "end_col": end_col,
                "row_tile": row_tile,
                "col_tile": col_tile
            })
    
    return tiles


def stitch_tiles(tiles: list, max_row: int, max_col: int) -> Image.Image:
    """Stitch tiles back into one full-sheet image."""
    if len(tiles) == 1:
        return tiles[0]["image"]
    
    # Calculate tile grid dimensions
    col_tiles = max(t["col_tile"] for t in tiles) + 1
    row_tiles = max(t["row_tile"] for t in tiles) + 1
    
    # Calculate total image size
    # Get width/height of each tile position
    widths = {}
    heights = {}
    for t in tiles:
        widths[t["col_tile"]] = max(widths.get(t["col_tile"], 0), t["image"].width)
        heights[t["row_tile"]] = max(heights.get(t["row_tile"], 0), t["image"].height)
    
    total_width = sum(widths.get(c, 0) for c in range(col_tiles))
    total_height = sum(heights.get(r, 0) for r in range(row_tiles))
    
    # Cap stitched image
    total_width = min(total_width, 12000)
    total_height = min(total_height, 8000)
    
    stitched = Image.new('RGB', (total_width, total_height), 'white')
    
    for t in tiles:
        x_offset = sum(widths.get(c, 0) for c in range(t["col_tile"]))
        y_offset = sum(heights.get(r, 0) for r in range(t["row_tile"]))
        stitched.paste(t["image"], (x_offset, y_offset))
    
    return stitched


def image_to_bytes(img: Image.Image, max_bytes: int = MAX_IMAGE_BYTES, max_dim: int = 7900) -> bytes:
    """Convert PIL Image to PNG bytes, resizing if necessary to meet size and dimension limits."""
    # First enforce max dimension (Bedrock limit is 8000px)
    if img.width > max_dim or img.height > max_dim:
        scale = min(max_dim / img.width, max_dim / img.height)
        new_width = int(img.width * scale)
        new_height = int(img.height * scale)
        img = img.resize((new_width, new_height), Image.LANCZOS)
    
    buf = BytesIO()
    img.save(buf, format='PNG')
    data = buf.getvalue()
    
    # If too large in bytes, resize further
    while len(data) > max_bytes and img.width > 400:
        # Reduce by 30%
        new_width = int(img.width * 0.7)
        new_height = int(img.height * 0.7)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        buf = BytesIO()
        img.save(buf, format='PNG')
        data = buf.getvalue()
    
    return data


def call_vlm(image_bytes: bytes, prompt: str, max_retries: int = 3) -> str:
    """Call Bedrock Claude Sonnet VLM with an image."""
    for attempt in range(max_retries):
        try:
            response = bedrock_client.converse(
                modelId=BEDROCK_MODEL,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "image": {
                                "format": "png",
                                "source": {"bytes": image_bytes}
                            }
                        },
                        {
                            "text": prompt
                        }
                    ]
                }],
                inferenceConfig={
                    "maxTokens": 8000,
                    "temperature": 0.1
                }
            )
            return response['output']['message']['content'][0]['text']
        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = 5 * (attempt + 1)
                print(f"    VLM call failed (attempt {attempt+1}): {e}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                return f"[VLM ERROR after {max_retries} attempts: {str(e)}]"


def call_text_llm(prompt: str, max_retries: int = 3) -> str:
    """Call Bedrock Claude Sonnet for text-only analysis."""
    for attempt in range(max_retries):
        try:
            response = bedrock_client.converse(
                modelId=BEDROCK_MODEL,
                messages=[{
                    "role": "user",
                    "content": [{"text": prompt}]
                }],
                inferenceConfig={
                    "maxTokens": 8000,
                    "temperature": 0.1
                }
            )
            return response['output']['message']['content'][0]['text']
        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = 5 * (attempt + 1)
                print(f"    LLM call failed (attempt {attempt+1}): {e}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                return f"[LLM ERROR after {max_retries} attempts: {str(e)}]"


def extract_structural_data(ws, max_row: int, max_col: int) -> dict:
    """Extract comprehensive structural data from a worksheet."""
    # Get all non-empty cells (capped for very large sheets)
    cells = []
    row_limit = min(max_row, 100)
    col_limit = min(max_col, 200)
    
    for row in range(1, row_limit + 1):
        for col in range(1, col_limit + 1):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cells.append({
                    "row": row,
                    "col": col,
                    "col_letter": get_column_letter(col),
                    "value": str(cell.value)[:200],
                    "is_bold": bool(cell.font and cell.font.bold),
                    "has_border": bool(cell.border and (cell.border.top or cell.border.bottom)),
                })
    
    # Detect merged cells
    merges = []
    for mc in ws.merged_cells.ranges:
        merges.append(str(mc))
    
    return {
        "cell_count": len(cells),
        "cells": cells[:500],  # Cap for LLM context
        "merged_ranges": merges[:50],
        "max_row": max_row,
        "max_col": max_col
    }


def build_vlm_prompt(sheet_name: str, classification: dict, structural_summary: str) -> str:
    """Build a VLM prompt for sheet analysis."""
    prompt = f"""You are analyzing a screenshot of an Excel sheet from a Japanese enterprise design document.

Sheet name: {sheet_name}
Detected type: {classification['sheet_type']}
Sheet size: {classification.get('max_row', '?')} rows x {classification.get('max_col', '?')} columns

Structural context (from Excel cell extraction):
{structural_summary}

Please analyze this image and extract the following information in structured Markdown format:

## 1. Sheet Overview
Describe the purpose of this sheet.

## 2. Visual Structure
Describe the main visual layout.

## 3. Extracted Tables
Extract important tables as Markdown tables (preserve Japanese text exactly).

## 4. Field Mapping / Data Mapping
If this contains mapping information, extract it:
| Source System | Source Table | Source Field | Target System | Target Table | Target Field | Transformation Rule | Notes |
|---|---|---|---|---|---|---|---|

## 5. Business Rules
List business rules, conditions, filters, validation rules.

## 6. Flowchart Explanation
If there's a flowchart/process, explain the flow.

## 7. Mermaid Diagram
If applicable, provide a valid Mermaid flowchart diagram (flowchart TD or LR).

## 8. Uncertain or Ambiguous Points
List any unclear or ambiguous areas.

Important:
- Preserve Japanese text as-is
- Do not invent data not visible in the image
- Mark uncertain items clearly
- If text is truncated in the image, note it"""
    
    return prompt


def process_sheet(wb_path: Path, ws_name: str, sheet_idx: int, 
                  output_base: Path, wb_name: str) -> dict:
    """Process a single worksheet: screenshots + structural extraction + VLM analysis."""
    
    safe_name = safe_filename(ws_name)
    sheet_dir = output_base / wb_name / "sheets" / f"{sheet_idx:02d}_{safe_name}"
    screenshots_dir = sheet_dir / "screenshots"
    vlm_dir = sheet_dir / "vlm"
    result_dir = sheet_dir / "result"
    
    for d in [screenshots_dir, vlm_dir, result_dir]:
        d.mkdir(parents=True, exist_ok=True)
    
    print(f"  Processing sheet [{sheet_idx}]: {ws_name}")
    
    # Open workbook (need read_only=False for proper cell access)
    wb = openpyxl.load_workbook(str(wb_path), data_only=True, read_only=False)
    ws = wb[ws_name]
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    
    # Step 1: Classify sheet
    cell_sample = extract_cell_data(ws, max_row, max_col, sample_rows=10)
    classification = classify_sheet_heuristic(ws_name, max_row, max_col, cell_sample)
    classification["max_row"] = max_row
    classification["max_col"] = max_col
    
    print(f"    Type: {classification['sheet_type']}, Size: {max_row}x{max_col}, Wide: {classification['is_wide']}")
    
    # Step 2: Extract structural data
    structural = extract_structural_data(ws, max_row, max_col)
    
    # Step 3: Generate screenshots
    tiles = generate_tiles(ws, max_row, max_col)
    tile_paths = []
    for i, tile in enumerate(tiles):
        tile_path = screenshots_dir / f"tile_{i+1:03d}.png"
        tile["image"].save(str(tile_path))
        tile_paths.append(tile_path)
    
    print(f"    Generated {len(tiles)} tile(s)")
    
    # Step 4: Stitch tiles
    stitched_img = stitch_tiles(tiles, max_row, max_col)
    stitched_path = screenshots_dir / "stitched_full_sheet.png"
    stitched_img.save(str(stitched_path))
    print(f"    Stitched image: {stitched_img.width}x{stitched_img.height}px")
    
    # Step 5: VLM Analysis
    # Prepare structural summary for context
    struct_lines = []
    for cell in structural["cells"][:100]:
        struct_lines.append(f"  [{cell['col_letter']}{cell['row']}] = {cell['value'][:60]}")
    structural_summary = "\n".join(struct_lines[:80])
    
    vlm_prompt = build_vlm_prompt(ws_name, classification, structural_summary)
    
    # Try to send stitched image to VLM (resize if needed)
    img_bytes = image_to_bytes(stitched_img)
    print(f"    Image for VLM: {len(img_bytes)/1024:.0f}KB")
    
    vlm_response = call_vlm(img_bytes, vlm_prompt)
    
    # Save VLM response
    vlm_response_path = vlm_dir / "vlm_raw_response.md"
    with open(vlm_response_path, 'w') as f:
        f.write(f"# VLM Response for: {ws_name}\n\n")
        f.write(f"Model: {BEDROCK_MODEL}\n")
        f.write(f"Image size: {stitched_img.width}x{stitched_img.height}px ({len(img_bytes)/1024:.0f}KB)\n\n")
        f.write(vlm_response)
    
    # If sheet is very wide and has multiple column tiles, also do per-tile VLM
    if len(tiles) > 2 and classification["is_wide"]:
        tile_responses = []
        for i, tile in enumerate(tiles[:6]):  # Limit to first 6 tiles
            tile_bytes = image_to_bytes(tile["image"])
            tile_prompt = f"""This is tile {i+1} of sheet '{ws_name}' (rows {tile['start_row']}-{tile['end_row']}, cols {get_column_letter(tile['start_col'])}-{get_column_letter(tile['end_col'])}).
Extract all visible text, table headers, field names, and values. Preserve Japanese text exactly."""
            tile_resp = call_vlm(tile_bytes, tile_prompt)
            tile_responses.append(f"### Tile {i+1} (rows {tile['start_row']}-{tile['end_row']}, cols {get_column_letter(tile['start_col'])}-{get_column_letter(tile['end_col'])})\n\n{tile_resp}")
            time.sleep(3)  # Rate limiting between VLM calls
        
        tile_response_path = vlm_dir / "vlm_tile_responses.md"
        with open(tile_response_path, 'w') as f:
            f.write(f"# VLM Tile Responses for: {ws_name}\n\n")
            f.write("\n\n---\n\n".join(tile_responses))
    
    # Step 6: Generate result files
    # Parse plan
    parse_plan = {
        "sheet_name": ws_name,
        "classification": classification,
        "strategy": "vlm_with_structural" if classification["needs_tiling"] else "vlm_single_image",
        "tiles_generated": len(tiles),
        "structural_cells_extracted": structural["cell_count"],
        "merged_ranges": len(structural["merged_ranges"])
    }
    
    parse_plan_path = sheet_dir / "sheet_parse_plan.md"
    with open(parse_plan_path, 'w') as f:
        f.write(f"# Parse Plan: {ws_name}\n\n")
        f.write(f"## Classification\n")
        f.write(f"- Sheet type: {classification['sheet_type']}\n")
        f.write(f"- Has flowchart: {classification['has_flowchart']}\n")
        f.write(f"- Has mapping: {classification['has_mapping']}\n")
        f.write(f"- Is wide: {classification['is_wide']}\n")
        f.write(f"- Is tall: {classification['is_tall']}\n")
        f.write(f"- Needs tiling: {classification['needs_tiling']}\n")
        f.write(f"- Confidence: {classification['confidence']}\n\n")
        f.write(f"## Strategy\n")
        f.write(f"- Primary: VLM visual analysis of stitched full-sheet image\n")
        f.write(f"- Secondary: Structural cell extraction from openpyxl\n")
        f.write(f"- Tiles: {len(tiles)}\n")
        f.write(f"- Image dimensions: {stitched_img.width}x{stitched_img.height}px\n")
    
    # Sheet result - combine VLM response with structural data
    result_path = result_dir / "sheet_result.md"
    with open(result_path, 'w') as f:
        f.write(f"# Sheet: {ws_name}\n\n")
        f.write(f"## 1. Sheet Overview\n\n")
        # Extract from VLM response if possible
        f.write(f"Sheet type: {classification['sheet_type']}\n")
        f.write(f"Dimensions: {max_row} rows × {max_col} columns\n\n")
        f.write(f"## 2. Parsing Strategy\n\n")
        f.write(f"- Screenshot tiles: {len(tiles)}\n")
        f.write(f"- VLM model: {BEDROCK_MODEL}\n")
        f.write(f"- Combined structural + visual analysis\n\n")
        f.write(f"## 3. Visual Structure\n\n")
        f.write(f"(See VLM analysis below)\n\n")
        f.write(f"## 4. Extracted Tables\n\n")
        f.write(f"(See VLM analysis below)\n\n")
        f.write(f"## 5. Field Mapping / Data Mapping\n\n")
        if classification["has_mapping"]:
            f.write(f"| Source System | Source Table | Source Field | Target System | Target Table | Target Field | Transformation Rule | Notes |\n")
            f.write(f"|---|---|---|---|---|---|---|---|\n")
            f.write(f"(See detailed VLM extraction)\n\n")
        else:
            f.write(f"Not applicable for this sheet type.\n\n")
        f.write(f"## 6. Business Rules\n\n")
        f.write(f"(See VLM analysis)\n\n")
        f.write(f"## 7. Flowchart Explanation\n\n")
        if classification["has_flowchart"]:
            f.write(f"(See VLM analysis and .mmd file)\n\n")
        else:
            f.write(f"No flowchart detected in this sheet.\n\n")
        f.write(f"## 8. Mermaid Diagram\n\n")
        if classification["has_flowchart"]:
            f.write(f"See sheet_flowchart.mmd\n\n")
        else:
            f.write(f"Not applicable.\n\n")
        f.write(f"## 9. Evidence Files\n\n")
        f.write(f"- Stitched image: screenshots/stitched_full_sheet.png\n")
        for tp in tile_paths:
            f.write(f"- Tile: screenshots/{tp.name}\n")
        f.write(f"- VLM response: vlm/vlm_raw_response.md\n\n")
        f.write(f"## 10. Uncertain or Ambiguous Points\n\n")
        f.write(f"(See VLM analysis uncertain section)\n\n")
        f.write(f"## 11. Manual Review Checklist\n\n")
        f.write(f"- [ ] Verify VLM extraction accuracy against stitched image\n")
        if classification["has_mapping"]:
            f.write(f"- [ ] Verify mapping field extraction completeness\n")
        if classification["has_flowchart"]:
            f.write(f"- [ ] Verify Mermaid diagram correctness\n")
        f.write(f"- [ ] Check for missed content in wide columns\n\n")
        f.write(f"---\n\n## VLM Analysis Result\n\n")
        f.write(vlm_response)
    
    # Uncertain points
    uncertain_path = result_dir / "uncertain_points.md"
    with open(uncertain_path, 'w') as f:
        f.write(f"# Uncertain Points: {ws_name}\n\n")
        if classification["is_wide"]:
            f.write(f"- Sheet has {max_col} columns; some content may be truncated in tile rendering\n")
        if max_row > 100:
            f.write(f"- Sheet has {max_row} rows; lower rows may have reduced rendering quality\n")
        f.write(f"- VLM confidence may be lower for small/dense text areas\n")
        if "[VLM ERROR" in vlm_response:
            f.write(f"- VLM call failed; results are incomplete\n")
    
    wb.close()
    
    return {
        "sheet_name": ws_name,
        "sheet_idx": sheet_idx,
        "safe_name": safe_name,
        "classification": classification,
        "tiles": len(tiles),
        "stitched_size": f"{stitched_img.width}x{stitched_img.height}",
        "vlm_success": "[VLM ERROR" not in vlm_response,
        "output_dir": str(sheet_dir)
    }


def process_workbook(wb_path: Path, output_base: Path) -> dict:
    """Process an entire workbook."""
    wb_filename = wb_path.name
    wb_safe_name = safe_filename(wb_path.stem)
    
    print(f"\n{'='*60}")
    print(f"Processing workbook: {wb_filename}")
    print(f"{'='*60}")
    
    wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    # Create workbook output dir
    wb_dir = output_base / wb_safe_name
    wb_dir.mkdir(parents=True, exist_ok=True)
    
    # Workbook summary
    summary = {
        "filename": wb_filename,
        "safe_name": wb_safe_name,
        "sheet_count": len(sheet_names),
        "sheets": [],
        "processed": 0,
        "failed": 0
    }
    
    # Process each sheet
    for idx, ws_name in enumerate(sheet_names):
        # Check if already processed (resume support)
        safe_name = safe_filename(ws_name)
        sheet_result_file = output_base / wb_safe_name / "sheets" / f"{idx:02d}_{safe_name}" / "vlm" / "vlm_raw_response.md"
        if sheet_result_file.exists():
            # Check if VLM was successful (not an error)
            content = sheet_result_file.read_text()
            if "[VLM ERROR" not in content:
                print(f"  Skipping sheet [{idx}] {ws_name} (already processed)")
                summary["sheets"].append({
                    "sheet_name": ws_name,
                    "sheet_idx": idx,
                    "safe_name": safe_name,
                    "classification": classify_sheet_heuristic(ws_name, 0, 0, []),
                    "tiles": 0,
                    "stitched_size": "cached",
                    "vlm_success": True,
                    "output_dir": str(output_base / wb_safe_name / "sheets" / f"{idx:02d}_{safe_name}"),
                    "cached": True
                })
                summary["processed"] += 1
                continue
        
        try:
            result = process_sheet(wb_path, ws_name, idx, output_base, wb_safe_name)
            summary["sheets"].append(result)
            summary["processed"] += 1
            time.sleep(3)  # Rate limiting between VLM calls
        except Exception as e:
            print(f"    ERROR processing sheet {ws_name}: {e}")
            traceback.print_exc()
            summary["sheets"].append({
                "sheet_name": ws_name,
                "sheet_idx": idx,
                "error": str(e)
            })
            summary["failed"] += 1
    
    # Write workbook summary
    wb_summary_path = wb_dir / "workbook_summary.md"
    with open(wb_summary_path, 'w') as f:
        f.write(f"# Workbook: {wb_filename}\n\n")
        f.write(f"## Overview\n")
        f.write(f"- Total sheets: {len(sheet_names)}\n")
        f.write(f"- Processed: {summary['processed']}\n")
        f.write(f"- Failed: {summary['failed']}\n\n")
        f.write(f"## Sheet List\n\n")
        f.write(f"| # | Sheet Name | Type | Size | Tiles | VLM OK |\n")
        f.write(f"|---|---|---|---|---|---|\n")
        for s in summary["sheets"]:
            if "error" in s:
                f.write(f"| {s.get('sheet_idx', '?')} | {s['sheet_name']} | ERROR | - | - | ❌ |\n")
            else:
                f.write(f"| {s['sheet_idx']} | {s['sheet_name']} | {s['classification']['sheet_type']} | {s['classification'].get('max_row', '?')}×{s['classification'].get('max_col', '?')} | {s['tiles']} | {'✅' if s['vlm_success'] else '❌'} |\n")
    
    return summary


def process_mermaid_file(mmd_path: Path, output_base: Path) -> dict:
    """Process a standalone Mermaid file."""
    with open(mmd_path, 'r') as f:
        content = f.read()
    
    # Copy to the related workbook's output
    wb_safe_name = safe_filename("M社様_DSSスクリプト改修概要_フローチャート")
    out_dir = output_base / wb_safe_name / "sheets" / "01_フローチャート" / "result"
    out_dir.mkdir(parents=True, exist_ok=True)
    
    out_path = out_dir / "sheet_flowchart.mmd"
    with open(out_path, 'w') as f:
        f.write(content)
    
    return {
        "source": str(mmd_path),
        "output": str(out_path),
        "lines": content.count('\n') + 1,
        "size": len(content)
    }


def generate_inventory(output_base: Path, workbook_results: list, mmd_result: dict):
    """Generate inventory.md listing all source files."""
    inv_path = output_base / "inventory.md"
    with open(inv_path, 'w') as f:
        f.write("# Source File Inventory\n\n")
        f.write(f"Generated: {datetime.now().isoformat()}\n\n")
        f.write("## S3 Source: s3://s3-hulftchina-rd/サンプル20260519/\n\n")
        f.write("| File | Type | Size | Sheets |\n")
        f.write("|---|---|---|---|\n")
        for wb in workbook_results:
            f.write(f"| {wb['filename']} | Excel | - | {wb['sheet_count']} |\n")
        if mmd_result:
            f.write(f"| flowchart.mmd | Mermaid | {mmd_result['size']} bytes | - |\n")


def generate_review_index(output_base: Path, workbook_results: list):
    """Generate manual_review_index.md."""
    idx_path = output_base / "manual_review_index.md"
    with open(idx_path, 'w') as f:
        f.write("# Manual Review Index\n\n")
        f.write(f"Generated: {datetime.now().isoformat()}\n\n")
        f.write("| Workbook | Sheet | Sheet Type | Screenshot File | Markdown Result | Mermaid File | Mapping File | Review Priority | Reason |\n")
        f.write("|---|---|---|---|---|---|---|---|---|\n")
        
        for wb in workbook_results:
            for s in wb["sheets"]:
                if "error" in s:
                    f.write(f"| {wb['filename']} | {s['sheet_name']} | ERROR | - | - | - | - | High | Processing error |\n")
                    continue
                
                cls = s["classification"]
                safe_name = s["safe_name"]
                wb_safe = wb["safe_name"]
                sheet_dir = f"{wb_safe}/sheets/{s['sheet_idx']:02d}_{safe_name}"
                
                screenshot = f"{sheet_dir}/screenshots/stitched_full_sheet.png"
                result_md = f"{sheet_dir}/result/sheet_result.md"
                mermaid = f"{sheet_dir}/result/sheet_flowchart.mmd" if cls["has_flowchart"] else "-"
                mapping = f"{sheet_dir}/result/extracted_mappings.md" if cls["has_mapping"] else "-"
                
                # Determine priority
                if cls["has_flowchart"] or (cls["is_wide"] and cls["has_mapping"]):
                    priority = "High"
                    reason = "Complex " + ("flowchart" if cls["has_flowchart"] else "wide mapping")
                elif cls["has_mapping"]:
                    priority = "Medium"
                    reason = "Mapping table"
                else:
                    priority = "Low"
                    reason = cls["sheet_type"]
                
                f.write(f"| {wb['filename'][:30]} | {s['sheet_name'][:30]} | {cls['sheet_type']} | {screenshot} | {result_md} | {mermaid} | {mapping} | {priority} | {reason} |\n")


def generate_run_summary(output_base: Path, workbook_results: list, mmd_result: dict, 
                          start_time: datetime, end_time: datetime):
    """Generate run_summary.md."""
    total_sheets = sum(wb["sheet_count"] for wb in workbook_results)
    total_processed = sum(wb["processed"] for wb in workbook_results)
    total_failed = sum(wb["failed"] for wb in workbook_results)
    total_tiles = sum(s.get("tiles", 0) for wb in workbook_results for s in wb["sheets"] if "error" not in s)
    
    high_risk = []
    for wb in workbook_results:
        for s in wb["sheets"]:
            if "error" in s:
                high_risk.append(f"{wb['filename']} / {s['sheet_name']} (ERROR)")
            elif s["classification"]["is_wide"] and s["classification"]["has_mapping"]:
                high_risk.append(f"{wb['filename']} / {s['sheet_name']} (wide mapping)")
            elif s["classification"]["has_flowchart"]:
                high_risk.append(f"{wb['filename']} / {s['sheet_name']} (flowchart)")
    
    summary_path = output_base / "run_summary.md"
    with open(summary_path, 'w') as f:
        f.write("# Run Summary\n\n")
        f.write(f"- Start time: {start_time.isoformat()}\n")
        f.write(f"- End time: {end_time.isoformat()}\n")
        f.write(f"- Duration: {(end_time - start_time).total_seconds():.0f} seconds\n\n")
        f.write("## Source Files\n\n")
        f.write(f"- S3 prefix: s3://s3-hulftchina-rd/サンプル20260519/\n")
        f.write(f"- Excel workbooks: {len(workbook_results)}\n")
        f.write(f"- Mermaid files: {'1' if mmd_result else '0'}\n")
        f.write(f"- Skipped files: 0\n\n")
        f.write("## Processing Results\n\n")
        f.write(f"- Total sheets: {total_sheets}\n")
        f.write(f"- Successfully processed: {total_processed}\n")
        f.write(f"- Failed: {total_failed}\n")
        f.write(f"- Total screenshot tiles: {total_tiles}\n")
        f.write(f"- Stitched images: {total_processed}\n")
        f.write(f"- Markdown result files: {total_processed}\n")
        f.write(f"- VLM calls made: ~{total_processed + total_tiles}\n\n")
        f.write("## High-Risk Sheets (Manual Review Required)\n\n")
        for hr in high_risk:
            f.write(f"- {hr}\n")
        f.write(f"\n## Known Limitations\n\n")
        f.write(f"- Excel shapes/connectors/arrows are not extractable via openpyxl (DrawingML limitation)\n")
        f.write(f"- Very wide sheets (100+ columns) may have text truncated in rendered tiles\n")
        f.write(f"- VLM analysis quality depends on image resolution and text density\n")
        f.write(f"- Japanese text rendering requires DroidSansFallbackFull font\n")
        f.write(f"- Cell values are truncated at 200 chars for structural extraction\n")


def main():
    """Main pipeline entry point."""
    start_time = datetime.now()
    print(f"Fresh Parse Pipeline starting at {start_time.isoformat()}", flush=True)
    print(f"Output directory: {OUTPUT_DIR}", flush=True)
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Find all Excel workbooks
    xlsx_files = list(DOWNLOADS_DIR.rglob("*.xlsx"))
    mmd_files = list(DOWNLOADS_DIR.rglob("*.mmd"))
    
    print(f"\nFound {len(xlsx_files)} Excel workbook(s) and {len(mmd_files)} Mermaid file(s)")
    
    workbook_results = []
    
    # Process each workbook
    for wb_path in xlsx_files:
        result = process_workbook(wb_path, OUTPUT_DIR)
        workbook_results.append(result)
    
    # Process Mermaid files
    mmd_result = None
    if mmd_files:
        mmd_result = process_mermaid_file(mmd_files[0], OUTPUT_DIR)
        print(f"\nProcessed Mermaid file: {mmd_files[0].name} ({mmd_result['lines']} lines)")
    
    end_time = datetime.now()
    
    # Generate summary files
    generate_inventory(OUTPUT_DIR, workbook_results, mmd_result)
    generate_review_index(OUTPUT_DIR, workbook_results)
    generate_run_summary(OUTPUT_DIR, workbook_results, mmd_result, start_time, end_time)
    
    print(f"\n{'='*60}")
    print(f"Pipeline complete!")
    print(f"Duration: {(end_time - start_time).total_seconds():.0f} seconds")
    print(f"Output: {OUTPUT_DIR}")
    print(f"{'='*60}")
    
    return workbook_results


if __name__ == "__main__":
    main()
