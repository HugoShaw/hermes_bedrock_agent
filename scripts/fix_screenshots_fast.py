#!/usr/bin/env python3
"""
Fast Screenshot Fix Pipeline
==============================
Two-phase approach:
Phase 1: LibreOffice full-workbook PDF export (one call per workbook) → per-page PNGs
Phase 2: Fix Pillow tile renders using NotoSansCJK font (correct ASCII + CJK rendering)

For page-to-sheet mapping, we use a heuristic:
- WB1 has 2 sheets: 概要 (overview) and フローチャート (flowchart) 
- WB2 has 27 sheets, mostly wide mapping tables

Since exact page-to-sheet mapping is complex, we:
1. Export full PDF pages as "lo_page_NNN.png" reference images
2. Re-render Pillow tiles with correct font
3. Use the LibreOffice pages as additional evidence alongside tile renders
"""

import subprocess
import os
import sys
import shutil
import math
import time
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import openpyxl

# Config
FONT_PATH = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
FONT_SIZE = 12
CELL_WIDTH_PX = 120
CELL_HEIGHT_PX = 22
DPI = 200
MAX_DIM = 7900

OUTPUT_BASE = os.path.expanduser("~/projects/hermes_bedrock_agent/outputs/fresh_parse_20260519")
SOURCE_DIR = "/tmp/s3_downloads/サンプル20260519"


def render_tile(ws, max_row, max_col, start_row, end_row, start_col, end_col, font):
    """Render a tile of Excel data with correct font."""
    end_row = min(end_row, max_row)
    end_col = min(end_col, max_col)
    
    num_rows = end_row - start_row + 1
    num_cols = end_col - start_col + 1
    
    img_width = min(num_cols * CELL_WIDTH_PX, 12000)
    img_height = min(num_rows * CELL_HEIGHT_PX, 12000)
    
    img = Image.new('RGB', (img_width, img_height), 'white')
    draw = ImageDraw.Draw(img)
    
    for row_idx in range(num_rows):
        actual_row = start_row + row_idx
        y = row_idx * CELL_HEIGHT_PX
        draw.line([(0, y), (img_width, y)], fill='#CCCCCC', width=1)
        
        for col_idx in range(num_cols):
            actual_col = start_col + col_idx
            x = col_idx * CELL_WIDTH_PX
            
            if row_idx == 0:
                draw.line([(x, 0), (x, img_height)], fill='#CCCCCC', width=1)
            
            try:
                cell = ws.cell(actual_row, actual_col)
                value = cell.value
                if value is not None:
                    text = str(value)[:20]
                    is_header = actual_row <= 3 or (cell.font and cell.font.bold)
                    color = '#000066' if is_header else '#333333'
                    draw.text((x + 2, y + 3), text, fill=color, font=font)
            except Exception:
                pass
    
    draw.line([(0, img_height - 1), (img_width - 1, img_height - 1)], fill='#CCCCCC', width=1)
    draw.line([(img_width - 1, 0), (img_width - 1, img_height - 1)], fill='#CCCCCC', width=1)
    
    return img


def process_workbook(xlsx_path, wb_name):
    """Process one workbook: LO export + Pillow tile fix."""
    wb_dir = os.path.join(OUTPUT_BASE, wb_name)
    print(f"\n{'='*60}")
    print(f"Workbook: {wb_name}")
    print(f"{'='*60}")
    
    # Phase 1: LibreOffice full-workbook PDF export
    tmp_dir = f"/tmp/lo_fix/{wb_name}"
    os.makedirs(tmp_dir, exist_ok=True)
    
    print(f"  Phase 1: LibreOffice PDF export...")
    result = subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'pdf',
         '--outdir', tmp_dir, xlsx_path],
        capture_output=True, text=True, timeout=120
    )
    
    # Find generated PDF
    pdfs = [f for f in os.listdir(tmp_dir) if f.endswith('.pdf')]
    if pdfs:
        pdf_path = os.path.join(tmp_dir, pdfs[0])
        # Convert to PNGs
        page_prefix = os.path.join(tmp_dir, "page")
        subprocess.run(
            ['pdftoppm', '-png', '-r', str(DPI), pdf_path, page_prefix],
            capture_output=True, timeout=120
        )
        pages = sorted([f for f in os.listdir(tmp_dir) if f.startswith("page") and f.endswith('.png')])
        print(f"    Generated {len(pages)} PDF pages")
        
        # Save LO reference pages to a lo_reference folder in the workbook dir
        lo_ref_dir = os.path.join(wb_dir, "lo_reference_pages")
        os.makedirs(lo_ref_dir, exist_ok=True)
        for p in pages:
            shutil.copy2(os.path.join(tmp_dir, p), os.path.join(lo_ref_dir, p))
        print(f"    Saved LO reference pages to {lo_ref_dir}")
    else:
        print(f"    WARNING: PDF export failed")
        pages = []
    
    # Phase 2: Re-render Pillow tiles with correct font
    print(f"  Phase 2: Re-rendering tiles with NotoSansCJK font...")
    
    try:
        font = ImageFont.truetype(FONT_PATH, FONT_SIZE)
        print(f"    Font loaded: {FONT_PATH}")
    except Exception as e:
        print(f"    ERROR: Cannot load font: {e}")
        return
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        
        if max_row == 0 or max_col == 0:
            print(f"    Sheet {idx} ({sheet_name}): empty, skipping")
            continue
        
        safe_name = sheet_name.replace('/', '_').replace('\\', '_').replace(' ', '_')[:40]
        sheet_dir = os.path.join(wb_dir, "sheets", f"{idx:02d}_{safe_name}")
        screenshots_dir = os.path.join(sheet_dir, "screenshots")
        os.makedirs(screenshots_dir, exist_ok=True)
        
        print(f"    Sheet {idx}: {sheet_name} ({max_row}x{max_col})")
        
        # Generate tiles
        cols_per_tile = 50
        rows_per_tile = 150
        
        num_col_tiles = math.ceil(max_col / cols_per_tile)
        num_row_tiles = math.ceil(max_row / rows_per_tile)
        
        tiles = []
        tile_idx = 0
        
        for row_tile in range(num_row_tiles):
            for col_tile in range(num_col_tiles):
                start_row = row_tile * rows_per_tile + 1
                end_row = min((row_tile + 1) * rows_per_tile, max_row)
                start_col = col_tile * cols_per_tile + 1
                end_col = min((col_tile + 1) * cols_per_tile, max_col)
                
                tile_idx += 1
                tile_path = os.path.join(screenshots_dir, f"tile_{tile_idx:03d}.png")
                
                img = render_tile(ws, max_row, max_col, start_row, end_row, start_col, end_col, font)
                img.save(tile_path, 'PNG')
                
                tiles.append({
                    'path': tile_path,
                    'row_tile': row_tile,
                    'col_tile': col_tile,
                    'width': img.width,
                    'height': img.height,
                })
                img.close()
        
        # Stitch tiles into full sheet
        if tiles:
            max_row_tile = max(t['row_tile'] for t in tiles)
            max_col_tile = max(t['col_tile'] for t in tiles)
            
            # Calculate layout
            row_heights = []
            for rt in range(max_row_tile + 1):
                h = max(t['height'] for t in tiles if t['row_tile'] == rt)
                row_heights.append(h)
            
            col_widths = []
            for ct in range(max_col_tile + 1):
                w = max(t['width'] for t in tiles if t['col_tile'] == ct)
                col_widths.append(w)
            
            total_w = sum(col_widths)
            total_h = sum(row_heights)
            
            stitched = Image.new('RGB', (total_w, total_h), 'white')
            
            for t in tiles:
                tile_img = Image.open(t['path'])
                x = sum(col_widths[:t['col_tile']])
                y = sum(row_heights[:t['row_tile']])
                stitched.paste(tile_img, (x, y))
                tile_img.close()
            
            # Resize if needed
            w, h = stitched.size
            if max(w, h) > MAX_DIM:
                ratio = MAX_DIM / max(w, h)
                stitched = stitched.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
            
            stitched_path = os.path.join(screenshots_dir, "stitched_full_sheet.png")
            stitched.save(stitched_path, 'PNG')
            stitched.close()
            
            print(f"      {tile_idx} tiles → stitched ({total_w}x{total_h})")
    
    wb.close()
    shutil.rmtree(tmp_dir, ignore_errors=True)


def main():
    os.makedirs("/tmp/lo_fix", exist_ok=True)
    
    workbooks = [
        ("01_基本設計/M社様_DSSスクリプト改修概要_フローチャート.xlsx",
         "M社様_DSSスクリプト改修概要_フローチャート"),
        ("02_詳細設計/MW_IFマッピング定義書_205_発注情報(登録・変更・取消).xlsx",
         "MW_IFマッピング定義書_205_発注情報(登録・変更・取消)"),
    ]
    
    start = time.time()
    for rel_path, wb_name in workbooks:
        xlsx_path = os.path.join(SOURCE_DIR, rel_path)
        if not os.path.exists(xlsx_path):
            print(f"SKIP: {xlsx_path}")
            continue
        process_workbook(xlsx_path, wb_name)
    
    elapsed = time.time() - start
    print(f"\n{'='*60}")
    print(f"Total time: {elapsed:.0f}s")
    print(f"{'='*60}")
    
    shutil.rmtree("/tmp/lo_fix", ignore_errors=True)


if __name__ == '__main__':
    main()
