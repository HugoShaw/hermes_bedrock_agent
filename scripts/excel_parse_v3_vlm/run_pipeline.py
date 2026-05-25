#!/usr/bin/env python3
"""
Excel Parse Pipeline v3 — VLM-Enhanced
=======================================

Key improvements over v2:
1. VLM-based sheet classification and layout understanding
2. Tiled image rendering for large sheets (avoids partial screenshot problems)
3. VLM-validated markdown generation with semantic understanding
4. Merged VLM + deterministic parsing for higher quality outputs
5. Better handling of multi-table sheets, metadata sections, and mapping tables

Architecture:
  S3 scan → download → workbook inventory → sheet rendering (tiles)
  → VLM classification → deterministic extraction → VLM validation
  → structured output → graph nodes/edges → KB chunks → quality report → S3 sync
"""

import os
import sys
import json
import hashlib
import base64
import time
import re
import io
import copy
import traceback
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import boto3
from botocore.config import Config
from PIL import Image, ImageDraw, ImageFont

# Load .env
from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env")

# ==============================================================================
# CONFIGURATION
# ==============================================================================

INPUT_DIR = PROJECT_ROOT / "data" / "input" / "sample_20260519"
OUTPUT_DIR = PROJECT_ROOT / "data" / "outputs" / "excel_parse_pipeline_v3_vlm" / "sample_20260519"
S3_BUCKET = "s3-hulftchina-rd"
S3_INPUT_PREFIX = "サンプル20260519/"
S3_OUTPUT_PREFIX = "output/sample_20260519/excel_parse_pipeline_v3_vlm/"

BEDROCK_VLM_MODEL = os.getenv("BEDROCK_VLM_MODEL_ID", "jp.anthropic.claude-sonnet-4-6")
BEDROCK_TEXT_MODEL = os.getenv("BEDROCK_TEXT_MODEL_ID", "jp.anthropic.claude-sonnet-4-6")
AWS_REGION = os.getenv("AWS_DEFAULT_REGION", "ap-northeast-1")

# VLM call settings
VLM_MAX_TOKENS = 8000
VLM_DELAY_BETWEEN_CALLS = 3  # seconds between VLM calls to avoid throttling

# Tile rendering settings
TILE_MAX_COLS = 30  # Max columns per tile for wide sheets
TILE_MAX_ROWS = 60  # Max rows per tile for long sheets
CELL_WIDTH_PX = 100  # Default cell width in pixels
CELL_HEIGHT_PX = 22  # Default cell height in pixels
MAX_CELL_TEXT_LEN = 50  # Truncate cell text for rendering

# ==============================================================================
# BEDROCK CLIENT
# ==============================================================================

_bedrock_client = None

def get_bedrock_client():
    global _bedrock_client
    if _bedrock_client is None:
        _bedrock_client = boto3.client(
            "bedrock-runtime",
            region_name=AWS_REGION,
            config=Config(read_timeout=600, retries={"max_attempts": 3})
        )
    return _bedrock_client


def call_vlm(prompt: str, images: list[bytes] = None, max_tokens: int = VLM_MAX_TOKENS) -> str:
    """Call Bedrock VLM with optional images. Returns text response."""
    client = get_bedrock_client()
    
    content = []
    if images:
        for img_bytes in images:
            # Bedrock Converse API expects raw bytes, NOT base64
            content.append({
                "image": {
                    "format": "png",
                    "source": {"bytes": img_bytes}
                }
            })
    content.append({"text": prompt})
    
    messages = [{"role": "user", "content": content}]
    
    try:
        response = client.converse(
            modelId=BEDROCK_VLM_MODEL,
            messages=messages,
            inferenceConfig={"maxTokens": max_tokens, "temperature": 0.1}
        )
        output = response.get("output", {})
        message = output.get("message", {})
        resp_content = message.get("content", [])
        text_parts = [block["text"] for block in resp_content if "text" in block]
        return "\n".join(text_parts)
    except Exception as e:
        print(f"  [VLM ERROR] {e}")
        return f"[VLM_ERROR: {str(e)[:200]}]"


def call_text_llm(prompt: str, max_tokens: int = 4000) -> str:
    """Call Bedrock text LLM (no images). Returns text response."""
    client = get_bedrock_client()
    messages = [{"role": "user", "content": [{"text": prompt}]}]
    
    try:
        response = client.converse(
            modelId=BEDROCK_TEXT_MODEL,
            messages=messages,
            inferenceConfig={"maxTokens": max_tokens, "temperature": 0.1}
        )
        output = response.get("output", {})
        message = output.get("message", {})
        resp_content = message.get("content", [])
        text_parts = [block["text"] for block in resp_content if "text" in block]
        return "\n".join(text_parts)
    except Exception as e:
        print(f"  [LLM ERROR] {e}")
        return f"[LLM_ERROR: {str(e)[:200]}]"


# ==============================================================================
# S3 IO
# ==============================================================================

_s3_client = None

def get_s3_client():
    global _s3_client
    if _s3_client is None:
        _s3_client = boto3.client("s3", region_name=AWS_REGION)
    return _s3_client


def s3_sync_outputs(output_dir: Path, s3_prefix: str):
    """Upload all output files to S3."""
    s3 = get_s3_client()
    count = 0
    for root, dirs, files in os.walk(output_dir):
        for fname in files:
            local_path = Path(root) / fname
            rel_path = local_path.relative_to(output_dir)
            s3_key = s3_prefix + str(rel_path)
            try:
                s3.upload_file(str(local_path), S3_BUCKET, s3_key)
                count += 1
            except Exception as e:
                print(f"  [S3 UPLOAD ERROR] {s3_key}: {e}")
    return count


# ==============================================================================
# EXCEL SHEET IMAGE RENDERER
# ==============================================================================

def render_sheet_to_image(ws, start_row=1, end_row=None, start_col=1, end_col=None,
                          cell_width=CELL_WIDTH_PX, cell_height=CELL_HEIGHT_PX,
                          max_width=4000, max_height=6000) -> bytes:
    """
    Render a portion of an Excel worksheet to a PNG image.
    Returns PNG bytes.
    
    This creates a visual representation showing cell values, merged cells,
    and basic formatting. Not pixel-perfect but good enough for VLM analysis.
    """
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    
    # Clamp to reasonable bounds
    end_row = min(end_row, start_row + TILE_MAX_ROWS - 1)
    end_col = min(end_col, start_col + TILE_MAX_COLS - 1)
    
    num_rows = end_row - start_row + 1
    num_cols = end_col - start_col + 1
    
    # Calculate image dimensions
    img_width = min(num_cols * cell_width + 50, max_width)  # +50 for row labels
    img_height = min(num_rows * cell_height + 30, max_height)  # +30 for col headers
    
    # Create image
    img = Image.new("RGB", (img_width, img_height), "white")
    draw = ImageDraw.Draw(img)
    
    # Try to load a font with Japanese support
    try:
        # DroidSansFallbackFull supports CJK characters
        font = ImageFont.truetype("/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf", 11)
        small_font = ImageFont.truetype("/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf", 9)
    except Exception:
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 11)
            small_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 9)
        except Exception:
            font = ImageFont.load_default()
            small_font = font
    
    header_h = 25  # Height of column header area
    label_w = 45   # Width of row label area
    
    # Draw column headers
    for ci in range(num_cols):
        col = start_col + ci
        col_letter = get_column_letter(col)
        x = label_w + ci * cell_width
        draw.rectangle([x, 0, x + cell_width - 1, header_h - 1], fill="#E8E8E8", outline="#CCCCCC")
        draw.text((x + 2, 4), col_letter, fill="black", font=small_font)
    
    # Draw row labels
    for ri in range(num_rows):
        row = start_row + ri
        y = header_h + ri * cell_height
        draw.rectangle([0, y, label_w - 1, y + cell_height - 1], fill="#E8E8E8", outline="#CCCCCC")
        draw.text((2, y + 3), str(row), fill="black", font=small_font)
    
    # Get merged cell ranges for this area
    merged_ranges = []
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= end_row and merged.max_row >= start_row and
            merged.min_col <= end_col and merged.max_col >= start_col):
            merged_ranges.append(merged)
    
    # Track which cells are part of a merge (not the top-left)
    merged_cells_set = set()
    for mr in merged_ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    merged_cells_set.add((r, c))
    
    # Draw cells
    for ri in range(num_rows):
        row = start_row + ri
        for ci in range(num_cols):
            col = start_col + ci
            
            if (row, col) in merged_cells_set:
                continue  # Skip non-top-left merged cells
            
            x = label_w + ci * cell_width
            y = header_h + ri * cell_height
            
            # Check if this is a merged cell start
            cell_w = cell_width
            cell_h = cell_height
            for mr in merged_ranges:
                if mr.min_row == row and mr.min_col == col:
                    merge_cols = min(mr.max_col, end_col) - col + 1
                    merge_rows = min(mr.max_row, end_row) - row + 1
                    cell_w = merge_cols * cell_width
                    cell_h = merge_rows * cell_height
                    break
            
            # Get cell value and formatting
            cell = ws.cell(row, col)
            value = cell.value
            
            # Determine fill color
            fill_color = "white"
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
                try:
                    rgb = cell.fill.fgColor.rgb
                    if isinstance(rgb, str) and len(rgb) >= 6:
                        fill_color = f"#{rgb[-6:]}"
                except Exception:
                    pass
            
            # Draw cell
            draw.rectangle([x, y, x + cell_w - 1, y + cell_h - 1], fill=fill_color, outline="#AAAAAA")
            
            # Draw value
            if value is not None:
                text = str(value)[:MAX_CELL_TEXT_LEN]
                text_color = "black"
                if cell.font and cell.font.bold:
                    text_color = "#000080"  # Dark blue for bold
                draw.text((x + 2, y + 3), text, fill=text_color, font=font)
    
    # Convert to PNG bytes
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def render_sheet_overview(ws, max_cols=80, max_rows=100) -> bytes:
    """Render a zoomed-out overview of the entire sheet (smaller cells)."""
    actual_cols = min(ws.max_column, max_cols)
    actual_rows = min(ws.max_row, max_rows)
    
    # Use smaller cell sizes for overview
    cell_w = max(40, min(80, 3200 // actual_cols))
    cell_h = max(14, min(20, 2400 // actual_rows))
    
    return render_sheet_to_image(
        ws, start_row=1, end_row=actual_rows,
        start_col=1, end_col=actual_cols,
        cell_width=cell_w, cell_height=cell_h,
        max_width=4000, max_height=4000
    )


def render_sheet_tiles(ws) -> list[dict]:
    """
    Render a sheet in non-overlapping tiles.
    Returns list of {tile_id, start_row, end_row, start_col, end_col, image_bytes}
    """
    tiles = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    tile_id = 0
    row = 1
    while row <= max_row:
        end_row = min(row + TILE_MAX_ROWS - 1, max_row)
        col = 1
        while col <= max_col:
            end_col = min(col + TILE_MAX_COLS - 1, max_col)
            
            # Check if this tile has any content
            has_content = False
            for r in range(row, end_row + 1):
                for c in range(col, end_col + 1):
                    if ws.cell(r, c).value is not None:
                        has_content = True
                        break
                if has_content:
                    break
            
            if has_content:
                img_bytes = render_sheet_to_image(
                    ws, start_row=row, end_row=end_row,
                    start_col=col, end_col=end_col
                )
                tiles.append({
                    "tile_id": tile_id,
                    "start_row": row,
                    "end_row": end_row,
                    "start_col": col,
                    "end_col": end_col,
                    "start_col_letter": get_column_letter(col),
                    "end_col_letter": get_column_letter(end_col),
                    "image_bytes": img_bytes
                })
                tile_id += 1
            
            col = end_col + 1
        row = end_row + 1
    
    return tiles


# ==============================================================================
# DETERMINISTIC EXTRACTION HELPERS
# ==============================================================================

def extract_sheet_cell_matrix(ws, start_row=1, end_row=None, start_col=1, end_col=None) -> list[list]:
    """Extract cell values as a 2D matrix."""
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    
    matrix = []
    for r in range(start_row, end_row + 1):
        row_data = []
        for c in range(start_col, end_col + 1):
            v = ws.cell(r, c).value
            row_data.append(v)
        matrix.append(row_data)
    return matrix


def extract_merged_cells_info(ws) -> list[dict]:
    """Extract all merged cell ranges with their values."""
    merged_info = []
    for mr in ws.merged_cells.ranges:
        val = ws.cell(mr.min_row, mr.min_col).value
        merged_info.append({
            "range": str(mr),
            "min_row": mr.min_row,
            "max_row": mr.max_row,
            "min_col": mr.min_col,
            "max_col": mr.max_col,
            "value": str(val) if val is not None else None
        })
    return merged_info


def detect_non_empty_regions(ws) -> list[dict]:
    """
    Detect contiguous non-empty rectangular regions using gap analysis.
    Returns regions with their bounding coordinates.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    
    if max_row == 0 or max_col == 0:
        return []
    
    # Build occupancy map
    occupied_rows = set()
    occupied_cols = set()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if ws.cell(r, c).value is not None:
                occupied_rows.add(r)
                occupied_cols.add(c)
    
    if not occupied_rows or not occupied_cols:
        return []
    
    # Find row gaps (2+ consecutive empty rows)
    sorted_rows = sorted(occupied_rows)
    row_groups = []
    current_group = [sorted_rows[0]]
    for i in range(1, len(sorted_rows)):
        if sorted_rows[i] - sorted_rows[i-1] > 2:  # Gap of 2+ empty rows
            row_groups.append((current_group[0], current_group[-1]))
            current_group = [sorted_rows[i]]
        else:
            current_group.append(sorted_rows[i])
    row_groups.append((current_group[0], current_group[-1]))
    
    # Find column gaps (3+ consecutive empty columns) for wide sheets
    sorted_cols = sorted(occupied_cols)
    col_groups = []
    current_group = [sorted_cols[0]]
    gap_threshold = 3 if max_col > 50 else 5
    for i in range(1, len(sorted_cols)):
        if sorted_cols[i] - sorted_cols[i-1] > gap_threshold:
            col_groups.append((current_group[0], current_group[-1]))
            current_group = [sorted_cols[i]]
        else:
            current_group.append(sorted_cols[i])
    col_groups.append((current_group[0], current_group[-1]))
    
    # Combine into regions
    regions = []
    region_id = 0
    for rg in row_groups:
        for cg in col_groups:
            # Check if there's actual content in this intersection
            has_content = False
            for r in range(rg[0], min(rg[1] + 1, rg[0] + 5)):  # Sample first 5 rows
                for c in range(cg[0], min(cg[1] + 1, cg[0] + 5)):
                    if ws.cell(r, c).value is not None:
                        has_content = True
                        break
                if has_content:
                    break
            
            if has_content:
                regions.append({
                    "region_id": f"R{region_id}",
                    "start_row": rg[0],
                    "end_row": rg[1],
                    "start_col": cg[0],
                    "end_col": cg[1],
                    "start_col_letter": get_column_letter(cg[0]),
                    "end_col_letter": get_column_letter(cg[1]),
                    "row_count": rg[1] - rg[0] + 1,
                    "col_count": cg[1] - cg[0] + 1
                })
                region_id += 1
    
    return regions


def extract_metadata_section(ws, max_row=20) -> dict:
    """Extract key-value metadata from the top section of a sheet."""
    metadata = {}
    for r in range(1, min(max_row + 1, ws.max_row + 1)):
        for c in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if val is not None:
                key = f"R{r}C{c}"
                metadata[key] = str(val)
    return metadata


def find_header_row(ws, start_col=1, end_col=None, search_start=1, search_end=25) -> Optional[int]:
    """Find the most likely header row by looking for 'No' or dense text rows."""
    if end_col is None:
        end_col = ws.max_column
    
    for r in range(search_start, min(search_end + 1, ws.max_row + 1)):
        for c in range(start_col, min(end_col + 1, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if val is not None and str(val).strip() == "No":
                return r
    return None


# ==============================================================================
# VLM-BASED SHEET CLASSIFICATION
# ==============================================================================

def classify_sheet_with_vlm(ws_name: str, overview_image: bytes, cell_summary: str) -> dict:
    """
    Use VLM to classify a sheet and understand its layout.
    Returns classification dict with sheet_type, regions, and notes.
    """
    prompt = f"""You are analyzing a Japanese enterprise Excel design document sheet.

Sheet name: {ws_name}
Cell content summary (first 20 rows, showing non-empty cells):
{cell_summary}

Please analyze this sheet and provide a JSON classification:

1. sheet_type: one of [mapping_spec, flowchart, overview, data_retrieval_condition, api_call_sequence, change_history, development_spec, code_conversion, error_table, unknown]
2. description: brief description of what this sheet contains (in English)
3. regions: list of logical regions you can identify, each with:
   - region_name: human-readable name
   - purpose: what this region contains
   - approximate_rows: [start, end]
   - approximate_cols: [start_letter, end_letter]
4. has_mapping_table: boolean - does this contain source-to-target field mappings?
5. has_metadata_section: boolean - is there a metadata header section at the top?
6. key_observations: list of important notes about this sheet's structure

Reply ONLY with valid JSON, no markdown code fences."""

    response = call_vlm(prompt, images=[overview_image])
    
    # Try to parse JSON from response
    try:
        # Strip markdown fences if present
        text = response.strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```\s*$", "", text)
        return json.loads(text)
    except json.JSONDecodeError:
        return {
            "sheet_type": "unknown",
            "description": f"VLM classification failed to parse: {response[:200]}",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": False,
            "key_observations": ["VLM response was not valid JSON"]
        }


# ==============================================================================
# MAPPING TABLE EXTRACTION (DETERMINISTIC + VLM VALIDATION)
# ==============================================================================

def extract_mapping_sheet_deterministic(ws) -> dict:
    """
    Extract mapping table data deterministically using the known structure:
    - Metadata section: rows 1-19
    - Header row: row 20-22 (contains "No")
    - Source table: left side
    - Target table: right side
    - マッピング元 column in target table links to source row numbers
    """
    max_row = ws.max_row
    max_col = ws.max_column
    
    result = {
        "metadata": {},
        "source_fields": [],
        "target_fields": [],
        "mappings": [],
        "uncertainties": []
    }
    
    # Extract metadata
    result["metadata"] = extract_metadata_section(ws)
    
    # Find "No" columns to detect source/target boundary
    no_columns = []
    for check_row in range(19, min(24, max_row + 1)):
        for col in range(1, max_col + 1):
            val = ws.cell(check_row, col).value
            if val is not None and str(val).strip() == "No":
                no_columns.append((col, check_row))
    
    if len(no_columns) < 2:
        result["uncertainties"].append({
            "type": "boundary_detection_failed",
            "message": f"Found {len(no_columns)} 'No' columns, expected 2+",
            "no_columns_found": no_columns
        })
        return result
    
    # Sort by column position and deduplicate
    no_columns.sort(key=lambda x: x[0])
    unique_cols = []
    seen_cols = set()
    for col, row in no_columns:
        if col not in seen_cols:
            unique_cols.append((col, row))
            seen_cols.add(col)
    
    if len(unique_cols) < 2:
        result["uncertainties"].append({
            "type": "boundary_detection_failed",
            "message": f"After dedup: {len(unique_cols)} unique 'No' columns"
        })
        return result
    
    source_no_col, source_header_row = unique_cols[0]
    target_no_col, target_header_row = unique_cols[1]
    
    # Compute boundary
    boundary_col = (source_no_col + target_no_col) // 2
    
    # Extract source headers
    source_headers = {}
    for c in range(source_no_col, boundary_col):
        val = ws.cell(source_header_row, c).value
        if val is not None:
            source_headers[c] = str(val).strip()
    
    # Extract target headers
    target_headers = {}
    for c in range(target_no_col, max_col + 1):
        val = ws.cell(target_header_row, c).value
        if val is not None:
            target_headers[c] = str(val).strip()
    
    # Find key columns by header name
    source_name_col = None
    source_variable_col = None
    target_name_col = None
    target_variable_col = None
    mapping_source_col = None  # マッピング元
    edit_content_col = None   # 編集内容
    
    for c, h in source_headers.items():
        if h == "項目名称" and source_name_col is None:
            source_name_col = c
        if h == "変数" and source_variable_col is None:
            source_variable_col = c
    
    for c, h in target_headers.items():
        if h == "項目名称" and target_name_col is None:
            target_name_col = c
        if h == "変数" and target_variable_col is None:
            target_variable_col = c
        if h == "マッピング元" and mapping_source_col is None:
            mapping_source_col = c
        if h == "編集内容" and edit_content_col is None:
            edit_content_col = c
    
    # Data start row
    data_start = max(source_header_row, target_header_row) + 1
    
    # Extract source fields
    source_field_map = {}  # No -> field info
    for r in range(data_start, max_row + 1):
        no_val = ws.cell(r, source_no_col).value
        if no_val is None:
            continue
        try:
            no_int = int(no_val)
        except (ValueError, TypeError):
            continue
        
        field_name = ws.cell(r, source_name_col).value if source_name_col else None
        variable = ws.cell(r, source_variable_col).value if source_variable_col else None
        
        field = {
            "no": no_int,
            "row": r,
            "field_name": str(field_name) if field_name else None,
            "variable": str(variable) if variable else None,
        }
        
        # Extract all source column values
        for c, h in source_headers.items():
            if c != source_no_col and c != source_name_col and c != source_variable_col:
                val = ws.cell(r, c).value
                if val is not None:
                    field[f"col_{h}"] = str(val)
        
        result["source_fields"].append(field)
        source_field_map[str(no_int)] = field
    
    # Extract target fields and mappings
    for r in range(data_start, max_row + 1):
        no_val = ws.cell(r, target_no_col).value
        if no_val is None:
            continue
        try:
            no_int = int(no_val)
        except (ValueError, TypeError):
            continue
        
        field_name = ws.cell(r, target_name_col).value if target_name_col else None
        variable = ws.cell(r, target_variable_col).value if target_variable_col else None
        mapping_ref = ws.cell(r, mapping_source_col).value if mapping_source_col else None
        edit_content = ws.cell(r, edit_content_col).value if edit_content_col else None
        
        target_field = {
            "no": no_int,
            "row": r,
            "field_name": str(field_name) if field_name else None,
            "variable": str(variable) if variable else None,
        }
        
        # Extract all target column values
        for c, h in target_headers.items():
            if c not in (target_no_col, target_name_col, target_variable_col, mapping_source_col, edit_content_col):
                val = ws.cell(r, c).value
                if val is not None:
                    target_field[f"col_{h}"] = str(val)
        
        result["target_fields"].append(target_field)
        
        # Process mapping reference
        if mapping_ref is not None:
            ref_str = str(mapping_ref).strip()
            mapping = {
                "target_no": no_int,
                "target_field_name": str(field_name) if field_name else None,
                "target_variable": str(variable) if variable else None,
                "source_ref_raw": ref_str,
                "edit_content": str(edit_content) if edit_content else None,
                "row": r,
                "confidence": 0.0,
                "ref_type": "unknown",
                "resolved_sources": []
            }
            
            # Classify and resolve reference
            if ref_str == "-" or ref_str == "ー":
                mapping["ref_type"] = "no_mapping"
                mapping["confidence"] = 0.9
            elif ref_str == "固定値":
                mapping["ref_type"] = "fixed_value"
                mapping["confidence"] = 0.85
            elif ref_str in ("ヘッダ", "明細"):
                mapping["ref_type"] = "section_reference"
                mapping["confidence"] = 0.8
            elif ref_str == "マッピング元":
                # This is a sub-header echo, not real data
                mapping["ref_type"] = "sub_header_echo"
                mapping["confidence"] = 0.3
                result["uncertainties"].append({
                    "type": "sub_header_echo",
                    "row": r,
                    "message": f"Row {r} has 'マッピング元' as data — likely a sub-header separator"
                })
            elif "\n" in ref_str:
                # Multi-reference
                refs = [x.strip() for x in ref_str.split("\n") if x.strip()]
                resolved = []
                for ref in refs:
                    if ref in source_field_map:
                        resolved.append(source_field_map[ref].get("field_name", f"Source#{ref}"))
                    elif ref == "固定値":
                        resolved.append("[固定値]")
                    else:
                        resolved.append(f"[unresolved:{ref}]")
                mapping["ref_type"] = "multi_source"
                mapping["resolved_sources"] = [r for r in resolved if r is not None]
                mapping["confidence"] = 0.7 if all("[unresolved" not in (r or "") for r in resolved) else 0.5
            else:
                # Try single numeric reference
                if ref_str in source_field_map:
                    mapping["ref_type"] = "single_source"
                    mapping["resolved_sources"] = [source_field_map[ref_str].get("field_name", f"Source#{ref_str}")]
                    mapping["confidence"] = 0.9
                elif ref_str.startswith("※"):
                    # Note-prefixed reference
                    num = re.sub(r"[※\(\)（）]", "", ref_str).strip()
                    if num in source_field_map:
                        mapping["ref_type"] = "conditional_source"
                        mapping["resolved_sources"] = [source_field_map[num].get("field_name", f"Source#{num}")]
                        mapping["confidence"] = 0.75
                    else:
                        mapping["ref_type"] = "unknown"
                        mapping["confidence"] = 0.4
                        result["uncertainties"].append({
                            "type": "unresolved_reference",
                            "row": r,
                            "source_ref": ref_str,
                            "target_field": str(field_name) if field_name else None
                        })
                else:
                    mapping["ref_type"] = "unknown"
                    mapping["confidence"] = 0.4
                    result["uncertainties"].append({
                        "type": "unresolved_reference",
                        "row": r,
                        "source_ref": ref_str,
                        "target_field": str(field_name) if field_name else None
                    })
            
            result["mappings"].append(mapping)
    
    return result


# ==============================================================================
# VLM VALIDATION OF EXTRACTION
# ==============================================================================

def vlm_validate_mapping_extraction(ws_name: str, tile_image: bytes,
                                     cell_text: str, extraction_result: dict) -> dict:
    """
    Use VLM to validate deterministic extraction results.
    Checks for missed fields, wrong boundaries, or misclassifications.
    """
    # Summarize extraction for VLM
    src_count = len(extraction_result.get("source_fields", []))
    tgt_count = len(extraction_result.get("target_fields", []))
    map_count = len(extraction_result.get("mappings", []))
    unc_count = len(extraction_result.get("uncertainties", []))
    
    prompt = f"""You are validating the extraction results from a Japanese enterprise mapping sheet.

Sheet: {ws_name}

Extraction summary:
- Source fields found: {src_count}
- Target fields found: {tgt_count}
- Mappings resolved: {map_count}
- Uncertainties: {unc_count}

Cell content around the header area:
{cell_text[:3000]}

Looking at the image (which shows part of the sheet), please verify:
1. Does the extraction count seem reasonable for what you see?
2. Are there any obvious tables or regions that might be missed?
3. Is the source/target boundary correctly identified?

Reply with JSON:
{{
  "validation_status": "ok" | "issues_found",
  "source_count_reasonable": true/false,
  "target_count_reasonable": true/false,
  "missed_regions": [],
  "notes": "..."
}}

Reply ONLY with valid JSON."""

    response = call_vlm(prompt, images=[tile_image])
    
    try:
        text = response.strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```\s*$", "", text)
        return json.loads(text)
    except json.JSONDecodeError:
        return {
            "validation_status": "parse_error",
            "notes": response[:300]
        }


# ==============================================================================
# MARKDOWN GENERATION (ENHANCED WITH STRUCTURE UNDERSTANDING)
# ==============================================================================

def generate_sheet_markdown(ws, ws_name: str, sheet_classification: dict,
                           extraction_result: dict = None) -> str:
    """
    Generate high-quality markdown for a sheet, using both deterministic data
    and VLM classification insights.
    """
    md_parts = []
    md_parts.append(f"# {ws_name}\n")
    
    sheet_type = sheet_classification.get("sheet_type", "unknown")
    description = sheet_classification.get("description", "")
    md_parts.append(f"**Sheet Type:** {sheet_type}")
    md_parts.append(f"**Description:** {description}")
    md_parts.append(f"**Dimensions:** {ws.max_row} rows × {ws.max_column} columns")
    md_parts.append(f"**Merged Cells:** {len(ws.merged_cells.ranges)}\n")
    
    # Add key observations
    observations = sheet_classification.get("key_observations", [])
    if observations:
        md_parts.append("## Key Observations\n")
        for obs in observations:
            md_parts.append(f"- {obs}")
        md_parts.append("")
    
    if sheet_type == "mapping_spec" and extraction_result:
        # Use structured extraction for mapping sheets
        md_parts.append(generate_mapping_sheet_markdown(ws, ws_name, extraction_result))
    elif sheet_type == "data_retrieval_condition":
        md_parts.append(generate_condition_sheet_markdown(ws))
    elif sheet_type == "flowchart":
        md_parts.append(generate_flowchart_sheet_markdown(ws))
    else:
        # Generic: dump as tables based on detected regions
        md_parts.append(generate_generic_sheet_markdown(ws))
    
    return "\n".join(md_parts)


def generate_mapping_sheet_markdown(ws, ws_name: str, extraction: dict) -> str:
    """Generate detailed markdown for a mapping sheet with source, target, and mappings."""
    parts = []
    
    # Metadata section
    metadata = extraction.get("metadata", {})
    if metadata:
        parts.append("## Metadata\n")
        # Group metadata into meaningful sections
        important_keys = {}
        for key, val in metadata.items():
            if val and len(val) > 1 and val not in ("ー", "-"):
                important_keys[key] = val
        
        # Extract specific known fields
        meta_table = []
        for key, val in sorted(important_keys.items()):
            meta_table.append(f"| {key} | {val.replace(chr(10), ' ')} |")
        
        if meta_table:
            parts.append("| Cell | Value |")
            parts.append("|------|-------|")
            parts.extend(meta_table[:30])  # Cap at 30 entries
            parts.append("")
    
    # Source fields table
    source_fields = extraction.get("source_fields", [])
    if source_fields:
        parts.append(f"## Source Fields ({len(source_fields)} fields)\n")
        parts.append("| No | Field Name | Variable |")
        parts.append("|----|-----------|----------|")
        for f in source_fields[:100]:  # Cap display
            no = f.get("no", "")
            name = (f.get("field_name") or "").replace("|", "\\|").replace("\n", " ")
            var = (f.get("variable") or "").replace("|", "\\|").replace("\n", " ")
            parts.append(f"| {no} | {name} | {var} |")
        if len(source_fields) > 100:
            parts.append(f"\n*... and {len(source_fields) - 100} more fields*\n")
        parts.append("")
    
    # Target fields table
    target_fields = extraction.get("target_fields", [])
    if target_fields:
        parts.append(f"## Target Fields ({len(target_fields)} fields)\n")
        parts.append("| No | Field Name | Variable |")
        parts.append("|----|-----------|----------|")
        for f in target_fields[:100]:
            no = f.get("no", "")
            name = (f.get("field_name") or "").replace("|", "\\|").replace("\n", " ")
            var = (f.get("variable") or "").replace("|", "\\|").replace("\n", " ")
            parts.append(f"| {no} | {name} | {var} |")
        if len(target_fields) > 100:
            parts.append(f"\n*... and {len(target_fields) - 100} more fields*\n")
        parts.append("")
    
    # Mappings table
    mappings = extraction.get("mappings", [])
    if mappings:
        parts.append(f"## Field Mappings ({len(mappings)} entries)\n")
        parts.append("| Target No | Target Field | Source Ref | Resolved Sources | Edit Content | Confidence |")
        parts.append("|-----------|-------------|-----------|-----------------|--------------|------------|")
        for m in mappings[:150]:
            tno = m.get("target_no", "")
            tname = (m.get("target_field_name") or "").replace("|", "\\|").replace("\n", " ")
            sref = (m.get("source_ref_raw") or "").replace("|", "\\|").replace("\n", ", ")
            resolved = ", ".join(s for s in m.get("resolved_sources", []) if s).replace("|", "\\|")
            edit = (m.get("edit_content") or "").replace("|", "\\|").replace("\n", " ")[:60]
            conf = m.get("confidence", 0)
            parts.append(f"| {tno} | {tname} | {sref} | {resolved} | {edit} | {conf:.1f} |")
        if len(mappings) > 150:
            parts.append(f"\n*... and {len(mappings) - 150} more mappings*\n")
        parts.append("")
    
    # Uncertainties
    uncertainties = extraction.get("uncertainties", [])
    if uncertainties:
        parts.append(f"## Uncertainties ({len(uncertainties)})\n")
        for u in uncertainties:
            parts.append(f"- **{u.get('type', 'unknown')}** (row {u.get('row', '?')}): {u.get('message', u.get('source_ref', ''))}")
        parts.append("")
    
    return "\n".join(parts)


def generate_condition_sheet_markdown(ws) -> str:
    """Generate markdown for data retrieval condition sheets."""
    parts = ["## Data Retrieval Conditions\n"]
    
    # These sheets are typically small, dump all non-empty cells
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append(str(v).replace("\n", " "))
        if row_vals:
            parts.append(f"Row {r}: " + " | ".join(row_vals))
    
    return "\n".join(parts)


def generate_flowchart_sheet_markdown(ws) -> str:
    """Generate markdown for flowchart sheets (mostly shapes, limited cell data)."""
    parts = ["## Flowchart Sheet\n"]
    parts.append("*Note: This sheet primarily contains Excel shapes/drawings.*")
    parts.append("*openpyxl cannot extract shapes. See manual Mermaid file if available.*\n")
    
    # Dump whatever text cells exist
    has_content = False
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                parts.append(f"- Cell {get_column_letter(c)}{r}: {v}")
                has_content = True
    
    if not has_content:
        parts.append("*(No text cell content — all content is in shapes/drawings)*")
    
    return "\n".join(parts)


def generate_generic_sheet_markdown(ws) -> str:
    """Generate markdown for generic/unknown sheet types."""
    parts = []
    
    # Detect regions
    regions = detect_non_empty_regions(ws)
    
    if not regions:
        parts.append("*(Empty sheet)*\n")
        return "\n".join(parts)
    
    parts.append(f"## Content ({len(regions)} region(s) detected)\n")
    
    for region in regions:
        sr, er = region["start_row"], region["end_row"]
        sc, ec = region["start_col"], region["end_col"]
        scl, ecl = region["start_col_letter"], region["end_col_letter"]
        
        parts.append(f"### Region: {scl}{sr}:{ecl}{er} ({region['row_count']}r × {region['col_count']}c)\n")
        
        # Extract as table
        # First, find headers (first non-empty row in region)
        header_row = None
        for r in range(sr, min(er + 1, sr + 5)):
            count = sum(1 for c in range(sc, ec + 1) if ws.cell(r, c).value is not None)
            if count >= 2:  # At least 2 non-empty cells = likely header
                header_row = r
                break
        
        if header_row:
            # Extract headers
            headers = []
            for c in range(sc, ec + 1):
                val = ws.cell(header_row, c).value
                headers.append(str(val).replace("|", "\\|").replace("\n", " ") if val else "")
            
            # Filter out empty trailing headers
            while headers and not headers[-1]:
                headers.pop()
            
            if headers:
                parts.append("| " + " | ".join(headers) + " |")
                parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
                
                # Data rows
                row_count = 0
                for r in range(header_row + 1, er + 1):
                    row_vals = []
                    for c in range(sc, sc + len(headers)):
                        val = ws.cell(r, c).value
                        row_vals.append(str(val).replace("|", "\\|").replace("\n", " ")[:80] if val else "")
                    
                    # Skip completely empty rows
                    if any(v for v in row_vals):
                        parts.append("| " + " | ".join(row_vals) + " |")
                        row_count += 1
                        if row_count >= 100:
                            parts.append(f"\n*... truncated at 100 rows (total ~{er - header_row} data rows)*\n")
                            break
                
                parts.append("")
        else:
            # No clear header — dump as key-value pairs
            for r in range(sr, min(er + 1, sr + 50)):
                vals = []
                for c in range(sc, ec + 1):
                    val = ws.cell(r, c).value
                    if val is not None:
                        vals.append(f"{get_column_letter(c)}{r}={str(val)[:60]}")
                if vals:
                    parts.append("  " + " | ".join(vals))
            parts.append("")
    
    return "\n".join(parts)


# ==============================================================================
# MERMAID PARSING
# ==============================================================================

def parse_mermaid_file(filepath: Path) -> dict:
    """Parse a Mermaid flowchart file and extract nodes and edges."""
    content = filepath.read_text(encoding="utf-8")
    
    nodes = []
    edges = []
    
    # Parse node definitions: A[label] or A{label} or A(label) or A((label))
    # and edge definitions: A --> B or A -->|label| B
    lines = content.split("\n")
    
    node_labels = {}  # id -> label
    
    for line in lines:
        line = line.strip()
        if not line or line.startswith("%%") or line.startswith("graph") or line.startswith("flowchart"):
            continue
        
        # Match edges: A --> B, A -->|text| B, A -- text --> B
        edge_patterns = [
            r"(\w+)\s*-->?\|([^|]*)\|\s*(\w+)",  # A -->|label| B
            r"(\w+)\s*--\s*(.+?)\s*-->\s*(\w+)",  # A -- label --> B
            r"(\w+)\s*-->\s*(\w+)",  # A --> B (no label)
            r"(\w+)\s*--->\s*(\w+)",  # A ---> B
            r"(\w+)\s*-\.->?\s*(\w+)",  # A -.-> B (dotted)
            r"(\w+)\s*==>\s*(\w+)",  # A ==> B (thick)
        ]
        
        edge_found = False
        for pattern in edge_patterns:
            match = re.search(pattern, line)
            if match:
                groups = match.groups()
                if len(groups) == 3:
                    edges.append({"from": groups[0], "to": groups[2], "label": groups[1]})
                elif len(groups) == 2:
                    edges.append({"from": groups[0], "to": groups[1], "label": ""})
                edge_found = True
                break
        
        # Match node definitions in the same line or standalone
        node_patterns = [
            r"(\w+)\[\"([^\"]+)\"\]",  # A["label"]
            r"(\w+)\[([^\]]+)\]",  # A[label]
            r"(\w+)\{\"([^\"]+)\"\}",  # A{"label"}
            r"(\w+)\{([^\}]+)\}",  # A{label}
            r"(\w+)\(\(([^\)]+)\)\)",  # A((label))
            r"(\w+)\(\"([^\"]+)\"\)",  # A("label")
            r"(\w+)\(([^\)]+)\)",  # A(label)
        ]
        
        for pattern in node_patterns:
            for match in re.finditer(pattern, line):
                node_id = match.group(1)
                label = match.group(2)
                if node_id not in node_labels:
                    node_labels[node_id] = label
    
    # Build nodes list
    for node_id, label in node_labels.items():
        node_type = "process"
        if node_id in [e["from"] for e in edges if e.get("label") and ("Yes" in e["label"] or "No" in e["label"] or "はい" in e["label"] or "いいえ" in e["label"])]:
            node_type = "decision"
        nodes.append({
            "id": node_id,
            "label": label,
            "type": node_type,
            "source": "manual_mermaid",
            "confidence": 1.0
        })
    
    return {
        "nodes": nodes,
        "edges": edges,
        "raw_content": content,
        "node_count": len(nodes),
        "edge_count": len(edges)
    }


# ==============================================================================
# GRAPH BUILDER
# ==============================================================================

def build_graph_nodes_edges(workbook_name: str, sheets_data: list[dict],
                           mermaid_data: dict = None) -> tuple[list[dict], list[dict]]:
    """Build GraphRAG nodes and edges from all extracted data."""
    nodes = []
    edges = []
    
    # Workbook node
    wb_id = f"wb_{hashlib.md5(workbook_name.encode()).hexdigest()[:8]}"
    nodes.append({
        "id": wb_id,
        "type": "Workbook",
        "label": workbook_name,
        "properties": {"file_name": workbook_name}
    })
    
    for sheet_data in sheets_data:
        sheet_name = sheet_data["sheet_name"]
        sheet_type = sheet_data.get("classification", {}).get("sheet_type", "unknown")
        sh_id = f"sh_{hashlib.md5((workbook_name + sheet_name).encode()).hexdigest()[:8]}"
        
        nodes.append({
            "id": sh_id,
            "type": "Sheet",
            "label": sheet_name,
            "properties": {"sheet_type": sheet_type, "workbook": workbook_name}
        })
        edges.append({"from": wb_id, "to": sh_id, "type": "HAS_SHEET"})
        
        extraction = sheet_data.get("extraction", {})
        
        # Source fields
        for field in extraction.get("source_fields", []):
            f_id = f"sf_{hashlib.md5((sheet_name + str(field.get('no', ''))).encode()).hexdigest()[:8]}"
            nodes.append({
                "id": f_id,
                "type": "SourceField",
                "label": field.get("field_name") or f"Source#{field.get('no')}",
                "properties": {
                    "no": field.get("no"),
                    "variable": field.get("variable"),
                    "sheet": sheet_name
                }
            })
            edges.append({"from": sh_id, "to": f_id, "type": "HAS_FIELD"})
        
        # Target fields
        for field in extraction.get("target_fields", []):
            f_id = f"tf_{hashlib.md5((sheet_name + 'T' + str(field.get('no', ''))).encode()).hexdigest()[:8]}"
            nodes.append({
                "id": f_id,
                "type": "TargetField",
                "label": field.get("field_name") or f"Target#{field.get('no')}",
                "properties": {
                    "no": field.get("no"),
                    "variable": field.get("variable"),
                    "sheet": sheet_name
                }
            })
            edges.append({"from": sh_id, "to": f_id, "type": "HAS_FIELD"})
        
        # Mappings as edges
        for mapping in extraction.get("mappings", []):
            if mapping.get("ref_type") in ("no_mapping", "sub_header_echo"):
                continue
            t_no = mapping.get("target_no")
            t_id = f"tf_{hashlib.md5((sheet_name + 'T' + str(t_no)).encode()).hexdigest()[:8]}"
            
            for source_name in mapping.get("resolved_sources", []):
                if source_name and "[unresolved" not in source_name:
                    # Find source field id
                    s_id = None
                    for sf in extraction.get("source_fields", []):
                        if sf.get("field_name") == source_name:
                            s_id = f"sf_{hashlib.md5((sheet_name + str(sf.get('no', ''))).encode()).hexdigest()[:8]}"
                            break
                    if s_id:
                        edges.append({
                            "from": s_id,
                            "to": t_id,
                            "type": "MAPS_TO",
                            "properties": {
                                "confidence": mapping.get("confidence", 0),
                                "ref_type": mapping.get("ref_type"),
                                "edit_content": mapping.get("edit_content")
                            }
                        })
    
    # Mermaid flowchart nodes and edges
    if mermaid_data:
        for node in mermaid_data.get("nodes", []):
            n_id = f"fn_{node['id']}"
            nodes.append({
                "id": n_id,
                "type": "FlowNode",
                "label": node["label"],
                "properties": {
                    "node_type": node.get("type", "process"),
                    "source": "manual_mermaid",
                    "confidence": 1.0
                }
            })
        
        for edge in mermaid_data.get("edges", []):
            edges.append({
                "from": f"fn_{edge['from']}",
                "to": f"fn_{edge['to']}",
                "type": "FLOW_TO",
                "properties": {"label": edge.get("label", "")}
            })
    
    return nodes, edges


# ==============================================================================
# KB CHUNK BUILDER
# ==============================================================================

def build_kb_chunks(workbook_name: str, sheets_data: list[dict],
                    mermaid_data: dict = None) -> dict[str, str]:
    """Build markdown KB chunks for vector retrieval."""
    chunks = {}
    
    # Workbook summary
    summary_parts = [f"# Workbook: {workbook_name}\n"]
    summary_parts.append(f"Total sheets: {len(sheets_data)}\n")
    summary_parts.append("## Sheet Overview\n")
    for sd in sheets_data:
        stype = sd.get("classification", {}).get("sheet_type", "unknown")
        desc = sd.get("classification", {}).get("description", "")
        summary_parts.append(f"- **{sd['sheet_name']}** ({stype}): {desc}")
    chunks["workbook_summary.md"] = "\n".join(summary_parts)
    
    # Per-sheet chunks
    for sd in sheets_data:
        sheet_name = sd["sheet_name"]
        safe_name = re.sub(r'[^\w\-]', '_', sheet_name)[:60]
        md = sd.get("markdown", "")
        if md:
            chunks[f"sheet_{safe_name}.md"] = md
    
    # Mapping records chunk
    mapping_parts = [f"# Field Mappings: {workbook_name}\n"]
    total_mappings = 0
    for sd in sheets_data:
        extraction = sd.get("extraction", {})
        mappings = extraction.get("mappings", [])
        if mappings:
            mapping_parts.append(f"\n## {sd['sheet_name']}\n")
            mapping_parts.append("| Target Field | Source Reference | Resolved | Confidence |")
            mapping_parts.append("|-------------|-----------------|----------|-----------|")
            for m in mappings:
                if m.get("ref_type") == "sub_header_echo":
                    continue
                tname = (m.get("target_field_name") or "?").replace("|", "/").replace("\n", " ")
                sref = (m.get("source_ref_raw") or "").replace("|", "/").replace("\n", ", ")
                resolved = ", ".join(s for s in m.get("resolved_sources", []) if s).replace("|", "/")
                conf = m.get("confidence", 0)
                mapping_parts.append(f"| {tname} | {sref} | {resolved} | {conf:.1f} |")
                total_mappings += 1
    
    if total_mappings > 0:
        chunks["mapping_records.md"] = "\n".join(mapping_parts)
    
    # Flowchart chunk
    if mermaid_data:
        fc_parts = ["# Flowchart (from manual Mermaid)\n"]
        fc_parts.append(f"Nodes: {mermaid_data['node_count']}, Edges: {mermaid_data['edge_count']}\n")
        fc_parts.append("## Nodes\n")
        for node in mermaid_data.get("nodes", []):
            fc_parts.append(f"- **{node['id']}**: {node['label']} ({node.get('type', 'process')})")
        fc_parts.append("\n## Flow Edges\n")
        for edge in mermaid_data.get("edges", []):
            label = f" [{edge['label']}]" if edge.get("label") else ""
            fc_parts.append(f"- {edge['from']} → {edge['to']}{label}")
        fc_parts.append("\n## Mermaid Source\n")
        fc_parts.append("```mermaid")
        fc_parts.append(mermaid_data.get("raw_content", ""))
        fc_parts.append("```")
        chunks["flowchart_records.md"] = "\n".join(fc_parts)
    
    return chunks


# ==============================================================================
# QUALITY REPORT
# ==============================================================================

def generate_quality_report(sheets_data: list[dict], mermaid_data: dict,
                           vlm_calls: int, vlm_validations: list[dict]) -> tuple[dict, str]:
    """Generate quality report in JSON and markdown format."""
    stats = {
        "workbooks_processed": 0,
        "sheets_processed": len(sheets_data),
        "source_fields_extracted": 0,
        "target_fields_extracted": 0,
        "mappings_extracted": 0,
        "uncertainties": 0,
        "vlm_calls_made": vlm_calls,
        "vlm_validations": len(vlm_validations),
        "mermaid_nodes": mermaid_data["node_count"] if mermaid_data else 0,
        "mermaid_edges": mermaid_data["edge_count"] if mermaid_data else 0,
    }
    
    all_uncertainties = []
    confidence_dist = {"high_0.8_to_1.0": 0, "medium_0.5_to_0.8": 0, "low_below_0.5": 0}
    
    for sd in sheets_data:
        extraction = sd.get("extraction", {})
        stats["source_fields_extracted"] += len(extraction.get("source_fields", []))
        stats["target_fields_extracted"] += len(extraction.get("target_fields", []))
        
        for m in extraction.get("mappings", []):
            stats["mappings_extracted"] += 1
            conf = m.get("confidence", 0)
            if conf >= 0.8:
                confidence_dist["high_0.8_to_1.0"] += 1
            elif conf >= 0.5:
                confidence_dist["medium_0.5_to_0.8"] += 1
            else:
                confidence_dist["low_below_0.5"] += 1
        
        for u in extraction.get("uncertainties", []):
            all_uncertainties.append({
                "sheet": sd["sheet_name"],
                **u
            })
    
    stats["uncertainties"] = len(all_uncertainties)
    
    report_json = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "pipeline_version": "v3_vlm",
        "statistics": stats,
        "confidence_distribution": confidence_dist,
        "uncertainties": all_uncertainties,
        "vlm_validations": vlm_validations
    }
    
    # Markdown version
    md_parts = ["# Quality Report — Excel Parse Pipeline v3 (VLM-Enhanced)\n"]
    md_parts.append(f"Generated: {report_json['generated_at']}\n")
    md_parts.append("## Statistics\n")
    md_parts.append("| Metric | Value |")
    md_parts.append("|--------|-------|")
    for k, v in stats.items():
        md_parts.append(f"| {k} | {v} |")
    md_parts.append("")
    md_parts.append("## Confidence Distribution\n")
    for k, v in confidence_dist.items():
        md_parts.append(f"- {k}: {v}")
    md_parts.append("")
    
    if all_uncertainties:
        md_parts.append(f"## Uncertainties ({len(all_uncertainties)})\n")
        for u in all_uncertainties:
            md_parts.append(f"- [{u.get('type', '?')}] {u.get('sheet', '?')} row {u.get('row', '?')}: {u.get('message', u.get('source_ref', ''))}")
    
    if vlm_validations:
        md_parts.append(f"\n## VLM Validations ({len(vlm_validations)})\n")
        for v in vlm_validations:
            status = v.get("validation_status", "?")
            sheet = v.get("sheet", "?")
            md_parts.append(f"- **{sheet}**: {status} — {v.get('notes', '')[:100]}")
    
    return report_json, "\n".join(md_parts)


# ==============================================================================
# MAIN PIPELINE
# ==============================================================================

def run_pipeline(no_s3_sync=False, skip_vlm=False):
    """Run the full VLM-enhanced Excel parse pipeline."""
    
    print("=" * 70)
    print("  EXCEL PARSE PIPELINE v3 — VLM-Enhanced")
    print("=" * 70)
    print(f"  Input:  {INPUT_DIR}")
    print(f"  Output: {OUTPUT_DIR}")
    print(f"  VLM Model: {BEDROCK_VLM_MODEL}")
    print(f"  Skip VLM: {skip_vlm}")
    print("=" * 70)
    
    start_time = time.time()
    vlm_call_count = 0
    vlm_validations = []
    
    # Create output directories
    for subdir in ["markdown", "structured", "graph", "kb_chunks", "images", "parse_plans"]:
        (OUTPUT_DIR / subdir).mkdir(parents=True, exist_ok=True)
    
    # ===== STAGE 1: Source File Inventory =====
    print("\n[Stage 1] Source File Inventory")
    
    input_files = []
    mermaid_files = []
    
    for root, dirs, files in os.walk(INPUT_DIR):
        for fname in files:
            fpath = Path(root) / fname
            rel = fpath.relative_to(INPUT_DIR)
            ext = fpath.suffix.lower()
            
            if ext in (".xlsx", ".xlsm", ".xls"):
                input_files.append({"path": str(fpath), "relative": str(rel), "type": "excel"})
            elif ext in (".mmd", ".mermaid"):
                mermaid_files.append({"path": str(fpath), "relative": str(rel), "type": "mermaid"})
    
    print(f"  Excel files: {len(input_files)}")
    print(f"  Mermaid files: {len(mermaid_files)}")
    
    # Save manifest
    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_dir": str(INPUT_DIR),
        "excel_files": input_files,
        "mermaid_files": mermaid_files
    }
    (OUTPUT_DIR / "source_files_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    
    # ===== STAGE 2: Parse Mermaid Files =====
    print("\n[Stage 2] Parse Mermaid Files")
    
    mermaid_data = None
    if mermaid_files:
        mmd_path = Path(mermaid_files[0]["path"])
        print(f"  Parsing: {mmd_path.name}")
        mermaid_data = parse_mermaid_file(mmd_path)
        print(f"  Nodes: {mermaid_data['node_count']}, Edges: {mermaid_data['edge_count']}")
        
        # Save mermaid parse results
        (OUTPUT_DIR / "structured" / "flow_nodes.jsonl").write_text(
            "\n".join(json.dumps(n, ensure_ascii=False) for n in mermaid_data["nodes"]),
            encoding="utf-8"
        )
        (OUTPUT_DIR / "structured" / "flow_edges.jsonl").write_text(
            "\n".join(json.dumps(e, ensure_ascii=False) for e in mermaid_data["edges"]),
            encoding="utf-8"
        )
    
    # ===== STAGE 3: Process Each Workbook =====
    print("\n[Stage 3] Process Workbooks")
    
    all_sheets_data = []
    workbook_count = 0
    
    for file_info in input_files:
        fpath = Path(file_info["path"])
        workbook_name = fpath.name
        print(f"\n  --- Processing: {workbook_name} ---")
        
        wb = openpyxl.load_workbook(str(fpath), data_only=True, read_only=False)
        workbook_count += 1
        
        for ws_idx, ws_name in enumerate(wb.sheetnames):
            ws = wb[ws_name]
            print(f"\n    Sheet {ws_idx}: {ws_name} ({ws.max_row}r × {ws.max_column}c)")
            
            sheet_data = {
                "workbook": workbook_name,
                "sheet_name": ws_name,
                "sheet_index": ws_idx,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_cells": len(ws.merged_cells.ranges),
                "classification": {},
                "extraction": {},
                "vlm_validation": None,
                "markdown": ""
            }
            
            # ----- 3a: Generate cell summary for classification -----
            cell_summary_lines = []
            for r in range(1, min(25, ws.max_row + 1)):
                row_vals = []
                for c in range(1, min(40, ws.max_column + 1)):
                    v = ws.cell(r, c).value
                    if v is not None:
                        row_vals.append(f"{get_column_letter(c)}{r}={str(v)[:30]}")
                if row_vals:
                    cell_summary_lines.append(f"  Row {r}: " + " | ".join(row_vals[:8]))
            cell_summary = "\n".join(cell_summary_lines[:20])
            
            # ----- 3b: VLM Classification -----
            if not skip_vlm and ws.max_row > 0 and ws.max_column > 0:
                # First try heuristic classification
                heuristic_class = heuristic_classify_sheet(ws_name, ws)
                
                # Only call VLM if heuristic couldn't classify
                if heuristic_class.get("sheet_type") == "unknown":
                    print(f"      Rendering overview image...")
                    try:
                        overview_img = render_sheet_overview(ws)
                        # Save overview image
                        safe_sheet = re.sub(r'[^\w\-]', '_', ws_name)[:50]
                        img_path = OUTPUT_DIR / "images" / f"{safe_sheet}_overview.png"
                        img_path.write_bytes(overview_img)
                        
                        print(f"      Calling VLM for classification...")
                        classification = classify_sheet_with_vlm(ws_name, overview_img, cell_summary)
                        vlm_call_count += 1
                        time.sleep(VLM_DELAY_BETWEEN_CALLS)
                    except Exception as e:
                        print(f"      [WARNING] VLM classification failed: {e}")
                        classification = heuristic_class
                else:
                    # Heuristic is confident — still render image for reference
                    classification = heuristic_class
                    try:
                        safe_sheet = re.sub(r'[^\w\-]', '_', ws_name)[:50]
                        img_path = OUTPUT_DIR / "images" / f"{safe_sheet}_overview.png"
                        overview_img = render_sheet_overview(ws)
                        img_path.write_bytes(overview_img)
                    except Exception:
                        pass
            else:
                # Heuristic classification without VLM
                classification = heuristic_classify_sheet(ws_name, ws)
            
            sheet_data["classification"] = classification
            sheet_type = classification.get("sheet_type", "unknown")
            print(f"      Classification: {sheet_type}")
            
            # ----- 3c: Deterministic Extraction -----
            extraction = {}
            if sheet_type == "mapping_spec" or (ws.max_column > 100 and "マッピング" in ws_name):
                print(f"      Extracting mapping table...")
                extraction = extract_mapping_sheet_deterministic(ws)
                sf = len(extraction.get("source_fields", []))
                tf = len(extraction.get("target_fields", []))
                mp = len(extraction.get("mappings", []))
                uc = len(extraction.get("uncertainties", []))
                print(f"      → Source: {sf}, Target: {tf}, Mappings: {mp}, Uncertainties: {uc}")
            
            sheet_data["extraction"] = extraction
            
            # ----- 3d: VLM Validation (for first 3 mapping sheets only) -----
            mapping_sheet_count = sum(1 for s in all_sheets_data if s.get("classification", {}).get("sheet_type") == "mapping_spec")
            if (not skip_vlm and sheet_type == "mapping_spec" and 
                extraction.get("source_fields") and mapping_sheet_count < 3):
                print(f"      VLM validation of extraction...")
                try:
                    # Render header area tile for validation
                    header_tile = render_sheet_to_image(
                        ws, start_row=19, end_row=min(30, ws.max_row),
                        start_col=1, end_col=min(30, ws.max_column)
                    )
                    validation = vlm_validate_mapping_extraction(
                        ws_name, header_tile, cell_summary, extraction
                    )
                    vlm_call_count += 1
                    validation["sheet"] = ws_name
                    vlm_validations.append(validation)
                    sheet_data["vlm_validation"] = validation
                    print(f"      Validation: {validation.get('validation_status', '?')}")
                    time.sleep(VLM_DELAY_BETWEEN_CALLS)
                except Exception as e:
                    print(f"      [WARNING] VLM validation failed: {e}")
            
            # ----- 3e: Generate Markdown -----
            print(f"      Generating markdown...")
            md = generate_sheet_markdown(ws, ws_name, classification, extraction)
            sheet_data["markdown"] = md
            
            # Save individual markdown
            safe_wb = re.sub(r'[^\w\-]', '_', workbook_name)[:60]
            safe_sh = re.sub(r'[^\w\-]', '_', ws_name)[:50]
            md_filename = f"{safe_wb}__{safe_sh}.md"
            (OUTPUT_DIR / "markdown" / md_filename).write_text(md, encoding="utf-8")
            
            all_sheets_data.append(sheet_data)
        
        wb.close()
        
        # Save combined markdown for this workbook
        combined_md = "\n\n---\n\n".join(
            sd["markdown"] for sd in all_sheets_data
            if sd["workbook"] == workbook_name and sd["markdown"]
        )
        sanitized_wb = re.sub(r'[^\w\-]', '_', workbook_name)[:60]
        combined_name = f"_combined_{sanitized_wb}.md"
        (OUTPUT_DIR / "markdown" / combined_name).write_text(combined_md, encoding="utf-8")
    
    # ===== STAGE 4: Save Structured Outputs =====
    print("\n\n[Stage 4] Save Structured Outputs")
    
    # Source fields
    all_source_fields = []
    all_target_fields = []
    all_mappings = []
    all_uncertainties = []
    
    for sd in all_sheets_data:
        extraction = sd.get("extraction", {})
        for f in extraction.get("source_fields", []):
            all_source_fields.append({"workbook": sd["workbook"], "sheet": sd["sheet_name"], **f})
        for f in extraction.get("target_fields", []):
            all_target_fields.append({"workbook": sd["workbook"], "sheet": sd["sheet_name"], **f})
        for m in extraction.get("mappings", []):
            all_mappings.append({"workbook": sd["workbook"], "sheet": sd["sheet_name"], **m})
        for u in extraction.get("uncertainties", []):
            all_uncertainties.append({"workbook": sd["workbook"], "sheet": sd["sheet_name"], **u})
    
    def save_jsonl(data, filename):
        path = OUTPUT_DIR / "structured" / filename
        path.write_text(
            "\n".join(json.dumps(item, ensure_ascii=False) for item in data),
            encoding="utf-8"
        )
        return len(data)
    
    sf_count = save_jsonl(all_source_fields, "source_fields.jsonl")
    tf_count = save_jsonl(all_target_fields, "target_fields.jsonl")
    mp_count = save_jsonl(all_mappings, "mappings.jsonl")
    uc_count = save_jsonl(all_uncertainties, "uncertainties.jsonl")
    
    print(f"  source_fields.jsonl: {sf_count}")
    print(f"  target_fields.jsonl: {tf_count}")
    print(f"  mappings.jsonl: {mp_count}")
    print(f"  uncertainties.jsonl: {uc_count}")
    
    # Parse plans (save sheet classifications)
    parse_plans = []
    for sd in all_sheets_data:
        parse_plans.append({
            "workbook": sd["workbook"],
            "sheet_name": sd["sheet_name"],
            "sheet_index": sd["sheet_index"],
            "classification": sd["classification"],
            "dimensions": {"rows": sd["max_row"], "cols": sd["max_column"]},
            "merged_cells": sd["merged_cells"]
        })
    
    (OUTPUT_DIR / "parse_plans" / "sheet_classifications.jsonl").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False) for p in parse_plans),
        encoding="utf-8"
    )
    
    # ===== STAGE 5: Build Graph =====
    print("\n[Stage 5] Build Graph Nodes & Edges")
    
    # Group sheets by workbook
    wb_sheets = {}
    for sd in all_sheets_data:
        wb_sheets.setdefault(sd["workbook"], []).append(sd)
    
    all_nodes = []
    all_edges = []
    for wb_name, sheets in wb_sheets.items():
        nodes, edges = build_graph_nodes_edges(wb_name, sheets, mermaid_data)
        all_nodes.extend(nodes)
        all_edges.extend(edges)
    
    (OUTPUT_DIR / "graph" / "nodes.jsonl").write_text(
        "\n".join(json.dumps(n, ensure_ascii=False) for n in all_nodes),
        encoding="utf-8"
    )
    (OUTPUT_DIR / "graph" / "edges.jsonl").write_text(
        "\n".join(json.dumps(e, ensure_ascii=False) for e in all_edges),
        encoding="utf-8"
    )
    print(f"  Nodes: {len(all_nodes)}, Edges: {len(all_edges)}")
    
    # ===== STAGE 6: Build KB Chunks =====
    print("\n[Stage 6] Build KB Chunks")
    
    for wb_name, sheets in wb_sheets.items():
        chunks = build_kb_chunks(wb_name, sheets, mermaid_data)
        for chunk_name, content in chunks.items():
            (OUTPUT_DIR / "kb_chunks" / chunk_name).write_text(content, encoding="utf-8")
        print(f"  {wb_name}: {len(chunks)} chunks")
    
    # ===== STAGE 7: Quality Report =====
    print("\n[Stage 7] Generate Quality Report")
    
    report_json, report_md = generate_quality_report(
        all_sheets_data, mermaid_data, vlm_call_count, vlm_validations
    )
    report_json["statistics"]["workbooks_processed"] = workbook_count
    
    (OUTPUT_DIR / "quality_report.json").write_text(
        json.dumps(report_json, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUTPUT_DIR / "quality_report.md").write_text(report_md, encoding="utf-8")
    
    # ===== STAGE 8: S3 Sync =====
    if not no_s3_sync:
        print("\n[Stage 8] S3 Sync")
        try:
            count = s3_sync_outputs(OUTPUT_DIR, S3_OUTPUT_PREFIX)
            print(f"  Uploaded {count} files to s3://{S3_BUCKET}/{S3_OUTPUT_PREFIX}")
        except Exception as e:
            print(f"  [S3 SYNC ERROR] {e}")
    else:
        print("\n[Stage 8] S3 Sync — SKIPPED (--no-s3-sync)")
    
    # ===== DONE =====
    elapsed = time.time() - start_time
    print("\n" + "=" * 70)
    print(f"  PIPELINE COMPLETE in {elapsed:.1f}s")
    print(f"  Sheets processed: {len(all_sheets_data)}")
    print(f"  VLM calls: {vlm_call_count}")
    print(f"  Source fields: {sf_count}")
    print(f"  Target fields: {tf_count}")
    print(f"  Mappings: {mp_count}")
    print(f"  Uncertainties: {uc_count}")
    print(f"  Graph: {len(all_nodes)} nodes, {len(all_edges)} edges")
    print(f"  Output: {OUTPUT_DIR}")
    print("=" * 70)


def heuristic_classify_sheet(ws_name: str, ws) -> dict:
    """Classify sheet without VLM using name-based heuristics."""
    name_lower = ws_name.lower()
    
    if "マッピング" in ws_name and ws.max_column > 50:
        return {
            "sheet_type": "mapping_spec",
            "description": "Wide mapping sheet with source/target field tables",
            "regions": [],
            "has_mapping_table": True,
            "has_metadata_section": True,
            "key_observations": ["Classified by sheet name pattern + wide column count"]
        }
    elif "データ取得条件" in ws_name:
        return {
            "sheet_type": "data_retrieval_condition",
            "description": "Data retrieval/query condition specification",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": True,
            "key_observations": []
        }
    elif "フローチャート" in ws_name:
        return {
            "sheet_type": "flowchart",
            "description": "Process flowchart (Excel shapes)",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": False,
            "key_observations": ["Content is in Excel shapes, not cells"]
        }
    elif "変更履歴" in ws_name:
        return {
            "sheet_type": "change_history",
            "description": "Document change history/revision log",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": False,
            "key_observations": []
        }
    elif "API呼出" in ws_name:
        return {
            "sheet_type": "api_call_sequence",
            "description": "API call sequence specification",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": True,
            "key_observations": []
        }
    elif "DataSpider" in ws_name or "補足" in ws_name:
        return {
            "sheet_type": "development_spec",
            "description": "Development specification or supplementary notes",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": False,
            "key_observations": []
        }
    elif "概要" in ws_name:
        return {
            "sheet_type": "overview",
            "description": "Document overview/summary",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": False,
            "key_observations": []
        }
    else:
        return {
            "sheet_type": "unknown",
            "description": "Could not classify by name heuristic",
            "regions": [],
            "has_mapping_table": False,
            "has_metadata_section": False,
            "key_observations": []
        }


# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Excel Parse Pipeline v3 (VLM-Enhanced)")
    parser.add_argument("--no-s3-sync", action="store_true", help="Skip S3 sync")
    parser.add_argument("--skip-vlm", action="store_true", help="Skip VLM calls (use heuristics only)")
    
    args = parser.parse_args()
    run_pipeline(no_s3_sync=args.no_s3_sync, skip_vlm=args.skip_vlm)
