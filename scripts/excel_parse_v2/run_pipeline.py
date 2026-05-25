#!/usr/bin/env python3
"""
Excel Parse Pipeline v2 - Reliable extraction of Japanese enterprise design documents.

Key design principles:
1. Every sheet is dumped as readable markdown (minimum requirement)
2. Wide mapping sheets are recognized and parsed as source+target pairs
3. Manual Mermaid files are authoritative for flowcharts
4. Structured JSONL for graph/vector KB preparation
5. Uncertainty is marked, not silently dropped

Usage:
    python scripts/excel_parse_v2/run_pipeline.py [--no-s3-sync] [--no-llm]
"""
import sys
import os
import json
import time
import logging
import argparse
import hashlib
from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Optional

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

# Load .env
from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env")

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import boto3

logger = logging.getLogger("excel_parse_v2")

# ============================================================================
# Configuration
# ============================================================================

class Config:
    S3_BUCKET = os.getenv("S3_BUCKET", "s3-hulftchina-rd")
    S3_INPUT_PREFIX = os.getenv("S3_RAW_PREFIX", "サンプル20260519")
    S3_OUTPUT_PREFIX = "output/sample_20260519/excel_parse_pipeline_v2"
    AWS_REGION = os.getenv("AWS_REGION", "ap-northeast-1")
    
    LOCAL_INPUT_DIR = PROJECT_ROOT / "data" / "input" / "sample_20260519"
    OUTPUT_DIR = PROJECT_ROOT / "data" / "outputs" / "excel_parse_pipeline_v2" / "sample_20260519"
    
    # LLM settings
    BEDROCK_MODEL_ID = os.getenv("BEDROCK_TEXT_MODEL_ID", "jp.anthropic.claude-sonnet-4-6")
    USE_LLM = True
    USE_VLM = False  # VLM not needed for this pipeline
    
    # Extraction limits
    MAX_ROWS_PER_SHEET = 500
    MAX_COLS_PER_SHEET = 250


# ============================================================================
# S3 I/O
# ============================================================================

class S3IO:
    def __init__(self, config: Config):
        self.config = config
        self.s3 = boto3.client("s3", region_name=config.AWS_REGION)
    
    def list_files(self, prefix: str) -> list[dict]:
        """List all files under prefix."""
        results = []
        paginator = self.s3.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=self.config.S3_BUCKET, Prefix=prefix):
            for obj in page.get("Contents", []):
                if obj["Size"] > 0:
                    results.append({
                        "key": obj["Key"],
                        "size": obj["Size"],
                        "last_modified": obj["LastModified"].isoformat(),
                    })
        return results
    
    def download_file(self, key: str, local_path: Path):
        """Download a single file from S3."""
        local_path.parent.mkdir(parents=True, exist_ok=True)
        self.s3.download_file(self.config.S3_BUCKET, key, str(local_path))
    
    def sync_upload(self, local_dir: Path, prefix: str):
        """Upload all files from local_dir to S3 prefix."""
        uploaded = 0
        for fpath in local_dir.rglob("*"):
            if fpath.is_file():
                relative = fpath.relative_to(local_dir)
                key = f"{prefix}/{relative}"
                content_type = self._guess_content_type(fpath)
                self.s3.upload_file(
                    str(fpath), self.config.S3_BUCKET, key,
                    ExtraArgs={"ContentType": content_type}
                )
                uploaded += 1
        return uploaded
    
    def _guess_content_type(self, path: Path) -> str:
        ext = path.suffix.lower()
        return {
            ".json": "application/json",
            ".jsonl": "application/jsonl",
            ".md": "text/markdown",
            ".html": "text/html",
            ".csv": "text/csv",
            ".mmd": "text/plain",
        }.get(ext, "application/octet-stream")


# ============================================================================
# Source Scanner
# ============================================================================

def scan_and_download(config: Config) -> dict:
    """Scan S3 and ensure all files are available locally."""
    s3io = S3IO(config)
    
    # List all files
    all_files = s3io.list_files(config.S3_INPUT_PREFIX + "/")
    
    manifest = {
        "scan_timestamp": datetime.now(timezone.utc).isoformat(),
        "s3_prefix": f"s3://{config.S3_BUCKET}/{config.S3_INPUT_PREFIX}/",
        "excel_files": [],
        "mermaid_files": [],
        "other_files": [],
    }
    
    for f in all_files:
        key = f["key"]
        ext = Path(key).suffix.lower()
        
        # Determine local path
        relative = key[len(config.S3_INPUT_PREFIX) + 1:]  # strip prefix/
        local_path = config.LOCAL_INPUT_DIR / relative
        
        entry = {
            "key": key,
            "local_path": str(local_path),
            "size": f["size"],
            "last_modified": f["last_modified"],
            "filename": Path(key).name,
        }
        
        if ext in (".xlsx", ".xlsm", ".xls"):
            manifest["excel_files"].append(entry)
        elif ext in (".mmd", ".mermaid"):
            manifest["mermaid_files"].append(entry)
        else:
            manifest["other_files"].append(entry)
        
        # Download if not present or size differs
        if not local_path.exists() or local_path.stat().st_size != f["size"]:
            logger.info(f"  Downloading: {key}")
            s3io.download_file(key, local_path)
    
    manifest["scan_summary"] = {
        "excel_files": len(manifest["excel_files"]),
        "mermaid_files": len(manifest["mermaid_files"]),
        "other_files": len(manifest["other_files"]),
        "total_size_bytes": sum(f["size"] for f in all_files),
    }
    
    return manifest


# ============================================================================
# Workbook Atlas Builder - structural facts only
# ============================================================================

def build_sheet_atlas(ws, sheet_name: str, config: Config) -> dict:
    """Extract structural facts from a worksheet."""
    max_row = min(ws.max_row or 1, config.MAX_ROWS_PER_SHEET)
    max_col = min(ws.max_column or 1, config.MAX_COLS_PER_SHEET)
    
    # Merged cells
    merged = [str(mc) for mc in ws.merged_cells.ranges]
    
    # Collect all non-empty cells (for atlas, limit to first 200 rows)
    cells = {}
    for row in range(1, min(max_row + 1, 201)):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if cell.value is not None:
                coord = f"{get_column_letter(col)}{row}"
                cells[coord] = {
                    "value": str(cell.value)[:200],  # truncate long values
                    "row": row,
                    "col": col,
                }
    
    # Detect occupied rows and columns
    occupied_rows = set()
    occupied_cols = set()
    for info in cells.values():
        occupied_rows.add(info["row"])
        occupied_cols.add(info["col"])
    
    return {
        "sheet_name": sheet_name,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cell_count": len(merged),
        "merged_cells": merged[:50],  # limit for large sheets
        "non_empty_cell_count": len(cells),
        "occupied_row_count": len(occupied_rows),
        "occupied_col_count": len(occupied_cols),
        "sample_cells": cells,  # limited to first 200 rows
    }


# ============================================================================
# Sheet Type Classifier
# ============================================================================

def classify_sheet_type(sheet_name: str, atlas: dict) -> dict:
    """Classify sheet type based on name patterns and structure."""
    name_lower = sheet_name.lower()
    
    # Name-based heuristics (very reliable for Japanese enterprise docs)
    if "マッピング" in sheet_name or "mapping" in name_lower:
        if "データ取得条件" in sheet_name:
            return {"type": "data_retrieval_condition", "confidence": 0.95}
        return {"type": "mapping_sheet", "confidence": 0.95}
    elif "変更履歴" in sheet_name or "履歴" in sheet_name:
        return {"type": "change_history", "confidence": 0.95}
    elif "フローチャート" in sheet_name or "flowchart" in name_lower:
        return {"type": "flowchart", "confidence": 0.95}
    elif "概要" in sheet_name or "overview" in name_lower:
        return {"type": "overview", "confidence": 0.95}
    elif "API" in sheet_name or "呼出" in sheet_name:
        return {"type": "api_call_sequence", "confidence": 0.90}
    elif "DataSpider" in sheet_name or "開発仕様" in sheet_name:
        return {"type": "development_spec", "confidence": 0.90}
    elif "補足" in sheet_name:
        return {"type": "supplementary_notes", "confidence": 0.90}
    elif "データ取得条件" in sheet_name:
        return {"type": "data_retrieval_condition", "confidence": 0.90}
    
    # Structure-based fallback
    max_col = atlas.get("max_column", 1)
    max_row = atlas.get("max_row", 1)
    
    if max_col > 100:
        return {"type": "mapping_sheet", "confidence": 0.70}
    elif max_row < 5:
        return {"type": "supplementary_notes", "confidence": 0.60}
    elif max_col < 10 and max_row > 20:
        return {"type": "list_table", "confidence": 0.60}
    
    return {"type": "unknown", "confidence": 0.3}


# ============================================================================
# Markdown Dump - THE CORE reliable extraction
# ============================================================================

def sheet_to_markdown(ws, sheet_name: str, wb_name: str, config: Config) -> str:
    """Convert a worksheet to a readable markdown representation.
    
    This is the MINIMUM REQUIREMENT - every sheet gets a readable markdown dump.
    Handles:
    - Wide tables by splitting into logical groups
    - Merged cells
    - Multi-region sheets
    """
    max_row = min(ws.max_row or 1, config.MAX_ROWS_PER_SHEET)
    max_col = min(ws.max_column or 1, config.MAX_COLS_PER_SHEET)
    
    lines = []
    lines.append(f"# {sheet_name}")
    lines.append(f"")
    lines.append(f"**Workbook:** {wb_name}")
    lines.append(f"**Dimensions:** {max_row} rows × {max_col} columns")
    lines.append(f"**Merged cells:** {len(ws.merged_cells.ranges)}")
    lines.append(f"")
    
    if max_col <= 1 and max_row <= 1:
        lines.append("*Empty or minimal sheet*")
        return "\n".join(lines)
    
    # For very wide sheets (mapping sheets), split into logical column groups
    if max_col > 30:
        return _wide_sheet_to_markdown(ws, sheet_name, wb_name, max_row, max_col, lines)
    else:
        return _normal_sheet_to_markdown(ws, sheet_name, max_row, max_col, lines)


def _normal_sheet_to_markdown(ws, sheet_name: str, max_row: int, max_col: int, lines: list) -> str:
    """Render a normal-width sheet as a markdown table."""
    # Find actual used range
    first_row, last_row = max_row, 1
    first_col, last_col = max_col, 1
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if ws.cell(row, col).value is not None:
                first_row = min(first_row, row)
                last_row = max(last_row, row)
                first_col = min(first_col, col)
                last_col = max(last_col, col)
    
    if first_row > last_row:
        lines.append("*No data found*")
        return "\n".join(lines)
    
    # Render as markdown table
    lines.append(f"## Data (rows {first_row}-{last_row}, cols {get_column_letter(first_col)}-{get_column_letter(last_col)})")
    lines.append("")
    
    # Build table
    headers = []
    for col in range(first_col, last_col + 1):
        headers.append(get_column_letter(col))
    
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    
    for row in range(first_row, last_row + 1):
        cells = []
        for col in range(first_col, last_col + 1):
            v = ws.cell(row, col).value
            if v is None:
                cells.append("")
            else:
                # Clean for markdown table
                s = str(v).replace("|", "\\|").replace("\n", "<br/>")[:100]
                cells.append(s)
        lines.append("| " + " | ".join(cells) + " |")
    
    return "\n".join(lines)


def _wide_sheet_to_markdown(ws, sheet_name: str, wb_name: str, max_row: int, max_col: int, lines: list) -> str:
    """Render a wide mapping sheet by detecting column groups.
    
    Japanese enterprise mapping sheets typically have:
    - Metadata header (rows 1-9): document name, IF-ID, etc.
    - Configuration section (rows 10-20): source/target config 
    - Column headers (row 21 or similar): field definitions
    - Data rows (row 22+): actual field mappings
    
    The sheet has a LEFT side (source system fields) and RIGHT side (target system fields).
    """
    # Step 1: Detect column groups by finding header rows
    # Look for the main header row (typically has the most non-empty cells in a single row)
    header_candidates = {}
    for row in range(1, min(max_row + 1, 30)):
        count = 0
        for col in range(1, max_col + 1):
            if ws.cell(row, col).value is not None:
                count += 1
        if count >= 3:
            header_candidates[row] = count
    
    # Step 2: Find column groups - detect large gaps between occupied columns
    # Use the row with most headers to identify groups
    if not header_candidates:
        # Fallback: just dump rows
        return _dump_rows_markdown(ws, sheet_name, max_row, max_col, lines)
    
    # Find all occupied columns across rows 1-30
    col_occupation = {}  # col -> list of non-empty row values
    for col in range(1, max_col + 1):
        values = []
        for row in range(1, min(max_row + 1, 30)):
            v = ws.cell(row, col).value
            if v is not None:
                values.append((row, str(v)[:80]))
        if values:
            col_occupation[col] = values
    
    # Detect column groups by finding gaps (>3 consecutive empty columns)
    occupied_cols = sorted(col_occupation.keys())
    if not occupied_cols:
        lines.append("*No data found*")
        return "\n".join(lines)
    
    groups = []
    current_group = [occupied_cols[0]]
    
    for i in range(1, len(occupied_cols)):
        gap = occupied_cols[i] - occupied_cols[i-1]
        if gap > 4:  # More than 4 empty columns = new group
            groups.append(current_group)
            current_group = [occupied_cols[i]]
        else:
            current_group.append(occupied_cols[i])
    groups.append(current_group)
    
    # Step 3: Render each column group
    lines.append(f"## Sheet Layout Analysis")
    lines.append(f"")
    lines.append(f"This is a wide sheet with **{len(groups)} column groups** detected:")
    lines.append(f"")
    
    for i, group in enumerate(groups):
        start_col = group[0]
        end_col = group[-1]
        lines.append(f"- Group {i+1}: Columns {get_column_letter(start_col)}-{get_column_letter(end_col)} ({end_col - start_col + 1} cols)")
    lines.append("")
    
    # Step 4: Render metadata section (rows 1-20) as key-value pairs
    lines.append("## Metadata Section")
    lines.append("")
    
    for row in range(1, min(21, max_row + 1)):
        row_vals = []
        for col in range(1, max_col + 1):
            v = ws.cell(row, col).value
            if v is not None:
                row_vals.append(f"**{get_column_letter(col)}{row}**: {str(v)[:150]}")
        if row_vals:
            lines.append(f"Row {row}: " + " | ".join(row_vals))
    lines.append("")
    
    # Step 5: Render each column group as a separate table
    # Find the actual header row for each group
    for gi, group in enumerate(groups):
        start_col = group[0]
        end_col = group[-1]
        
        # Find header row for this group (row with most non-empty cells in this column range)
        best_header_row = 21  # default for mapping sheets
        best_count = 0
        for row in range(15, min(30, max_row + 1)):
            count = sum(1 for col in range(start_col, end_col + 1) if ws.cell(row, col).value is not None)
            if count > best_count:
                best_count = count
                best_header_row = row
        
        # Get headers
        headers = {}
        for col in range(start_col, end_col + 1):
            v = ws.cell(best_header_row, col).value
            if v is not None:
                headers[col] = str(v)[:50]
        
        if not headers:
            continue
        
        lines.append(f"## Column Group {gi+1}: {get_column_letter(start_col)}-{get_column_letter(end_col)}")
        lines.append(f"")
        lines.append(f"Header row: {best_header_row}")
        lines.append(f"Headers: {json.dumps(headers, ensure_ascii=False)}")
        lines.append(f"")
        
        # Render data as markdown table (limited cols)
        header_cols = sorted(headers.keys())
        if len(header_cols) > 20:
            header_cols = header_cols[:20]
            lines.append(f"*(Showing first 20 of {len(headers)} columns)*")
            lines.append("")
        
        # Table header
        h_labels = [headers.get(c, get_column_letter(c)) for c in header_cols]
        lines.append("| " + " | ".join(h_labels) + " |")
        lines.append("| " + " | ".join(["---"] * len(header_cols)) + " |")
        
        # Data rows
        data_start = best_header_row + 1
        data_end = min(max_row, data_start + 200)  # limit output
        
        row_count = 0
        for row in range(data_start, data_end + 1):
            cells = []
            has_data = False
            for col in header_cols:
                v = ws.cell(row, col).value
                if v is None:
                    cells.append("")
                else:
                    has_data = True
                    s = str(v).replace("|", "\\|").replace("\n", "<br/>")[:80]
                    cells.append(s)
            if has_data:
                lines.append("| " + " | ".join(cells) + " |")
                row_count += 1
        
        lines.append(f"")
        lines.append(f"*({row_count} data rows)*")
        lines.append("")
    
    return "\n".join(lines)


def _dump_rows_markdown(ws, sheet_name: str, max_row: int, max_col: int, lines: list) -> str:
    """Fallback: dump all non-empty rows as plain text."""
    lines.append("## Raw Data Dump")
    lines.append("")
    
    for row in range(1, max_row + 1):
        cells = []
        for col in range(1, max_col + 1):
            v = ws.cell(row, col).value
            if v is not None:
                cells.append(f"{get_column_letter(col)}={str(v)[:100]}")
        if cells:
            lines.append(f"**Row {row}:** {' | '.join(cells)}")
    
    return "\n".join(lines)


# ============================================================================
# Mapping Sheet Parser - Extract structured source/target/mapping data
# ============================================================================

def parse_mapping_sheet(ws, sheet_name: str, wb_name: str, config: Config) -> dict:
    """Parse a mapping sheet and extract structured field/mapping data.
    
    Japanese enterprise mapping sheets follow a pattern:
    - Top section: metadata (文書名, IF-ID, etc.)
    - Middle: configuration (送信元, 形式, テーブル名, etc.)
    - Two side-by-side tables: source fields (left) and target fields (right)
    - The target table has a "マッピング元" column referencing source row numbers
    """
    max_row = min(ws.max_row or 1, config.MAX_ROWS_PER_SHEET)
    max_col = min(ws.max_column or 1, config.MAX_COLS_PER_SHEET)
    
    result = {
        "sheet_name": sheet_name,
        "workbook": wb_name,
        "sheet_type": "mapping_sheet",
        "metadata": {},
        "source_config": {},
        "target_config": {},
        "source_fields": [],
        "target_fields": [],
        "mappings": [],
        "uncertainties": [],
    }
    
    # Step 1: Extract metadata from rows 1-9
    for row in range(1, 10):
        for col in range(1, max_col + 1):
            v = ws.cell(row, col).value
            if v is not None:
                key = str(v).strip()
                # Look for value in next columns
                for vcol in range(col + 1, min(col + 10, max_col + 1)):
                    val = ws.cell(row, vcol).value
                    if val is not None:
                        result["metadata"][key] = str(val).strip()
                        break
    
    # Step 2: Find source and target configuration sections (rows 10-20)
    for row in range(10, min(21, max_row + 1)):
        for col in range(1, max_col + 1):
            v = ws.cell(row, col).value
            if v is not None:
                key = str(v).strip()
                # Look for value in the next few columns
                for vcol in range(col + 1, min(col + 10, max_col + 1)):
                    val = ws.cell(row, vcol).value
                    if val is not None:
                        if col < 60:  # Left side = source config
                            result["source_config"][key] = str(val).strip()
                        else:  # Right side = target config
                            result["target_config"][key] = str(val).strip()
                        break
    
    # Step 3: Detect the main header rows
    # The header is typically around row 21-22 for these sheets
    # Source side headers are in one row, target side headers may be in the next row
    source_header_row = None
    target_header_row = None
    
    for row in range(15, min(30, max_row + 1)):
        vals_left = []
        vals_right = []
        for col in range(1, max_col + 1):
            v = ws.cell(row, col).value
            if v is not None:
                if col <= 60:
                    vals_left.append((col, str(v).strip()))
                else:
                    vals_right.append((col, str(v).strip()))
        
        # Source header typically starts with "No" in column A
        if vals_left and any("No" == v[1] for v in vals_left):
            if source_header_row is None:
                source_header_row = row
        
        # Target header row has "No" on the right side
        if vals_right and any("No" == v[1] for v in vals_right):
            if target_header_row is None:
                target_header_row = row
    
    if source_header_row is None:
        # Try another pattern: look for 項目名称
        for row in range(15, min(30, max_row + 1)):
            for col in range(1, 30):
                v = ws.cell(row, col).value
                if v is not None and "項目名称" in str(v):
                    source_header_row = row
                    break
            if source_header_row:
                break
    
    if source_header_row is None:
        result["uncertainties"].append({
            "type": "header_detection_failed",
            "detail": "Could not detect source header row",
            "severity": "high",
        })
        return result
    
    # Step 4: Parse header columns
    # Strategy: find the "No" column for both source and target tables
    # Source "No" is always in column A (col 1)
    # Target "No" is the second occurrence of "No" in the header row(s)
    source_headers = {}  # col -> header_name
    target_headers = {}  # col -> header_name
    
    # Find all "No" columns in header rows to determine source/target boundary
    no_columns = []
    for check_row in range(source_header_row, min(source_header_row + 3, max_row + 1)):
        for col in range(1, max_col + 1):
            v = ws.cell(check_row, col).value
            if v is not None and str(v).strip() == "No":
                no_columns.append((col, check_row))
    
    # Sort by column to ensure source (col A) comes first, target (col BK/BO) second
    no_columns.sort(key=lambda x: x[0])
    
    # Deduplicate: keep only distinct column positions (same col in multiple rows = same table)
    seen_cols = set()
    unique_no_columns = []
    for col, row in no_columns:
        if col not in seen_cols:
            seen_cols.add(col)
            unique_no_columns.append((col, row))
    no_columns = unique_no_columns
    
    # The boundary between source and target is halfway between first and second "No"
    if len(no_columns) >= 2:
        src_no_boundary = no_columns[0][0]
        tgt_no_start = no_columns[1][0]
        boundary_col = (src_no_boundary + tgt_no_start) // 2
        
        # Target header row from the target "No" position
        if target_header_row is None:
            target_header_row = no_columns[1][1]
        # Also update source_header_row if it came from target detection
        source_header_row = no_columns[0][1]
    else:
        # Fallback: use column 60 as boundary
        boundary_col = 60
    
    # Collect source headers from source_header_row (left of boundary)
    for col in range(1, boundary_col):
        v = ws.cell(source_header_row, col).value
        if v is not None:
            header_name = str(v).strip()
            if header_name:
                source_headers[col] = header_name
    
    # Collect target headers from target_header_row (or source_header_row if same)
    tgt_row = target_header_row if target_header_row else source_header_row
    for col in range(boundary_col, max_col + 1):
        v = ws.cell(tgt_row, col).value
        if v is not None:
            header_name = str(v).strip()
            if header_name:
                target_headers[col] = header_name
    
    # If source and target have different header rows, also check the other row for the target
    if target_header_row and target_header_row != source_header_row:
        # Check if source_header_row has additional target headers not in target_header_row
        for col in range(boundary_col, max_col + 1):
            v = ws.cell(source_header_row, col).value
            if v is not None and col not in target_headers:
                header_name = str(v).strip()
                if header_name:
                    target_headers[col] = header_name
    
    # Also check rows after header for additional target columns (like マッピング元, 編集内容)
    # These are sometimes in the data row (row after header) acting as sub-headers
    if target_header_row:
        extra_row = target_header_row + 1
        for col in range(boundary_col, max_col + 1):
            if col not in target_headers:
                v = ws.cell(extra_row, col).value
                # Only add if it looks like a header (contains Japanese/English text, not a number)
                if v is not None:
                    sv = str(v).strip()
                    if sv and not sv.isdigit() and len(sv) > 1:
                        # Check if this column has mostly text (headers) or numbers (data)
                        # Look at the same column a few rows down
                        check_v = ws.cell(extra_row + 3, col).value
                        if check_v is None or not str(check_v).strip().isdigit():
                            pass  # Might be header-like but skip for safety
    
    result["_source_headers"] = {get_column_letter(k): v for k, v in source_headers.items()}
    result["_target_headers"] = {get_column_letter(k): v for k, v in target_headers.items()}
    result["_source_header_row"] = source_header_row
    result["_target_header_row"] = target_header_row
    
    # Step 5: Identify key columns
    source_no_col = None
    source_name_col = None
    source_type_col = None
    source_length_col = None
    
    for col, name in source_headers.items():
        if name == "No":
            source_no_col = col
        elif "項目名称" in name:
            source_name_col = col
        elif name == "Type" or "タイプ" in name:
            source_type_col = col
        elif "長さ" in name:
            source_length_col = col
    
    target_no_col = None
    target_name_col = None
    target_variable_col = None  # 変数
    target_type_col = None
    target_required_col = None
    target_length_col = None
    target_mapping_source_col = None  # マッピング元
    target_mapping_detail_col = None  # 編集内容
    target_remark_col = None  # 備考
    
    for col, name in target_headers.items():
        if name == "No":
            target_no_col = col
        elif "項目名称" in name:
            target_name_col = col
        elif name == "変数":
            target_variable_col = col
        elif name == "Type" or "タイプ" in name:
            target_type_col = col
        elif "必須" in name:
            target_required_col = col
        elif "長さ" in name:
            target_length_col = col
        elif "マッピング元" in name:
            target_mapping_source_col = col
        elif "編集内容" in name and "DataSpider" not in name:
            target_mapping_detail_col = col
        elif "備考" in name:
            target_remark_col = col
    
    # Step 6: Extract source fields
    data_start = max(source_header_row, target_header_row or source_header_row) + 1
    
    source_field_map = {}  # no -> field dict
    
    for row in range(data_start, max_row + 1):
        # Check if row has data
        no_val = ws.cell(row, source_no_col).value if source_no_col else None
        name_val = ws.cell(row, source_name_col).value if source_name_col else None
        
        if no_val is None and name_val is None:
            continue
        
        field = {
            "no": str(no_val) if no_val is not None else None,
            "field_name": str(name_val).strip() if name_val else None,
            "type": None,
            "length": None,
            "row": row,
            "evidence": {"sheet": sheet_name, "row": row},
        }
        
        if source_type_col:
            v = ws.cell(row, source_type_col).value
            if v: field["type"] = str(v).strip()
        if source_length_col:
            v = ws.cell(row, source_length_col).value
            if v: field["length"] = str(v).strip()
        
        # Add any extra columns
        extras = {}
        for col, hname in source_headers.items():
            if col not in (source_no_col, source_name_col, source_type_col, source_length_col):
                v = ws.cell(row, col).value
                if v is not None:
                    extras[hname] = str(v).strip()[:200]
        if extras:
            field["extra"] = extras
        
        result["source_fields"].append(field)
        if field["no"]:
            source_field_map[field["no"]] = field
    
    # Step 7: Extract target fields and mappings
    for row in range(data_start, max_row + 1):
        no_val = ws.cell(row, target_no_col).value if target_no_col else None
        name_val = ws.cell(row, target_name_col).value if target_name_col else None
        
        if no_val is None and name_val is None:
            continue
        
        field = {
            "no": str(no_val) if no_val is not None else None,
            "field_name": str(name_val).strip() if name_val else None,
            "variable": None,
            "type": None,
            "required": None,
            "length": None,
            "mapping_source_ref": None,
            "mapping_detail": None,
            "remark": None,
            "row": row,
            "evidence": {"sheet": sheet_name, "row": row},
        }
        
        if target_variable_col:
            v = ws.cell(row, target_variable_col).value
            if v: field["variable"] = str(v).strip()
        if target_type_col:
            v = ws.cell(row, target_type_col).value
            if v: field["type"] = str(v).strip()
        if target_required_col:
            v = ws.cell(row, target_required_col).value
            if v: field["required"] = str(v).strip()
        if target_length_col:
            v = ws.cell(row, target_length_col).value
            if v: field["length"] = str(v).strip()
        if target_mapping_source_col:
            v = ws.cell(row, target_mapping_source_col).value
            if v: field["mapping_source_ref"] = str(v).strip()
        if target_mapping_detail_col:
            v = ws.cell(row, target_mapping_detail_col).value
            if v: field["mapping_detail"] = str(v).strip()
        if target_remark_col:
            v = ws.cell(row, target_remark_col).value
            if v: field["remark"] = str(v).strip()
        
        # Extra target columns
        extras = {}
        known_cols = {target_no_col, target_name_col, target_variable_col, target_type_col,
                      target_required_col, target_length_col,
                      target_mapping_source_col, target_mapping_detail_col, target_remark_col}
        for col, hname in target_headers.items():
            if col not in known_cols:
                v = ws.cell(row, col).value
                if v is not None:
                    extras[hname] = str(v).strip()[:200]
        if extras:
            field["extra"] = extras
        
        result["target_fields"].append(field)
        
        # Step 8: Build mapping record
        if field["mapping_source_ref"]:
            source_ref = field["mapping_source_ref"]
            source_field = source_field_map.get(source_ref)
            
            # Try to resolve multi-line references (e.g. "6\n8\n7" = multiple source rows)
            resolved_sources = []
            if not source_field and "\n" in source_ref:
                refs = [r.strip() for r in source_ref.split("\n") if r.strip()]
                for r in refs:
                    if r in source_field_map:
                        fname = source_field_map[r].get("field_name")
                        if fname:
                            resolved_sources.append(fname)
                    elif r.lstrip("※").lstrip("(").rstrip(")") in source_field_map:
                        clean_r = r.lstrip("※").lstrip("(").rstrip(")")
                        fname = source_field_map[clean_r].get("field_name")
                        if fname:
                            resolved_sources.append(fname)
            
            # Also try stripping special prefixes (※, (1), etc.)
            if not source_field and not resolved_sources:
                clean_ref = source_ref.lstrip("※").lstrip("(").split(")")[0] if "※" in source_ref else source_ref
                if clean_ref != source_ref:
                    source_field = source_field_map.get(clean_ref)
            
            mapping = {
                "target_no": field["no"],
                "target_field": field["field_name"],
                "source_ref": source_ref,
                "source_field": source_field["field_name"] if source_field else None,
                "source_fields_multi": resolved_sources if resolved_sources else None,
                "mapping_detail": field["mapping_detail"],
                "confidence": 0.9 if source_field else (0.7 if resolved_sources else 0.5),
                "evidence": {
                    "workbook": wb_name,
                    "sheet": sheet_name,
                    "row": row,
                    "target_col": get_column_letter(target_mapping_source_col) if target_mapping_source_col else None,
                },
            }
            
            if not source_field and not resolved_sources and source_ref:
                # Categorize the reference type for better confidence scoring
                ref_clean = source_ref.strip()
                if ref_clean in ("ヘッダ", "明細", "固定値"):
                    # Known section/fixed-value references - not errors, just different ref type
                    mapping["ref_type"] = ref_clean
                    mapping["confidence"] = 0.8  # Higher confidence - known pattern
                elif ref_clean == "-":
                    # Explicit "no mapping" marker
                    mapping["ref_type"] = "no_mapping"
                    mapping["confidence"] = 0.9  
                elif ref_clean == "マッピング元":
                    # Header echo - likely a sub-header row, not real data
                    mapping["ref_type"] = "header_echo"
                    mapping["confidence"] = 0.3
                    mapping["uncertainty"] = "Likely a sub-header row, not actual data"
                else:
                    mapping["uncertainty"] = f"Source ref '{source_ref}' not found in source field list"
                
                # Only log as uncertainty if truly unresolved
                if ref_clean not in ("ヘッダ", "明細", "固定値", "-"):
                    result["uncertainties"].append({
                        "type": "unresolved_mapping_reference",
                        "target_field": field["field_name"],
                        "source_ref": source_ref,
                        "row": row,
                    })
            
            result["mappings"].append(mapping)
    
    return result


# ============================================================================
# Data Retrieval Condition Parser
# ============================================================================

def parse_data_retrieval_sheet(ws, sheet_name: str, wb_name: str, config: Config) -> dict:
    """Parse a データ取得条件 sheet."""
    max_row = min(ws.max_row or 1, config.MAX_ROWS_PER_SHEET)
    max_col = min(ws.max_column or 1, config.MAX_COLS_PER_SHEET)
    
    result = {
        "sheet_name": sheet_name,
        "workbook": wb_name,
        "sheet_type": "data_retrieval_condition",
        "conditions": [],
        "parameters": [],
    }
    
    # These sheets typically have conditions listed in rows
    for row in range(1, max_row + 1):
        row_data = {}
        for col in range(1, max_col + 1):
            v = ws.cell(row, col).value
            if v is not None:
                row_data[get_column_letter(col)] = str(v).strip()
        if row_data:
            result["conditions"].append({"row": row, "data": row_data})
    
    return result


# ============================================================================
# Mermaid Parser
# ============================================================================

def parse_mermaid_file(file_path: str, related_workbook: str = "") -> dict:
    """Parse a Mermaid file as authoritative flowchart source."""
    content = Path(file_path).read_text(encoding="utf-8")
    
    result = {
        "source_file": file_path,
        "source_type": "manual_mermaid",
        "related_workbook": related_workbook,
        "confidence": 1.0,
        "content": content,
        "nodes": [],
        "edges": [],
        "subgraphs": [],
    }
    
    # Parse nodes, edges, and subgraphs
    lines = content.split("\n")
    current_subgraph = None
    
    for line in lines:
        line = line.strip().rstrip("\r")
        
        # Subgraph detection
        if line.startswith("subgraph "):
            parts = line[len("subgraph "):].split("[", 1)
            sg_id = parts[0].strip().strip('"')
            sg_label = ""
            if len(parts) > 1:
                sg_label = parts[1].rstrip("]").strip('"')
            current_subgraph = {"id": sg_id, "label": sg_label, "nodes": []}
            result["subgraphs"].append(current_subgraph)
            continue
        
        if line == "end":
            current_subgraph = None
            continue
        
        # Node detection: ID["label"] or ID{"label"} or ID(["label"])
        import re
        
        # Edge detection: A --> B or A -->|label| B
        edge_match = re.match(r'(\w+)\s*-->\s*(?:\|"?([^"|]*)"?\|\s*)?(\w+)', line)
        if edge_match:
            source_id = edge_match.group(1)
            label = edge_match.group(2) or ""
            target_id = edge_match.group(3)
            result["edges"].append({
                "source": source_id,
                "target": target_id,
                "label": label,
                "evidence": {"source_type": "manual_mermaid", "file": file_path},
            })
            continue
        
        # Node with label
        node_match = re.match(r'(\w+)\[[""](.+?)[""]?\]', line)
        if not node_match:
            node_match = re.match(r'(\w+)\{"?(.+?)"?\}', line)
        if not node_match:
            node_match = re.match(r'(\w+)\(\["?(.+?)"?\]\)', line)
        
        if node_match:
            node_id = node_match.group(1)
            node_label = node_match.group(2).replace("<br/>", "\n")
            node_type = "process"
            if "{" in line:
                node_type = "decision"
            
            node = {
                "id": node_id,
                "label": node_label,
                "type": node_type,
                "subgraph": current_subgraph["id"] if current_subgraph else None,
                "evidence": {"source_type": "manual_mermaid", "file": file_path},
            }
            result["nodes"].append(node)
            if current_subgraph:
                current_subgraph["nodes"].append(node_id)
    
    return result


# ============================================================================
# Graph Builder - Normalize to GraphRAG nodes/edges
# ============================================================================

def build_graph(workbook_results: list, mermaid_results: list) -> dict:
    """Build GraphRAG-compatible nodes and edges."""
    nodes = []
    edges = []
    
    for wb_result in workbook_results:
        wb_name = wb_result["workbook_name"]
        
        # Workbook node
        wb_node_id = f"wb_{hashlib.md5(wb_name.encode()).hexdigest()[:8]}"
        nodes.append({
            "id": wb_node_id,
            "type": "Workbook",
            "label": wb_name,
            "properties": {"filename": wb_name},
        })
        
        for sheet_result in wb_result.get("sheets", []):
            sheet_name = sheet_result["sheet_name"]
            sheet_type = sheet_result.get("sheet_type", "unknown")
            sheet_id = f"sh_{hashlib.md5(f'{wb_name}/{sheet_name}'.encode()).hexdigest()[:8]}"
            
            nodes.append({
                "id": sheet_id,
                "type": "Sheet",
                "label": sheet_name,
                "properties": {"sheet_type": sheet_type, "workbook": wb_name},
            })
            edges.append({
                "source": wb_node_id, "target": sheet_id,
                "type": "HAS_SHEET",
            })
            
            # Source fields
            for field in sheet_result.get("source_fields", []):
                if not field.get("field_name"):
                    continue
                fno = field.get("no", "")
                fname = field["field_name"]
                field_id = f"sf_{hashlib.md5(f'{sheet_name}/src/{fno}_{fname}'.encode()).hexdigest()[:8]}"
                nodes.append({
                    "id": field_id,
                    "type": "SourceField",
                    "label": field["field_name"],
                    "properties": {
                        "no": field.get("no"),
                        "data_type": field.get("type"),
                        "length": field.get("length"),
                        "sheet": sheet_name,
                    },
                })
                edges.append({
                    "source": sheet_id, "target": field_id,
                    "type": "HAS_FIELD",
                })
            
            # Target fields
            for field in sheet_result.get("target_fields", []):
                if not field.get("field_name"):
                    continue
                fno = field.get("no", "")
                fname = field["field_name"]
                field_id = f"tf_{hashlib.md5(f'{sheet_name}/tgt/{fno}_{fname}'.encode()).hexdigest()[:8]}"
                nodes.append({
                    "id": field_id,
                    "type": "TargetField",
                    "label": field["field_name"],
                    "properties": {
                        "no": field.get("no"),
                        "data_type": field.get("type"),
                        "required": field.get("required"),
                        "length": field.get("length"),
                        "sheet": sheet_name,
                    },
                })
                edges.append({
                    "source": sheet_id, "target": field_id,
                    "type": "HAS_FIELD",
                })
            
            # Mapping edges
            for mapping in sheet_result.get("mappings", []):
                tgt_field = mapping.get("target_field")
                if not tgt_field:
                    continue
                tgt_no = mapping.get("target_no", "")
                tgt_id = f"tf_{hashlib.md5(f'{sheet_name}/tgt/{tgt_no}_{tgt_field}'.encode()).hexdigest()[:8]}"
                
                if mapping.get("source_field"):
                    # Single resolved source
                    src_ref = mapping.get("source_ref", "")
                    src_field = mapping["source_field"]
                    src_id = f"sf_{hashlib.md5(f'{sheet_name}/src/{src_ref}_{src_field}'.encode()).hexdigest()[:8]}"
                    edges.append({
                        "source": src_id, "target": tgt_id,
                        "type": "MAPS_TO",
                        "properties": {
                            "mapping_detail": mapping.get("mapping_detail"),
                            "confidence": mapping.get("confidence", 0.5),
                        },
                    })
                elif mapping.get("source_fields_multi"):
                    # Multi-source resolved
                    for i, sf in enumerate(mapping["source_fields_multi"]):
                        src_ref_part = mapping.get("source_ref", "").split("\n")[i] if i < len(mapping.get("source_ref", "").split("\n")) else str(i)
                        src_id = f"sf_{hashlib.md5(f'{sheet_name}/src/{src_ref_part}_{sf}'.encode()).hexdigest()[:8]}"
                        edges.append({
                            "source": src_id, "target": tgt_id,
                            "type": "MAPS_TO",
                            "properties": {
                                "mapping_detail": mapping.get("mapping_detail"),
                                "confidence": mapping.get("confidence", 0.5),
                                "multi_source": True,
                            },
                        })
                elif mapping.get("ref_type") in ("ヘッダ", "明細"):
                    # Section reference - create edge from section node
                    section_name = mapping.get("ref_type")
                    section_id = f"sec_{hashlib.md5(f'{sheet_name}/{section_name}'.encode()).hexdigest()[:8]}"
                    # Create section node if not exists
                    nodes.append({
                        "id": section_id,
                        "type": "Section",
                        "label": f"{section_name} ({sheet_name})",
                        "properties": {"section_type": section_name, "sheet": sheet_name},
                    })
                    edges.append({
                        "source": section_id, "target": tgt_id,
                        "type": "MAPS_TO",
                        "properties": {
                            "mapping_detail": mapping.get("mapping_detail"),
                            "confidence": mapping.get("confidence", 0.5),
                            "ref_type": section_name,
                        },
                    })
    
    # Mermaid flowchart nodes
    for mermaid in mermaid_results:
        for node in mermaid.get("nodes", []):
            nodes.append({
                "id": f"fn_{node['id']}",
                "type": "FlowNode",
                "label": node["label"],
                "properties": {
                    "node_type": node.get("type"),
                    "subgraph": node.get("subgraph"),
                    "source_file": mermaid.get("source_file"),
                },
            })
        
        for edge in mermaid.get("edges", []):
            edges.append({
                "source": f"fn_{edge['source']}",
                "target": f"fn_{edge['target']}",
                "type": "FLOW_TO",
                "properties": {"label": edge.get("label")},
            })
    
    return {"nodes": nodes, "edges": edges}


# ============================================================================
# KB Chunk Builder - Markdown chunks for vector KB
# ============================================================================

def build_kb_chunks(workbook_results: list, mermaid_results: list, markdown_files: dict) -> dict:
    """Build markdown chunks suitable for vector knowledge base embedding."""
    chunks = {}
    
    # Workbook summaries
    summary_lines = ["# Workbook Summaries\n"]
    for wb in workbook_results:
        summary_lines.append(f"## {wb['workbook_name']}")
        summary_lines.append(f"- Sheets: {len(wb.get('sheets', []))}")
        summary_lines.append(f"- Source: {wb.get('source_file', 'N/A')}")
        total_mappings = sum(len(s.get("mappings", [])) for s in wb.get("sheets", []))
        summary_lines.append(f"- Total mappings: {total_mappings}")
        summary_lines.append("")
    chunks["workbook_summary.md"] = "\n".join(summary_lines)
    
    # Mapping records
    mapping_lines = ["# Mapping Records\n"]
    for wb in workbook_results:
        for sheet in wb.get("sheets", []):
            mappings = sheet.get("mappings", [])
            if mappings:
                mapping_lines.append(f"## {sheet['sheet_name']} ({wb['workbook_name']})")
                mapping_lines.append("")
                mapping_lines.append("| Target Field | Source Ref | Source Field | Detail | Confidence |")
                mapping_lines.append("|---|---|---|---|---|")
                for m in mappings:
                    # Escape newlines in table cells for proper markdown
                    tgt = str(m.get('target_field', '') or '').replace('\n', ' / ')
                    src_ref = str(m.get('source_ref', '') or '').replace('\n', ', ')
                    src_field = str(m.get('source_field', '') or '?').replace('\n', ' / ')
                    detail = str(m.get('mapping_detail', '') or '').replace('\n', ' / ')[:100]
                    # Also handle multi-source fields
                    if m.get('source_fields_multi'):
                        src_field = " + ".join(s for s in m['source_fields_multi'] if s)
                    mapping_lines.append(
                        f"| {tgt} | {src_ref} "
                        f"| {src_field} | {detail} "
                        f"| {m.get('confidence', 0):.1f} |"
                    )
                mapping_lines.append("")
    chunks["mapping_records.md"] = "\n".join(mapping_lines)
    
    # Flowchart records from Mermaid
    flow_lines = ["# Flowchart Records\n"]
    for mermaid in mermaid_results:
        flow_lines.append(f"## Flowchart: {Path(mermaid['source_file']).name}")
        flow_lines.append(f"- Source: {mermaid['source_type']} (confidence: {mermaid['confidence']})")
        flow_lines.append(f"- Nodes: {len(mermaid.get('nodes', []))}")
        flow_lines.append(f"- Edges: {len(mermaid.get('edges', []))}")
        flow_lines.append(f"- Subgraphs: {len(mermaid.get('subgraphs', []))}")
        flow_lines.append("")
        
        for sg in mermaid.get("subgraphs", []):
            flow_lines.append(f"### {sg['label'] or sg['id']}")
            for nid in sg.get("nodes", []):
                # Find node label
                node = next((n for n in mermaid.get("nodes", []) if n["id"] == nid), None)
                if node:
                    flow_lines.append(f"- {node['label']}")
            flow_lines.append("")
    chunks["flowchart_records.md"] = "\n".join(flow_lines)
    
    # Sheet markdown dumps (already generated)
    # These are individual sheet dumps that go into the kb_chunks directory
    
    return chunks


# ============================================================================
# Quality Report
# ============================================================================

def generate_quality_report(workbook_results: list, mermaid_results: list, manifest: dict) -> dict:
    """Generate quality metrics and issue tracking."""
    
    total_sheets = sum(len(wb.get("sheets", [])) for wb in workbook_results)
    total_source_fields = sum(
        len(s.get("source_fields", []))
        for wb in workbook_results
        for s in wb.get("sheets", [])
    )
    total_target_fields = sum(
        len(s.get("target_fields", []))
        for wb in workbook_results
        for s in wb.get("sheets", [])
    )
    total_mappings = sum(
        len(s.get("mappings", []))
        for wb in workbook_results
        for s in wb.get("sheets", [])
    )
    total_uncertainties = sum(
        len(s.get("uncertainties", []))
        for wb in workbook_results
        for s in wb.get("sheets", [])
    )
    
    # Confidence distribution
    all_mappings = [
        m for wb in workbook_results
        for s in wb.get("sheets", [])
        for m in s.get("mappings", [])
    ]
    high_conf = sum(1 for m in all_mappings if m.get("confidence", 0) >= 0.8)
    med_conf = sum(1 for m in all_mappings if 0.5 <= m.get("confidence", 0) < 0.8)
    low_conf = sum(1 for m in all_mappings if m.get("confidence", 0) < 0.5)
    
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "pipeline_version": "2.0.0",
        "statistics": {
            "workbooks_processed": len(workbook_results),
            "sheets_processed": total_sheets,
            "source_fields_extracted": total_source_fields,
            "target_fields_extracted": total_target_fields,
            "mappings_extracted": total_mappings,
            "uncertainties": total_uncertainties,
            "mermaid_files": len(mermaid_results),
            "mermaid_nodes": sum(len(m.get("nodes", [])) for m in mermaid_results),
            "mermaid_edges": sum(len(m.get("edges", [])) for m in mermaid_results),
        },
        "confidence_distribution": {
            "high_0.8_to_1.0": high_conf,
            "medium_0.5_to_0.8": med_conf,
            "low_below_0.5": low_conf,
        },
        "issues": [],
        "human_review_required": [],
    }
    
    # Collect issues
    for wb in workbook_results:
        for sheet in wb.get("sheets", []):
            for u in sheet.get("uncertainties", []):
                report["issues"].append({
                    "type": u.get("type", "unknown"),
                    "detail": u.get("detail", str(u)),
                    "location": f"{wb['workbook_name']}/{sheet['sheet_name']}",
                    "severity": u.get("severity", "medium"),
                })
                if u.get("severity") == "high":
                    report["human_review_required"].append({
                        "workbook": wb["workbook_name"],
                        "sheet": sheet["sheet_name"],
                        "issue": u.get("detail", str(u)),
                    })
    
    return report


# ============================================================================
# Main Pipeline Orchestrator
# ============================================================================

def run_pipeline(no_s3_sync: bool = False, no_llm: bool = True):
    """Run the full pipeline."""
    config = Config()
    if no_llm:
        config.USE_LLM = False
    
    # Setup output directory
    output_dir = config.OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )
    
    start_time = time.time()
    logger.info("=" * 70)
    logger.info("Excel Parse Pipeline v2 - Starting")
    logger.info(f"  Input: s3://{config.S3_BUCKET}/{config.S3_INPUT_PREFIX}/")
    logger.info(f"  Output: {output_dir}")
    logger.info("=" * 70)
    
    # ---- Stage 1: Scan and Download ----
    logger.info("Stage 1: Source Scan & Download")
    manifest = scan_and_download(config)
    
    manifest_path = output_dir / "source_files_manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    
    logger.info(f"  Found {manifest['scan_summary']['excel_files']} Excel, "
                f"{manifest['scan_summary']['mermaid_files']} Mermaid files")
    
    # ---- Stage 2: Process Each Workbook ----
    logger.info("Stage 2: Workbook Processing")
    
    workbook_results = []
    all_markdown = {}  # sheet_key -> markdown content
    
    for excel_file in manifest["excel_files"]:
        local_path = excel_file["local_path"]
        wb_name = excel_file["filename"]
        
        logger.info(f"  Processing: {wb_name}")
        
        wb_result = {
            "workbook_name": wb_name,
            "source_file": excel_file["key"],
            "local_path": local_path,
            "sheets": [],
        }
        
        try:
            wb = openpyxl.load_workbook(local_path, data_only=True, read_only=False)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"    Sheet: {sheet_name} ({ws.max_row}r × {ws.max_column}c)")
                
                # Build atlas
                atlas = build_sheet_atlas(ws, sheet_name, config)
                
                # Classify sheet type
                classification = classify_sheet_type(sheet_name, atlas)
                sheet_type = classification["type"]
                
                # Generate markdown dump (always - minimum requirement)
                md_content = sheet_to_markdown(ws, sheet_name, wb_name, config)
                md_key = f"{wb_name}/{sheet_name}"
                all_markdown[md_key] = md_content
                
                # Structured parsing based on type
                sheet_result = {
                    "sheet_name": sheet_name,
                    "sheet_type": sheet_type,
                    "classification_confidence": classification["confidence"],
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "source_fields": [],
                    "target_fields": [],
                    "mappings": [],
                    "uncertainties": [],
                }
                
                if sheet_type == "mapping_sheet":
                    parsed = parse_mapping_sheet(ws, sheet_name, wb_name, config)
                    sheet_result["source_fields"] = parsed.get("source_fields", [])
                    sheet_result["target_fields"] = parsed.get("target_fields", [])
                    sheet_result["mappings"] = parsed.get("mappings", [])
                    sheet_result["uncertainties"] = parsed.get("uncertainties", [])
                    sheet_result["metadata"] = parsed.get("metadata", {})
                    sheet_result["source_config"] = parsed.get("source_config", {})
                    sheet_result["target_config"] = parsed.get("target_config", {})
                    sheet_result["_source_headers"] = parsed.get("_source_headers", {})
                    sheet_result["_target_headers"] = parsed.get("_target_headers", {})
                elif sheet_type == "data_retrieval_condition":
                    parsed = parse_data_retrieval_sheet(ws, sheet_name, wb_name, config)
                    sheet_result["conditions"] = parsed.get("conditions", [])
                
                wb_result["sheets"].append(sheet_result)
                
                logger.info(f"      Type: {sheet_type} | Fields: {len(sheet_result['source_fields'])}+{len(sheet_result['target_fields'])} | Mappings: {len(sheet_result['mappings'])}")
            
            wb.close()
        except Exception as e:
            logger.error(f"  ERROR processing {wb_name}: {e}", exc_info=True)
            wb_result["error"] = str(e)
        
        workbook_results.append(wb_result)
    
    # ---- Stage 3: Parse Mermaid Files ----
    logger.info("Stage 3: Mermaid Parsing")
    mermaid_results = []
    
    for mermaid_file in manifest.get("mermaid_files", []):
        local_path = mermaid_file["local_path"]
        logger.info(f"  Parsing: {mermaid_file['filename']}")
        
        try:
            result = parse_mermaid_file(local_path, related_workbook="M社様_DSSスクリプト改修概要_フローチャート.xlsx")
            result["s3_key"] = mermaid_file["key"]
            mermaid_results.append(result)
            logger.info(f"    Nodes: {len(result['nodes'])}, Edges: {len(result['edges'])}, Subgraphs: {len(result['subgraphs'])}")
        except Exception as e:
            logger.error(f"  ERROR parsing {local_path}: {e}")
    
    # ---- Stage 4: Save Markdown Dumps ----
    logger.info("Stage 4: Saving Markdown Dumps")
    md_dir = output_dir / "markdown"
    md_dir.mkdir(parents=True, exist_ok=True)
    
    for key, content in all_markdown.items():
        # Sanitize filename
        safe_name = key.replace("/", "__").replace(" ", "_")
        md_path = md_dir / f"{safe_name}.md"
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(content)
    
    # Also create a combined markdown file per workbook
    for wb_result in workbook_results:
        wb_name = wb_result["workbook_name"]
        combined_lines = [f"# {wb_name}\n"]
        for sheet in wb_result.get("sheets", []):
            sheet_key = f"{wb_name}/{sheet['sheet_name']}"
            if sheet_key in all_markdown:
                combined_lines.append(all_markdown[sheet_key])
                combined_lines.append("\n---\n")
        
        safe_wb_name = wb_name.replace(" ", "_")
        combined_path = md_dir / f"_combined_{safe_wb_name}.md"
        with open(combined_path, "w", encoding="utf-8") as f:
            f.write("\n".join(combined_lines))
    
    logger.info(f"  Saved {len(all_markdown)} sheet markdown files")
    
    # ---- Stage 5: Save Structured Data ----
    logger.info("Stage 5: Structured Data Output")
    structured_dir = output_dir / "structured"
    structured_dir.mkdir(parents=True, exist_ok=True)
    
    # Fields JSONL
    with open(structured_dir / "source_fields.jsonl", "w", encoding="utf-8") as f:
        for wb in workbook_results:
            for sheet in wb.get("sheets", []):
                for field in sheet.get("source_fields", []):
                    record = {"workbook": wb["workbook_name"], "sheet": sheet["sheet_name"], **field}
                    f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    
    with open(structured_dir / "target_fields.jsonl", "w", encoding="utf-8") as f:
        for wb in workbook_results:
            for sheet in wb.get("sheets", []):
                for field in sheet.get("target_fields", []):
                    record = {"workbook": wb["workbook_name"], "sheet": sheet["sheet_name"], **field}
                    f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    
    # Mappings JSONL
    with open(structured_dir / "mappings.jsonl", "w", encoding="utf-8") as f:
        for wb in workbook_results:
            for sheet in wb.get("sheets", []):
                for mapping in sheet.get("mappings", []):
                    record = {"workbook": wb["workbook_name"], "sheet": sheet["sheet_name"], **mapping}
                    f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    
    # Mermaid/flowchart JSONL
    with open(structured_dir / "flow_nodes.jsonl", "w", encoding="utf-8") as f:
        for mermaid in mermaid_results:
            for node in mermaid.get("nodes", []):
                f.write(json.dumps(node, ensure_ascii=False) + "\n")
    
    with open(structured_dir / "flow_edges.jsonl", "w", encoding="utf-8") as f:
        for mermaid in mermaid_results:
            for edge in mermaid.get("edges", []):
                f.write(json.dumps(edge, ensure_ascii=False) + "\n")
    
    # Uncertainties JSONL
    with open(structured_dir / "uncertainties.jsonl", "w", encoding="utf-8") as f:
        for wb in workbook_results:
            for sheet in wb.get("sheets", []):
                for u in sheet.get("uncertainties", []):
                    record = {"workbook": wb["workbook_name"], "sheet": sheet["sheet_name"], **u}
                    f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    
    # ---- Stage 6: Graph Output ----
    logger.info("Stage 6: Graph Building")
    graph = build_graph(workbook_results, mermaid_results)
    
    graph_dir = output_dir / "graph"
    graph_dir.mkdir(parents=True, exist_ok=True)
    
    with open(graph_dir / "nodes.jsonl", "w", encoding="utf-8") as f:
        for node in graph["nodes"]:
            f.write(json.dumps(node, ensure_ascii=False) + "\n")
    
    with open(graph_dir / "edges.jsonl", "w", encoding="utf-8") as f:
        for edge in graph["edges"]:
            f.write(json.dumps(edge, ensure_ascii=False) + "\n")
    
    logger.info(f"  Graph: {len(graph['nodes'])} nodes, {len(graph['edges'])} edges")
    
    # ---- Stage 7: KB Chunks ----
    logger.info("Stage 7: KB Chunk Generation")
    kb_chunks = build_kb_chunks(workbook_results, mermaid_results, all_markdown)
    
    kb_dir = output_dir / "kb_chunks"
    kb_dir.mkdir(parents=True, exist_ok=True)
    
    for filename, content in kb_chunks.items():
        with open(kb_dir / filename, "w", encoding="utf-8") as f:
            f.write(content)
    
    # Also save full mermaid content as KB chunk
    for mermaid in mermaid_results:
        mermaid_kb_path = kb_dir / f"mermaid_{Path(mermaid['source_file']).stem}.md"
        with open(mermaid_kb_path, "w", encoding="utf-8") as f:
            f.write(f"# Flowchart: {Path(mermaid['source_file']).name}\n\n")
            f.write(f"Source: {mermaid['source_type']} (authoritative)\n\n")
            f.write("```mermaid\n")
            f.write(mermaid["content"])
            f.write("\n```\n")
    
    logger.info(f"  Generated {len(kb_chunks)} KB chunk files")
    
    # ---- Stage 8: Quality Report ----
    logger.info("Stage 8: Quality Report")
    quality = generate_quality_report(workbook_results, mermaid_results, manifest)
    
    with open(output_dir / "quality_report.json", "w", encoding="utf-8") as f:
        json.dump(quality, f, ensure_ascii=False, indent=2)
    
    # Human-readable quality report
    qr_lines = [
        "# Quality Report - Excel Parse Pipeline v2\n",
        f"Generated: {quality['generated_at']}",
        "",
        "## Statistics",
        "",
        "| Metric | Value |",
        "|--------|-------|",
    ]
    for k, v in quality["statistics"].items():
        qr_lines.append(f"| {k} | {v} |")
    qr_lines.append("")
    qr_lines.append("## Confidence Distribution")
    qr_lines.append("")
    for k, v in quality["confidence_distribution"].items():
        qr_lines.append(f"- {k}: {v}")
    qr_lines.append("")
    
    if quality["issues"]:
        qr_lines.append(f"## Issues ({len(quality['issues'])})")
        qr_lines.append("")
        for issue in quality["issues"][:50]:
            qr_lines.append(f"- [{issue['severity']}] {issue['type']}: {issue['detail']} @ {issue['location']}")
    
    if quality["human_review_required"]:
        qr_lines.append(f"\n## Human Review Required ({len(quality['human_review_required'])})")
        qr_lines.append("")
        for item in quality["human_review_required"]:
            qr_lines.append(f"- {item['workbook']}/{item['sheet']}: {item['issue']}")
    
    with open(output_dir / "quality_report.md", "w", encoding="utf-8") as f:
        f.write("\n".join(qr_lines))
    
    # ---- Stage 9: S3 Sync ----
    if not no_s3_sync:
        logger.info("Stage 9: S3 Sync")
        try:
            s3io = S3IO(config)
            uploaded = s3io.sync_upload(output_dir, config.S3_OUTPUT_PREFIX)
            logger.info(f"  Uploaded {uploaded} files to s3://{config.S3_BUCKET}/{config.S3_OUTPUT_PREFIX}/")
        except Exception as e:
            logger.warning(f"  S3 sync failed (non-fatal): {e}")
    else:
        logger.info("Stage 9: S3 Sync (SKIPPED)")
    
    # ---- Final Summary ----
    elapsed = time.time() - start_time
    logger.info("=" * 70)
    logger.info("Pipeline Complete!")
    logger.info(f"  Duration: {elapsed:.1f}s")
    logger.info(f"  Output: {output_dir}")
    logger.info(f"  Markdown sheets: {len(all_markdown)}")
    logger.info(f"  Source fields: {quality['statistics']['source_fields_extracted']}")
    logger.info(f"  Target fields: {quality['statistics']['target_fields_extracted']}")
    logger.info(f"  Mappings: {quality['statistics']['mappings_extracted']}")
    logger.info(f"  Uncertainties: {quality['statistics']['uncertainties']}")
    logger.info(f"  Mermaid nodes: {quality['statistics']['mermaid_nodes']}")
    logger.info(f"  Graph: {len(graph['nodes'])} nodes, {len(graph['edges'])} edges")
    logger.info("=" * 70)
    
    return {
        "status": "complete",
        "duration_seconds": elapsed,
        "output_dir": str(output_dir),
        "statistics": quality["statistics"],
    }


# ============================================================================
# Entry Point
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel Parse Pipeline v2")
    parser.add_argument("--no-s3-sync", action="store_true", help="Skip S3 output sync")
    parser.add_argument("--no-llm", action="store_true", default=True, help="Skip LLM calls (default: True)")
    args = parser.parse_args()
    
    result = run_pipeline(no_s3_sync=args.no_s3_sync, no_llm=args.no_llm)
    print(f"\nResult: {json.dumps(result, indent=2, ensure_ascii=False)}")
