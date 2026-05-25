#!/usr/bin/env python3
"""Extract mapping records from all mapping sheets with adaptive column detection."""

import json, os, re, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def detect_columns(ws, max_scan=30):
    """Auto-detect column positions by searching for header keywords."""
    cols = {}
    header_row = None
    
    # Search for the target table header row by looking for "No" + "項目名称" pattern
    # in the right half of the sheet (col > 50)
    for row_idx in range(18, max_scan):
        for col_idx in range(50, min(ws.max_column + 1, 192)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip() == "No":
                # Check if next non-empty cell in same row is 項目名称
                for next_col in range(col_idx + 1, col_idx + 10):
                    next_val = ws.cell(row=row_idx, column=next_col).value
                    if next_val and "項目名称" in str(next_val):
                        header_row = row_idx
                        cols["target_no"] = col_idx
                        cols["target_name"] = next_col
                        break
                if header_row:
                    break
        if header_row:
            break
    
    if not header_row:
        return None, None, None, None
    
    # Scan the header row for other known columns
    for col_idx in range(cols["target_name"] + 1, min(ws.max_column + 1, 192)):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            val_str = str(val).strip()
            if val_str == "変数" and "target_variable" not in cols:
                cols["target_variable"] = col_idx
            elif "Type" in val_str and "target_type" not in cols:
                cols["target_type"] = col_idx
            elif "必須" in val_str and "target_required" not in cols:
                cols["target_required"] = col_idx
            elif "長さ" in val_str and "target_length" not in cols:
                cols["target_length"] = col_idx
            elif "備考" in val_str and "target_notes" not in cols:
                cols["target_notes"] = col_idx
            elif "マッピング元" in val_str:
                cols["mapping_src"] = col_idx
            elif "編集内容" in val_str and "DataSpider" not in val_str:
                cols["edit_rule"] = col_idx
            elif "DataSpider" in val_str or ("編集内容" in val_str and "DataSpider" in val_str):
                cols["ds_rule"] = col_idx
    
    # Source table: look for "No" + "項目名称" in cols A area
    source_header_row = None
    for row_idx in range(18, max_scan):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip() == "No":
            source_header_row = row_idx
            break
    
    # Also handle case where source header is "(ヘッダレコード)" style
    if source_header_row is None:
        for row_idx in range(18, max_scan):
            val = ws.cell(row=row_idx, column=1).value
            if val and ("レコード" in str(val) or "ヘッダ" in str(val)):
                # Next row might have "No"
                next_val = ws.cell(row=row_idx + 1, column=1).value
                if next_val and str(next_val).strip() == "No":
                    source_header_row = row_idx + 1
                    break
    
    # Source field columns
    cols["source_no"] = 1  # A
    if source_header_row:
        for col_idx in range(2, 50):
            val = ws.cell(row=source_header_row, column=col_idx).value
            if val:
                val_str = str(val).strip()
                if "項目名称" in val_str and "source_name" not in cols:
                    cols["source_name"] = col_idx
                elif "Type" in val_str and "source_type" not in cols:
                    cols["source_type"] = col_idx
                elif "長さ" in val_str and "source_length" not in cols:
                    cols["source_length"] = col_idx
                elif "備考" in val_str and "source_notes" not in cols:
                    cols["source_notes"] = col_idx
    
    # Defaults for missing columns
    cols.setdefault("source_name", 3)
    cols.setdefault("source_type", 24)
    cols.setdefault("source_length", 29)
    cols.setdefault("source_notes", 47)
    cols.setdefault("target_variable", cols.get("target_name", 65) + 10)
    cols.setdefault("target_type", cols.get("target_name", 65) + 20)
    cols.setdefault("target_required", cols.get("target_type", 88) + 5)
    cols.setdefault("target_length", cols.get("target_required", 93) + 2)
    cols.setdefault("target_notes", cols.get("target_length", 95) + 5)
    cols.setdefault("mapping_src", cols.get("target_notes", 100) + 5)
    cols.setdefault("edit_rule", cols.get("mapping_src", 120) + 9)
    cols.setdefault("ds_rule", cols.get("edit_rule", 129) + 6)
    
    data_start_row = header_row + 1
    
    return source_header_row, header_row, data_start_row, cols


def extract_mapping_sheet(ws, sheet_name, wb_name, run_id):
    """Extract mapping records from a single sheet."""
    
    source_header_row, target_header_row, data_start_row, cols = detect_columns(ws)
    
    if not target_header_row:
        return [], {}
    
    # Determine mapping direction
    src_system = ""
    dst_system = ""
    for row_idx in range(8, 20):
        for col_idx in range(1, 5):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and "送信元" in str(val):
                src_system = str(ws.cell(row=row_idx, column=6).value or "")
                break
        for col_idx in range(60, 75):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and "送信先" in str(val):
                # Look for value in next columns
                for vc in range(col_idx + 1, col_idx + 10):
                    v = ws.cell(row=row_idx, column=vc).value
                    if v and str(v).strip() not in ("", "ー", "-"):
                        dst_system = str(v)
                        break
                break
    
    mapping_direction = f"{src_system} → {dst_system}" if src_system and dst_system else sheet_name
    
    # Extract source fields
    source_fields = {}
    for row_idx in range(data_start_row, ws.max_row + 1):
        no_val = ws.cell(row=row_idx, column=cols["source_no"]).value
        no_str = str(no_val).strip() if no_val is not None else ""
        try:
            if no_str and no_str.replace('.','').isdigit():
                key = str(int(float(no_str)))
            else:
                continue
        except (ValueError, TypeError):
            continue
        
        name_val = ws.cell(row=row_idx, column=cols["source_name"]).value
        source_fields[key] = {
            "source_no": key,
            "source_field_name": str(name_val).strip() if name_val else "",
            "source_type": str(ws.cell(row=row_idx, column=cols["source_type"]).value or ""),
            "source_length": str(ws.cell(row=row_idx, column=cols["source_length"]).value or ""),
            "row": row_idx
        }
    
    # Extract target fields and mappings
    current_section = "header"
    mapping_records = []
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        target_no_val = ws.cell(row=row_idx, column=cols["target_no"]).value
        
        # Section markers
        if target_no_val and isinstance(target_no_val, str):
            s = target_no_val.strip()
            if "明細" in s:
                current_section = "detail"
                continue
            if "ヘッダ" in s:
                current_section = "header"
                continue
            if s in ("No", "マッピング元") or "レコード" in s:
                continue
        
        target_no = target_no_val
        target_name = ws.cell(row=row_idx, column=cols["target_name"]).value
        
        if target_no is None and target_name is None:
            continue
        
        # Validate target_no is numeric
        if target_no is not None:
            no_str = str(target_no).strip()
            if not no_str.replace('.','').isdigit():
                continue
        
        target_type = ws.cell(row=row_idx, column=cols["target_type"]).value
        target_req = ws.cell(row=row_idx, column=cols["target_required"]).value
        target_len = ws.cell(row=row_idx, column=cols["target_length"]).value
        
        mapping_src_raw = ws.cell(row=row_idx, column=cols["mapping_src"]).value
        edit_rule = ws.cell(row=row_idx, column=cols["edit_rule"]).value
        ds_rule = ws.cell(row=row_idx, column=cols["ds_rule"]).value
        
        # Parse mapping source
        source_field_refs = []
        mapping_type = "unknown"
        fixed_value = ""
        
        src_empty = mapping_src_raw is None or str(mapping_src_raw).strip() in ("", "-", "ー")
        
        if src_empty:
            if ds_rule and "固定値" in str(ds_rule):
                mapping_type = "fixed_value"
                match = re.search(r'固定値[：:]?\s*[\"「](.+?)[\"」]', str(ds_rule))
                if match:
                    fixed_value = match.group(1)
            elif edit_rule and ("設定不要" in str(edit_rule) or "除く" in str(edit_rule)):
                mapping_type = "not_set"
            elif ds_rule and ("項目自体を除く" in str(ds_rule) or "設定不要" in str(ds_rule)):
                mapping_type = "not_set"
            elif ds_rule and "固定" in str(ds_rule):
                mapping_type = "fixed_value"
            else:
                mapping_type = "unknown"
        else:
            src_text = str(mapping_src_raw).strip()
            if "固定値" in src_text or "固定" in src_text:
                mapping_type = "conditional_rule"
            
            src_numbers = re.findall(r'(\d+)', src_text)
            for sn in src_numbers:
                if sn in source_fields:
                    source_field_refs.append({
                        "source_no": sn,
                        "source_field_name": source_fields[sn]["source_field_name"],
                        "source_type": source_fields[sn]["source_type"],
                        "source_length": source_fields[sn]["source_length"],
                        "evidence_cells": [f"A{source_fields[sn]['row']}", f"{get_column_letter(cols['source_name'])}{source_fields[sn]['row']}"]
                    })
        
        # Classify mapping type
        if mapping_type == "unknown":
            if not source_field_refs:
                if ds_rule and ("固定値" in str(ds_rule) or "固定" in str(ds_rule)):
                    mapping_type = "fixed_value"
                else:
                    mapping_type = "unknown"
            elif len(source_field_refs) == 1:
                combined = str(edit_rule or "") + str(ds_rule or "")
                if "コード変換" in combined:
                    mapping_type = "code_conversion"
                elif "連結" in combined or "結合" in combined:
                    mapping_type = "concatenation"
                elif "算出" in combined or "計算" in combined:
                    mapping_type = "calculation"
                elif "条件" in combined or "場合" in combined or "判定" in combined:
                    mapping_type = "conditional_rule"
                elif "固定値" in combined or "固定" in combined:
                    mapping_type = "fixed_value"
                elif "変換" in combined:
                    mapping_type = "date_format_conversion" if "日付" in combined else "code_conversion"
                else:
                    mapping_type = "direct_copy"
            else:
                combined = str(edit_rule or "") + str(ds_rule or "")
                if "連結" in combined or "結合" in combined:
                    mapping_type = "concatenation"
                elif "判定" in combined or "場合" in combined:
                    mapping_type = "conditional_rule"
                else:
                    mapping_type = "conditional_rule"
        
        confidence = 0.9
        if not source_field_refs and mapping_type == "unknown":
            confidence = 0.4
        elif mapping_type in ("conditional_rule",) and len(source_field_refs) > 3:
            confidence = 0.7
        elif mapping_type == "fixed_value" and not fixed_value:
            confidence = 0.6
        
        human_review = confidence < 0.8 or mapping_type in ("unknown", "conditional_rule")
        
        record = {
            "run_id": run_id,
            "workbook": wb_name,
            "sheet": sheet_name,
            "mapping_direction": mapping_direction,
            "record_type": current_section,
            "target_field_no": str(target_no) if target_no else "",
            "target_field_name": str(target_name).strip() if target_name else "",
            "target_field_japanese_name": str(target_name).strip() if target_name else "",
            "target_required": target_req is not None and str(target_req).strip() in ("〇", "○", "◯"),
            "target_data_type": str(target_type) if target_type else "",
            "target_length": str(target_len) if target_len else "",
            "source_fields": source_field_refs,
            "fixed_value": fixed_value,
            "mapping_type": mapping_type,
            "transformation_rule_summary": str(edit_rule)[:200] if edit_rule and str(edit_rule).strip() not in ("-", "None", "", "ー") else "",
            "dataspider_rule": str(ds_rule)[:300] if ds_rule and str(ds_rule).strip() not in ("-", "None", "", "ー") else "",
            "condition_logic": "",
            "business_meaning": "",
            "evidence": {
                "target_cells": [f"{get_column_letter(cols['target_no'])}{row_idx}", f"{get_column_letter(cols['target_name'])}{row_idx}"],
                "mapping_source_cells": [f"{get_column_letter(cols['mapping_src'])}{row_idx}"] if not src_empty else [],
                "source_cells": [ref["evidence_cells"][0] for ref in source_field_refs],
                "rule_cells": [f"{get_column_letter(cols['edit_rule'])}{row_idx}"] if edit_rule and str(edit_rule).strip() not in ("-", "None", "", "ー") else [],
                "dataspider_rule_cells": [f"{get_column_letter(cols['ds_rule'])}{row_idx}"] if ds_rule and str(ds_rule).strip() not in ("-", "None", "", "ー") else []
            },
            "confidence": confidence,
            "human_review_required": human_review
        }
        mapping_records.append(record)
    
    return mapping_records, source_fields


def main():
    run_id = open("/tmp/current_run_id.txt").read().strip()
    output_dir = open("/tmp/current_output_dir.txt").read().strip()
    
    wb_path = "data/input/sample_20260519/02_詳細設計/MW_IFマッピング定義書_205_発注情報(登録・変更・取消).xlsx"
    wb_name = os.path.basename(wb_path)
    wb = load_workbook(wb_path, data_only=True)
    
    mapping_sheets = [s for s in wb.sheetnames if "マッピングシート" in s]
    
    total_records = 0
    all_records = []
    sheet_summaries = []
    
    map_dir = os.path.join(output_dir, "mappings/mapping")
    os.makedirs(map_dir, exist_ok=True)
    
    for sheet_name in mapping_sheets:
        ws = wb[sheet_name]
        print(f"Processing: {sheet_name}")
        
        records, source_fields = extract_mapping_sheet(ws, sheet_name, wb_name, run_id)
        
        sheet_summaries.append({
            "sheet": sheet_name,
            "records": len(records),
            "source_fields": len(source_fields),
            "direction": records[0]["mapping_direction"] if records else ""
        })
        
        if records:
            safe_name = sheet_name.replace("/", "_").replace("\\", "_")
            jsonl_path = os.path.join(map_dir, f"{safe_name}.mapping.jsonl")
            with open(jsonl_path, "w") as f:
                for rec in records:
                    f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            
            md_path = os.path.join(map_dir, f"{safe_name}.mapping_review.md")
            with open(md_path, "w") as f:
                f.write(f"# Mapping Review: {sheet_name}\n\n")
                f.write(f"Total records: {len(records)}\n")
                f.write(f"Direction: {records[0]['mapping_direction']}\n")
                f.write(f"Source fields: {len(source_fields)}\n\n")
                f.write("| No | Section | Target Field | Source(s) | Type | Conf | Review |\n")
                f.write("|---|---|---|---|---|---|---|\n")
                for rec in records:
                    src_str = ",".join([s['source_no'] for s in rec["source_fields"]]) or "-"
                    f.write(f"| {rec['target_field_no']} | {rec['record_type']} | {rec['target_field_name'][:20]} | {src_str[:20]} | {rec['mapping_type'][:15]} | {rec['confidence']} | {'⚠️' if rec['human_review_required'] else '✓'} |\n")
            
            print(f"  → {len(records)} records, {len(source_fields)} source fields")
            total_records += len(records)
            all_records.extend(records)
        else:
            print(f"  → No mapping records (structure not detected)")
    
    # Save combined
    combined_path = os.path.join(map_dir, "_all_mappings_combined.jsonl")
    with open(combined_path, "w") as f:
        for rec in all_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    
    # Save summary
    summary_path = os.path.join(map_dir, "_extraction_summary.json")
    with open(summary_path, "w") as f:
        json.dump({
            "run_id": run_id,
            "total_records": total_records,
            "sheets_processed": len(mapping_sheets),
            "sheet_summaries": sheet_summaries
        }, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*60}")
    print(f"TOTAL: {total_records} mapping records from {len(mapping_sheets)} sheets")
    
    types = {}
    for r in all_records:
        types[r["mapping_type"]] = types.get(r["mapping_type"], 0) + 1
    print(f"\nMapping type distribution:")
    for t, c in sorted(types.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}")
    
    wb.close()
    return total_records, all_records


if __name__ == "__main__":
    main()
