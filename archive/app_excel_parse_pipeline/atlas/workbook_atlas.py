"""Workbook atlas: extract structural facts from Excel workbooks.

This module extracts ONLY physical/structural facts - no semantic interpretation.
It does NOT hardcode any business-specific headers or column names.
"""
import json
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter


def build_workbook_atlas(workbook_path: str) -> dict:
    """Build a complete structural atlas for a workbook.
    
    Extracts only objective physical facts:
    - Sheet names and dimensions
    - Used ranges
    - Merged cells
    - Cell values, styles, formulas
    - Row heights, column widths
    - Hidden rows/columns
    """
    path = Path(workbook_path)
    wb = openpyxl.load_workbook(str(path), data_only=False, read_only=False)

    atlas = {
        "workbook_id": path.stem,
        "workbook_name": path.name,
        "source_file": str(path),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "sheets": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_atlas = _build_sheet_atlas(ws, sheet_name)
        atlas["sheets"].append(sheet_atlas)

    wb.close()
    return atlas


def _build_sheet_atlas(ws, sheet_name: str) -> dict:
    """Build structural atlas for a single sheet."""
    # Basic dimensions
    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1

    used_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    # Merged cells
    merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]

    # Hidden rows and columns
    hidden_rows = []
    hidden_cols = []
    if hasattr(ws, 'row_dimensions'):
        for row_idx, rd in ws.row_dimensions.items():
            if rd.hidden:
                hidden_rows.append(row_idx)
    if hasattr(ws, 'column_dimensions'):
        for col_letter, cd in ws.column_dimensions.items():
            if cd.hidden:
                hidden_cols.append(col_letter)

    # Column widths
    col_widths = {}
    if hasattr(ws, 'column_dimensions'):
        for col_letter, cd in ws.column_dimensions.items():
            if cd.width and cd.width > 0:
                col_widths[col_letter] = cd.width

    # Row heights
    row_heights = {}
    if hasattr(ws, 'row_dimensions'):
        for row_idx, rd in ws.row_dimensions.items():
            if rd.height and rd.height > 0:
                row_heights[str(row_idx)] = rd.height

    # Extract cell data for non-empty cells (with styles)
    cells_data = _extract_cells(ws, min_row, max_row, min_col, max_col)

    # Detect non-empty regions
    non_empty_rows = set()
    non_empty_cols = set()
    for cell_info in cells_data:
        non_empty_rows.add(cell_info["row"])
        non_empty_cols.add(cell_info["col"])

    sheet_atlas = {
        "sheet_name": sheet_name,
        "sheet_index": list(ws.parent.sheetnames).index(sheet_name),
        "used_range": used_range,
        "dimensions": {
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
            "total_rows": max_row - min_row + 1,
            "total_cols": max_col - min_col + 1,
        },
        "merged_cells": merged_ranges,
        "hidden_rows": hidden_rows,
        "hidden_cols": hidden_cols,
        "col_widths": col_widths,
        "row_heights": row_heights,
        "non_empty_cell_count": len(cells_data),
        "non_empty_row_count": len(non_empty_rows),
        "non_empty_col_count": len(non_empty_cols),
        "cells": cells_data,
    }

    return sheet_atlas


def _extract_cells(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> list:
    """Extract all non-empty cell data with styles."""
    cells = []

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue

            cell_info = {
                "row": cell.row,
                "col": cell.column,
                "col_letter": get_column_letter(cell.column),
                "coordinate": cell.coordinate,
                "value": _safe_value(cell.value),
                "data_type": cell.data_type,
            }

            # Formula
            if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell_info["has_formula"] = True

            # Style info (border, fill, font)
            if cell.font:
                font_info = {}
                if cell.font.bold:
                    font_info["bold"] = True
                if cell.font.color and cell.font.color.rgb and cell.font.color.rgb != "00000000":
                    font_info["color"] = str(cell.font.color.rgb)
                if cell.font.size:
                    font_info["size"] = cell.font.size
                if font_info:
                    cell_info["font"] = font_info

            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = str(cell.fill.fgColor.rgb)
                if rgb != "00000000":
                    cell_info["fill_color"] = rgb

            if cell.border:
                borders = _get_border_info(cell.border)
                if borders:
                    cell_info["borders"] = borders

            # Comment
            if cell.comment:
                cell_info["comment"] = str(cell.comment.text)[:200]

            cells.append(cell_info)

    return cells


def _safe_value(value) -> Any:
    """Convert cell value to JSON-safe type."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)[:500]  # Truncate long strings


def _get_border_info(border) -> dict:
    """Extract border information."""
    borders = {}
    for side in ["left", "right", "top", "bottom"]:
        b = getattr(border, side, None)
        if b and b.style:
            borders[side] = b.style
    return borders


def save_workbook_atlas(atlas: dict, output_dir: Path) -> Path:
    """Save workbook atlas to file."""
    atlas_dir = output_dir / "atlas"
    atlas_dir.mkdir(parents=True, exist_ok=True)
    output_path = atlas_dir / "workbook_atlas.json"
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(atlas, f, ensure_ascii=False, indent=2, default=str)
    
    return output_path


def save_sheet_atlases(atlas: dict, output_dir: Path) -> Path:
    """Save individual sheet atlases as JSONL."""
    atlas_dir = output_dir / "atlas"
    atlas_dir.mkdir(parents=True, exist_ok=True)
    output_path = atlas_dir / "sheet_atlas.jsonl"
    
    with open(output_path, "w", encoding="utf-8") as f:
        for sheet in atlas["sheets"]:
            # Write a compact version without full cell data
            compact = {k: v for k, v in sheet.items() if k != "cells"}
            compact["cell_sample"] = sheet["cells"][:20] if sheet.get("cells") else []
            f.write(json.dumps(compact, ensure_ascii=False, default=str) + "\n")
    
    return output_path
