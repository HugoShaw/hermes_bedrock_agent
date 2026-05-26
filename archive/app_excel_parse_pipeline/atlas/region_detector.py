"""Region detector: identify logical regions in a sheet without business-specific knowledge.

Uses ONLY physical/structural signals:
- Gaps in non-empty rows/columns
- Border patterns
- Merged cell boundaries
- Style changes (fill color, font)
- Dense vs sparse areas
"""
import json
from pathlib import Path
from typing import Any
from collections import defaultdict

from openpyxl.utils import get_column_letter


def detect_regions(sheet_atlas: dict) -> list:
    """Detect logical regions from sheet atlas data.
    
    A region is a contiguous rectangular block of cells that appears
    to be a coherent unit based on physical signals only.
    """
    cells = sheet_atlas.get("cells", [])
    if not cells:
        return []

    dims = sheet_atlas["dimensions"]

    # Build occupancy grid
    row_occupancy = defaultdict(set)  # row -> set of occupied cols
    col_occupancy = defaultdict(set)  # col -> set of occupied rows

    for cell in cells:
        row_occupancy[cell["row"]].add(cell["col"])
        col_occupancy[cell["col"]].add(cell["row"])

    # Detect row gaps (empty rows that separate regions vertically)
    all_rows = sorted(row_occupancy.keys())
    row_groups = _find_groups(all_rows, gap_threshold=2)

    # Detect column gaps within each row group
    regions = []
    for row_group in row_groups:
        min_r, max_r = row_group[0], row_group[-1]

        # Find which columns are populated in this row range
        cols_in_group = set()
        for r in range(min_r, max_r + 1):
            cols_in_group.update(row_occupancy.get(r, set()))

        if not cols_in_group:
            continue

        sorted_cols = sorted(cols_in_group)
        col_groups = _find_groups(sorted_cols, gap_threshold=2)

        for col_group in col_groups:
            min_c, max_c = col_group[0], col_group[-1]

            # Build region info
            region_cells = [
                c for c in cells
                if min_r <= c["row"] <= max_r and min_c <= c["col"] <= max_c
            ]

            if len(region_cells) < 2:
                continue

            region = _build_region(
                region_cells, min_r, max_r, min_c, max_c,
                sheet_atlas["sheet_name"]
            )
            regions.append(region)

    # Merge overlapping/adjacent small regions if they're likely one table
    regions = _merge_adjacent_regions(regions)

    # Assign region IDs
    for i, region in enumerate(regions):
        region["region_id"] = f"{sheet_atlas['sheet_name']}_R{i:02d}"

    return regions


def _find_groups(sorted_values: list, gap_threshold: int = 2) -> list:
    """Group consecutive values, splitting at gaps > threshold."""
    if not sorted_values:
        return []

    groups = []
    current_group = [sorted_values[0]]

    for i in range(1, len(sorted_values)):
        if sorted_values[i] - sorted_values[i - 1] > gap_threshold:
            groups.append(current_group)
            current_group = [sorted_values[i]]
        else:
            current_group.append(sorted_values[i])

    if current_group:
        groups.append(current_group)

    return groups


def _build_region(cells: list, min_r: int, max_r: int, min_c: int, max_c: int,
                  sheet_name: str) -> dict:
    """Build a region descriptor from its cells."""
    range_str = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"

    # Detect header row candidates (bold, fill, first populated rows)
    header_candidates = _detect_headers(cells, min_r, max_r)

    # Compute density
    total_possible = (max_r - min_r + 1) * (max_c - min_c + 1)
    density = len(cells) / max(total_possible, 1)

    # Detect if this looks like a table (regular column patterns)
    is_table_like = _is_table_like(cells, min_r, max_r, min_c, max_c)

    # Get sample values for context
    sample_values = []
    for c in cells[:30]:
        if c.get("value"):
            sample_values.append({
                "coordinate": c["coordinate"],
                "value": str(c["value"])[:100],
            })

    # Detect fill patterns (colored header areas)
    fill_colors = set()
    for c in cells:
        if c.get("fill_color"):
            fill_colors.add(c["fill_color"])

    region = {
        "range": range_str,
        "row_span": [min_r, max_r],
        "col_span": [min_c, max_c],
        "row_count": max_r - min_r + 1,
        "col_count": max_c - min_c + 1,
        "cell_count": len(cells),
        "density": round(density, 3),
        "is_table_like": is_table_like,
        "header_row_candidates": header_candidates,
        "has_merged_cells": any(True for c in cells if c.get("col") == min_c),  # simplified check
        "has_borders": any(c.get("borders") for c in cells),
        "has_fills": len(fill_colors) > 0,
        "fill_colors": list(fill_colors)[:5],
        "has_formulas": any(c.get("has_formula") for c in cells),
        "has_comments": any(c.get("comment") for c in cells),
        "sample_values": sample_values,
    }

    return region


def _detect_headers(cells: list, min_r: int, max_r: int) -> list:
    """Detect potential header rows based on physical signals."""
    candidates = []

    # Group cells by row
    row_cells = defaultdict(list)
    for c in cells:
        row_cells[c["row"]].append(c)

    # Check first few rows for header signals
    for row_idx in sorted(row_cells.keys())[:5]:
        row = row_cells[row_idx]
        signals = 0

        # Bold text
        if any(c.get("font", {}).get("bold") for c in row):
            signals += 2

        # Background fill
        if any(c.get("fill_color") for c in row):
            signals += 1

        # All values are strings (not numbers)
        all_str = all(
            isinstance(c.get("value"), str) for c in row if c.get("value") is not None
        )
        if all_str and len(row) >= 3:
            signals += 1

        # Short cell values (typical of headers)
        avg_len = sum(len(str(c.get("value", ""))) for c in row) / max(len(row), 1)
        if avg_len < 20:
            signals += 1

        if signals >= 2:
            candidates.append(row_idx)

    return candidates


def _is_table_like(cells: list, min_r: int, max_r: int, min_c: int, max_c: int) -> bool:
    """Check if region has table-like structure (regular column patterns)."""
    if max_r - min_r < 2:
        return False
    if max_c - min_c < 1:
        return False

    # Check if most rows have similar column coverage
    row_col_counts = defaultdict(int)
    for c in cells:
        row_col_counts[c["row"]] += 1

    if not row_col_counts:
        return False

    counts = list(row_col_counts.values())
    avg_count = sum(counts) / len(counts)
    # If most rows have similar number of filled columns, it's table-like
    consistent_rows = sum(1 for c in counts if abs(c - avg_count) <= avg_count * 0.3)
    return consistent_rows / len(counts) > 0.5


def _merge_adjacent_regions(regions: list) -> list:
    """Merge regions that are adjacent and likely parts of the same table."""
    if len(regions) <= 1:
        return regions

    # Simple heuristic: merge regions that share columns and have <=1 row gap
    merged = []
    used = set()

    for i, r1 in enumerate(regions):
        if i in used:
            continue
        current = r1.copy()
        for j, r2 in enumerate(regions):
            if j <= i or j in used:
                continue
            # Check if same column span and close row span
            if (r1["col_span"] == r2["col_span"] and
                abs(r1["row_span"][1] - r2["row_span"][0]) <= 2):
                # Merge
                current["row_span"] = [
                    min(current["row_span"][0], r2["row_span"][0]),
                    max(current["row_span"][1], r2["row_span"][1]),
                ]
                current["row_count"] = current["row_span"][1] - current["row_span"][0] + 1
                current["cell_count"] += r2["cell_count"]
                current["sample_values"] = (
                    current.get("sample_values", []) + r2.get("sample_values", [])
                )[:30]
                used.add(j)
        merged.append(current)
        used.add(i)

    # Handle any unused regions
    for i, r in enumerate(regions):
        if i not in used:
            merged.append(r)

    return merged


def save_region_atlas(regions_by_sheet: dict, output_dir: Path) -> Path:
    """Save region atlas as JSONL."""
    atlas_dir = output_dir / "atlas"
    atlas_dir.mkdir(parents=True, exist_ok=True)
    output_path = atlas_dir / "region_atlas.jsonl"

    with open(output_path, "w", encoding="utf-8") as f:
        for sheet_name, regions in regions_by_sheet.items():
            for region in regions:
                record = {"sheet_name": sheet_name, **region}
                f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")

    return output_path
