"""Generic parse plan executor.

Executes a parse plan deterministically:
- Reads the parse plan
- Extracts tables by coordinates specified in the plan
- Extracts fields by region and column roles
- Resolves references between source and target fields
- Marks uncertain/unresolved records
"""
import json
import logging
from pathlib import Path
from typing import Any, Optional
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from ..config import PipelineConfig

logger = logging.getLogger(__name__)


class ParsePlanExecutor:
    """Executes a parse plan against an Excel workbook."""

    def __init__(self, workbook_path: str, parse_plan: dict, config: PipelineConfig):
        self.workbook_path = workbook_path
        self.plan = parse_plan
        self.config = config
        self.wb = None
        
        # Results
        self.tables = []
        self.fields = []
        self.mappings = []
        self.transformations = []
        self.conditions = []
        self.uncertain_records = []
        self.unresolved_references = []

    def execute(self) -> dict:
        """Execute the parse plan and return all extracted data."""
        # Use data_only=True to get calculated values, but NOT read_only 
        # because we need random cell access by coordinate
        self.wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=False)

        try:
            for sheet_plan in self.plan.get("sheets", []):
                self._execute_sheet(sheet_plan)
                # Safety: limit total extracted fields
                if len(self.fields) > 10000:
                    logger.warning("Field extraction capped at 10000 records")
                    break
        finally:
            self.wb.close()

        # Resolve cross-region references
        self._resolve_references()

        return {
            "tables": self.tables,
            "fields": self.fields,
            "mappings": self.mappings,
            "transformations": self.transformations,
            "conditions": self.conditions,
            "uncertain_records": self.uncertain_records,
            "unresolved_references": self.unresolved_references,
        }

    def _execute_sheet(self, sheet_plan: dict):
        """Execute plan for a single sheet."""
        sheet_name = sheet_plan["sheet_name"]
        
        if sheet_name not in self.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return

        ws = self.wb[sheet_name]

        for region in sheet_plan.get("regions", []):
            self._execute_region(ws, sheet_name, region)

    def _execute_region(self, ws, sheet_name: str, region_plan: dict):
        """Execute plan for a single region."""
        role = region_plan.get("semantic_role", "unknown")
        strategy = region_plan.get("extraction_strategy", {})
        strategy_type = strategy.get("type", "table_rows")

        if strategy_type == "skip":
            return

        if strategy_type == "table_rows":
            self._extract_table_rows(ws, sheet_name, region_plan)
        elif strategy_type == "mapping_rows":
            self._extract_mapping_rows(ws, sheet_name, region_plan)
        elif strategy_type == "note_text":
            self._extract_note_text(ws, sheet_name, region_plan)
        else:
            # Default to table extraction
            self._extract_table_rows(ws, sheet_name, region_plan)

    def _extract_table_rows(self, ws, sheet_name: str, region_plan: dict):
        """Extract table rows based on the region plan."""
        region_id = region_plan.get("region_id", "")
        range_str = region_plan.get("range", "")
        
        # Parse range
        row_start, row_end, col_start, col_end = self._parse_range(range_str, region_plan)
        
        # Get header rows
        header_rows = region_plan.get("header_rows", [])
        data_start = region_plan.get("data_start_row", row_start + 1)
        data_end = region_plan.get("data_end_row", row_end)

        # Get column roles
        columns = region_plan.get("columns", [])
        col_roles = {}
        for col_def in columns:
            col_letter = col_def.get("column", "")
            if col_letter:
                try:
                    col_idx = column_index_from_string(col_letter)
                    col_roles[col_idx] = col_def.get("role", "unknown")
                except Exception:
                    pass

        # Extract headers
        headers = {}
        for h_row in header_rows:
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=h_row, column=col)
                if cell.value is not None:
                    headers[col] = str(cell.value).strip()

        # Record the table
        table_record = {
            "table_id": f"{sheet_name}_{region_id}",
            "sheet_name": sheet_name,
            "region_id": region_id,
            "semantic_role": region_plan.get("semantic_role", "unknown"),
            "range": range_str,
            "header_rows": header_rows,
            "headers": {get_column_letter(k): v for k, v in headers.items()},
            "data_row_range": [data_start, data_end],
            "row_count": max(0, data_end - data_start + 1),
            "evidence": {
                "workbook": self.plan.get("workbook_name", ""),
                "sheet": sheet_name,
                "region_id": region_id,
                "cell_range": range_str,
            },
        }
        self.tables.append(table_record)

        # Extract fields (rows) - cap at 500 rows per region for performance
        max_data_rows = 500
        actual_end = min(data_end + 1, data_start + max_data_rows)
        for row_idx in range(data_start, actual_end):
            row_data = {}
            row_has_value = False

            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row_idx, column=col)
                if cell.value is not None:
                    row_has_value = True
                    col_letter = get_column_letter(col)
                    role = col_roles.get(col, "unknown")
                    header = headers.get(col, col_letter)

                    row_data[col_letter] = {
                        "value": _safe_value(cell.value),
                        "header": header,
                        "role": role,
                    }

            if row_has_value:
                field_record = {
                    "field_id": f"{sheet_name}_{region_id}_R{row_idx}",
                    "table_id": f"{sheet_name}_{region_id}",
                    "sheet_name": sheet_name,
                    "region_id": region_id,
                    "row": row_idx,
                    "data": row_data,
                    "confidence": 0.8 if col_roles else 0.5,
                    "evidence": {
                        "workbook": self.plan.get("workbook_name", ""),
                        "sheet": sheet_name,
                        "region_id": region_id,
                        "cell_range": f"{get_column_letter(col_start)}{row_idx}:{get_column_letter(col_end)}{row_idx}",
                        "row": row_idx,
                    },
                }
                self.fields.append(field_record)

    def _extract_mapping_rows(self, ws, sheet_name: str, region_plan: dict):
        """Extract mapping rows (source->target with transformation)."""
        # Similar to table rows but also creates mapping records
        self._extract_table_rows(ws, sheet_name, region_plan)

        # Post-process to create mapping records
        region_id = region_plan.get("region_id", "")
        columns = region_plan.get("columns", [])

        source_col = None
        target_col = None
        transform_col = None

        for col_def in columns:
            role = col_def.get("role", "")
            if role == "source_reference":
                source_col = col_def.get("column", "")
            elif role == "target_reference":
                target_col = col_def.get("column", "")
            elif role == "transformation_rule":
                transform_col = col_def.get("column", "")

        if source_col or target_col:
            for field in self.fields:
                if field.get("region_id") == region_id:
                    data = field.get("data", {})
                    source_val = data.get(source_col, {}).get("value", "") if source_col else ""
                    target_val = data.get(target_col, {}).get("value", "") if target_col else ""
                    transform_val = data.get(transform_col, {}).get("value", "") if transform_col else ""

                    if source_val or target_val:
                        mapping = {
                            "mapping_id": f"map_{field['field_id']}",
                            "source_field": str(source_val),
                            "target_field": str(target_val),
                            "transformation": str(transform_val) if transform_val else None,
                            "sheet_name": sheet_name,
                            "region_id": region_id,
                            "row": field.get("row"),
                            "confidence": 0.7,
                            "evidence": field.get("evidence", {}),
                        }
                        self.mappings.append(mapping)

    def _extract_note_text(self, ws, sheet_name: str, region_plan: dict):
        """Extract free-text notes from a region."""
        range_str = region_plan.get("range", "")
        row_start, row_end, col_start, col_end = self._parse_range(range_str, region_plan)

        text_parts = []
        for row_idx in range(row_start, min(row_end + 1, row_start + 100)):
            row_text = []
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row_idx, column=col)
                if cell.value is not None:
                    row_text.append(str(cell.value))
            if row_text:
                text_parts.append(" | ".join(row_text))

        if text_parts:
            self.uncertain_records.append({
                "record_type": "note_text",
                "sheet_name": sheet_name,
                "region_id": region_plan.get("region_id", ""),
                "text": "\n".join(text_parts),
                "confidence": 0.5,
                "reason": "note_block_needs_interpretation",
            })

    def _parse_range(self, range_str: str, region_plan: dict) -> tuple:
        """Parse range string or use region plan coordinates."""
        try:
            if ":" in range_str:
                parts = range_str.split(":")
                # Parse start
                start_col_str = ""
                start_row_str = ""
                for ch in parts[0]:
                    if ch.isalpha():
                        start_col_str += ch
                    else:
                        start_row_str += ch
                # Parse end
                end_col_str = ""
                end_row_str = ""
                for ch in parts[1]:
                    if ch.isalpha():
                        end_col_str += ch
                    else:
                        end_row_str += ch

                col_start = column_index_from_string(start_col_str) if start_col_str else 1
                row_start = int(start_row_str) if start_row_str else 1
                col_end = column_index_from_string(end_col_str) if end_col_str else col_start
                row_end = int(end_row_str) if end_row_str else row_start

                return row_start, row_end, col_start, col_end
        except Exception:
            pass

        # Fallback to row_span/col_span
        row_span = region_plan.get("row_span", [1, 100])
        col_span = region_plan.get("col_span", [1, 20])
        return row_span[0], row_span[1], col_span[0], col_span[1]

    def _resolve_references(self):
        """Resolve cross-references between source and target fields."""
        # Build a lookup by field name
        field_by_name = defaultdict(list)
        for field in self.fields:
            for col_data in field.get("data", {}).values():
                if col_data.get("role") == "field_name" and col_data.get("value"):
                    field_by_name[str(col_data["value"]).strip()].append(field)

        # Check mappings for unresolved references
        for mapping in self.mappings:
            source = mapping.get("source_field", "")
            target = mapping.get("target_field", "")

            if source and source not in field_by_name:
                self.unresolved_references.append({
                    "type": "missing_source",
                    "mapping_id": mapping["mapping_id"],
                    "reference": source,
                    "confidence": 0.4,
                })

            if target and target not in field_by_name:
                self.unresolved_references.append({
                    "type": "missing_target",
                    "mapping_id": mapping["mapping_id"],
                    "reference": target,
                    "confidence": 0.4,
                })


def save_execution_results(results: dict, output_dir: Path) -> dict:
    """Save all execution results to structured JSONL files."""
    structured_dir = output_dir / "structured"
    structured_dir.mkdir(parents=True, exist_ok=True)

    saved_files = {}

    for key in ["tables", "fields", "mappings", "transformations", 
                "conditions", "uncertain_records", "unresolved_references"]:
        records = results.get(key, [])
        output_path = structured_dir / f"{key}.jsonl"
        with open(output_path, "w", encoding="utf-8") as f:
            for record in records:
                f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
        saved_files[key] = {"path": str(output_path), "count": len(records)}

    return saved_files


def _safe_value(value) -> Any:
    """Convert cell value to JSON-safe type."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)[:500]
