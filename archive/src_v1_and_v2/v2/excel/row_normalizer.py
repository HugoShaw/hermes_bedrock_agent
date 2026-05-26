"""
Row normalizer — convert table rows to ExcelRowRecord instances.

Requirements:
- Use detected headers for column names
- Preserve original cell refs
- Preserve raw values and normalized string values
- Skip fully empty rows
- Preserve formulas separately
- Include workbook/sheet/table region metadata
"""
from __future__ import annotations

import logging
from typing import Any

from openpyxl.utils import get_column_letter

from hermes_bedrock_agent.v2.excel.excel_schema import ExcelRowRecord, ExcelTableRegion

logger = logging.getLogger(__name__)


class RowNormalizer:
    """Normalize Excel rows into ExcelRowRecord instances.

    Parameters
    ----------
    max_cell_text_length : int
        Maximum text length per cell value.
    max_rows_per_region : int
        Maximum rows to process per table region.
    """

    def __init__(
        self,
        max_cell_text_length: int = 500,
        max_rows_per_region: int = 2000,
    ) -> None:
        self.max_cell_text_length = max_cell_text_length
        self.max_rows_per_region = max_rows_per_region

    def normalize_region(
        self,
        ws: Any,
        region: ExcelTableRegion,
        workbook_id: str,
    ) -> list[ExcelRowRecord]:
        """Normalize all data rows in a table region.

        Parameters
        ----------
        ws : openpyxl Worksheet
        region : ExcelTableRegion
            Detected table region with header and data range info.
        workbook_id : str
            Parent workbook ID.

        Returns
        -------
        List of ExcelRowRecord instances.
        """
        if region.data_start_row is None or region.data_end_row is None:
            return []

        min_col = region.metadata.get("min_col", 1)
        max_col = region.metadata.get("max_col", 1)
        columns = region.columns

        records: list[ExcelRowRecord] = []
        rows_processed = 0

        for row_idx in range(region.data_start_row, region.data_end_row + 1):
            if rows_processed >= self.max_rows_per_region:
                logger.warning(
                    "Hit max_rows_per_region (%d) for region %s",
                    self.max_rows_per_region, region.cell_range,
                )
                break

            values: dict[str, Any] = {}
            normalized_values: dict[str, str] = {}
            source_cell_refs: dict[str, str] = {}
            is_empty = True

            for col_offset, col_idx in enumerate(range(min_col, max_col + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_name = columns[col_offset] if col_offset < len(columns) else get_column_letter(col_idx)
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

                if cell.value is not None:
                    is_empty = False
                    raw_value = cell.value
                    values[col_name] = raw_value

                    # Normalize to string
                    if isinstance(raw_value, str):
                        norm = raw_value.strip()[:self.max_cell_text_length]
                    else:
                        norm = str(raw_value)[:self.max_cell_text_length]
                    normalized_values[col_name] = norm
                    source_cell_refs[col_name] = cell_ref

            # Skip fully empty rows
            if is_empty:
                continue

            row_id = ExcelRowRecord.generate_id(
                region.sheet_id,
                row_idx,
                region.table_region_id,
            )

            record = ExcelRowRecord(
                row_id=row_id,
                workbook_id=workbook_id,
                sheet_id=region.sheet_id,
                table_region_id=region.table_region_id,
                sheet_name=region.sheet_name,
                row_number=row_idx,
                values=values,
                normalized_values=normalized_values,
                source_cell_refs=source_cell_refs,
                metadata={
                    "cell_range": region.cell_range,
                    "region_type": region.region_type,
                },
            )
            records.append(record)
            rows_processed += 1

        logger.info(
            "Normalized %d rows from region %s in sheet '%s'",
            len(records), region.cell_range, region.sheet_name,
        )
        return records
